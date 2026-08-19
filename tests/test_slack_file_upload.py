"""Slack file-upload lane (cq-b0a847ef0c8e) + xlsx export (cq-c51123b0ad07).

The upload lane had NEVER executed: `files:write` was ungranted, the handler
logged the refusal at WARNING and returned False, and the caller silently posted
inline. Every failure mode collapsed to one bool, so "permanently broken" and
"never reached" were indistinguishable -- from the logs AND from the requester.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from cora.tools import slack_file_upload as sfu
from cora.tools import table_export as tx

_REPO_ROOT = Path(__file__).resolve().parents[1]


class _Resp(dict):
    def __init__(self, data, headers=None):
        super().__init__(data)
        self.headers = headers or {}


class FakeSlack:
    """Minimal Slack stub. `scopes=None` models a probe that cannot conclude."""

    def __init__(self, scopes="chat:write,files:write", upload_error=None,
                 complete_error=None):
        self._scopes = scopes
        self.upload_error = upload_error
        self.complete_error = complete_error
        self.completed = []

    def auth_test(self):
        headers = {"x-oauth-scopes": [self._scopes]} if self._scopes is not None else {}
        return _Resp({"ok": True, "user_id": "U0BOT"}, headers)

    def files_getUploadURLExternal(self, filename, length):
        if self.upload_error:
            return _Resp({"ok": False, "error": self.upload_error})
        return _Resp({"ok": True, "upload_url": "https://files.slack.test/up",
                      "file_id": "F1"})

    def files_completeUploadExternal(self, **kwargs):
        if self.complete_error:
            return _Resp({"ok": False, "error": self.complete_error})
        self.completed.append(kwargs)
        return _Resp({"ok": True})


@pytest.fixture(autouse=True)
def _clean(monkeypatch):
    sfu.reset_scope_cache()
    monkeypatch.setenv("SLACK_BOT_TOKEN", "xoxb-test")
    yield
    sfu.reset_scope_cache()


@pytest.fixture
def put_ok(monkeypatch):
    captured = {}

    def _put(url, content=None, headers=None, timeout=None):
        captured["body"] = content
        captured["headers"] = headers

        class R:
            status_code = 200
        return R()

    import httpx
    monkeypatch.setattr(httpx, "put", _put)
    return captured


class TestScopeProbe:
    def test_probe_reads_the_live_grant_not_the_manifest(self):
        """The manifest is repo state and can lag the live app in EITHER
        direction; the header is what the token actually carries."""
        assert sfu.files_write_granted(FakeSlack("chat:write,files:write")) is True
        sfu.reset_scope_cache()
        assert sfu.files_write_granted(FakeSlack("chat:write")) is False

    def test_inconclusive_probe_is_unknown_not_denied(self):
        """A transient auth.test blip must not become a silent capability
        outage -- unknown means 'attempt it and find out'."""
        sfu.reset_scope_cache()
        assert sfu.files_write_granted(FakeSlack(scopes=None)) is None

    def test_unrecognizable_header_is_unknown_not_denied(self):
        """REGRESSION. Any object that merely responds to .get() -- a test
        double, a changed SDK response shape -- used to yield a confident
        "not granted" and silently disable the lane. That is the exact class of
        invisible degrade this module exists to end, so a header that does not
        parse as a scope list must read as UNKNOWN.
        """
        sfu.reset_scope_cache()
        assert sfu.files_write_granted(FakeSlack("<MagicMock id='140''>")) is None
        sfu.reset_scope_cache()
        assert sfu.files_write_granted(FakeSlack("  ")) is None

    def test_probe_is_cached(self):
        client = FakeSlack("files:write")
        calls = []
        orig = client.auth_test
        client.auth_test = lambda: (calls.append(1), orig())[1]
        sfu.files_write_granted(client)
        sfu.files_write_granted(client)
        assert len(calls) == 1


class TestUploadOutcomes:
    def test_success(self, put_ok):
        client = FakeSlack()
        outcome, _ = sfu.upload_bytes(client, "C1", "f.txt", b"hello", "T")
        assert outcome == sfu.OK
        assert client.completed[0]["channel_id"] == "C1"

    def test_missing_scope_is_reported_not_swallowed(self):
        """The whole point: the caller can now distinguish this from a generic
        failure and tell the requester an admin grant is what's missing."""
        outcome, _ = sfu.upload_bytes(FakeSlack(scopes="chat:write"), "C1",
                                      "f.txt", b"x", "T")
        assert outcome == sfu.NO_SCOPE
        note = sfu.requester_note(outcome)
        assert "files:write" in note
        # D-051: adding a bot scope requires REINSTALLING the app, which issues a
        # new bot token. "Harrison can grant it in the Slack app settings" reads
        # as a toggle and would send him looking for one that does not exist.
        assert "reinstalling" in note and "new bot token" in note

    def test_missing_scope_detected_at_the_api_when_the_probe_was_unknown(self, put_ok):
        outcome, _ = sfu.upload_bytes(
            FakeSlack(scopes=None, upload_error="missing_scope"), "C1",
            "f.txt", b"x", "T")
        assert outcome == sfu.NO_SCOPE

    def test_not_in_channel_gets_its_own_actionable_note(self, put_ok):
        outcome, _ = sfu.upload_bytes(
            FakeSlack(complete_error="not_in_channel"), "C1", "f.txt", b"x", "T")
        assert outcome == sfu.NOT_IN_CHANNEL
        assert "invite me" in sfu.requester_note(outcome).lower()

    def test_no_token(self, monkeypatch):
        monkeypatch.delenv("SLACK_BOT_TOKEN", raising=False)
        outcome, _ = sfu.upload_bytes(FakeSlack(), "C1", "f.txt", b"x", "T")
        assert outcome == sfu.NO_TOKEN

    def test_every_outcome_has_a_requester_note(self):
        for outcome in (sfu.NO_SCOPE, sfu.NO_TOKEN, sfu.NO_HTTPX,
                        sfu.NOT_IN_CHANNEL, sfu.FAILED):
            assert sfu.requester_note(outcome).strip()
        assert sfu.requester_note(sfu.OK) == ""

    def test_text_upload_sanitizes_content_and_title(self, put_ok):
        """W3-05: these bytes are PUT straight to Slack, bypassing the egress
        WebClient patch, so the sanitizer has to run here."""
        client = FakeSlack()
        sfu.upload_text(client, "C1", "T https://drive.google.com/file/d/abc/view",
                        "body https://drive.google.com/file/d/abc/view")
        assert b"drive.google.com" not in put_ok["body"]
        assert "drive.google.com" not in client.completed[0]["files"][0]["title"]


class TestManifest:
    def test_manifest_requests_files_write(self):
        cfg = json.loads((_REPO_ROOT / "slack-app-config" / "manifest.json")
                         .read_text(encoding="utf-8"))
        assert "files:write" in cfg["oauth_config"]["scopes"]["bot"]


class TestLegacyCallerContract:
    def test_upload_report_as_file_returns_an_outcome_pair(self, put_ok):
        """The bool was the defect. A caller that cannot tell WHY cannot say
        why, and a silent degrade reads as dead code."""
        from cora.tools import financial_client
        result = financial_client.upload_report_as_file(
            FakeSlack(), "C1", "Cash flow", "x" * 50)
        assert isinstance(result, tuple) and result[0] == sfu.OK

    def test_cashflow_caller_falls_back_visibly(self):
        src = (_REPO_ROOT / "src" / "cora" / "tools" / "tool_dispatch.py").read_text(
            encoding="utf-8")
        block = src[src.index("Feature 6: upload long reports as Slack files"):]
        block = block[:block.index("def _tool_osn_financial_pulse")]
        assert "requester_note(outcome)" in block
        assert "if uploaded:" not in block


class TestXlsxExport:
    def test_rows_from_report_reads_the_rendered_lines(self):
        text = ("Profit and Loss for F3E (2026-01-01 to 2026-03-31) [Accrual basis]:\n"
                "  • Income: 125,000.00\n"
                "  • Total Expenses: 90,000.00\n")
        rows = tx.rows_from_report(text)
        assert rows[0] == ["Section", "Item", "Value"]
        assert ["", "Income", "125,000.00"] in [
            [r[0], r[1], r[2]] for r in rows] or any(
            r[1] == "Income" and r[2] == "125,000.00" for r in rows)

    def test_caveat_lines_are_kept_not_dropped(self):
        """An export that silently omits the qualifier next to a partial figure
        is how a partial number gets read as a complete one."""
        text = "  • Revenue: 100\nNo data this week for: Gilbert. 3 of 4 stores.\n"
        flat = " ".join(c for r in tx.rows_from_report(text) for c in r)
        assert "3 of 4 stores" in flat

    def test_values_are_never_recomputed(self):
        """D-095: the sheet is a TRANSCRIPTION of what was displayed. Two
        numbers for one question on a money surface is worse than no export."""
        rows = tx.rows_from_report("  • Income: (1,234.56)\n")
        assert any("(1,234.56)" in c for r in rows for c in r)

    def test_sheet_name_is_sanitized_not_failed(self):
        """cq-ad74f3908e8d is this bug's sibling, where an over-long tab name
        silently fell back to markdown."""
        name = tx.safe_sheet_name("P&L [F3E]: 2026-01-01/2026-03-31 extra long title")
        assert len(name) <= 31
        for ch in "[]:*?/\\":
            assert ch not in name

    def test_build_xlsx_round_trips(self, tmp_path):
        from openpyxl import load_workbook
        payload = tx.build_xlsx([["A", "B"], ["1", "2"]], "Sheet")
        f = tmp_path / "o.xlsx"
        f.write_bytes(payload)
        ws = load_workbook(f).active
        assert [c.value for c in ws[1]] == ["A", "B"]
        assert [c.value for c in ws[2]] == ["1", "2"]

    def test_export_sanitizes_before_building_the_workbook(self, put_ok):
        """Binary payloads cannot be sanitized after the fact -- an xlsx would
        otherwise be a way to ship exactly what sanitize_text strips."""
        client = FakeSlack()
        outcome, _ = tx.deliver_report_as_xlsx(
            client, "C1", "Report", "  • Link: https://drive.google.com/file/d/abc/view")
        assert outcome == sfu.OK
        assert b"drive.google.com" not in put_ok["body"]

    def test_uploads_with_the_xlsx_content_type(self, put_ok):
        tx.deliver_report_as_xlsx(FakeSlack(), "C1", "Report", "  • A: 1")
        assert "spreadsheetml.sheet" in put_ok["headers"]["Content-Type"]


class TestExportWiring:
    def test_format_param_offered_on_every_qbo_report_tool(self):
        from cora.tools import tool_dispatch as td
        for name in ("qbo_get_profit_loss", "qbo_get_balance_sheet",
                     "qbo_get_ar_aging", "qbo_get_ap_aging",
                     "qbo_get_recent_transactions"):
            spec = next(t for t in td.TOOL_DEFINITIONS if t["name"] == name)
            fmt = spec["input_schema"]["properties"]["format"]
            assert fmt["enum"] == ["text", "xlsx"], name

    def test_export_is_opt_in(self):
        from cora.tools import tool_dispatch as td
        text = "  • Income: 1"
        assert td._maybe_export_xlsx({"_channel_id": "C1"}, "T", text) == text

    def test_export_skips_refusals(self, monkeypatch):
        """An .xlsx of an error message is noise that looks like data."""
        from cora.tools import tool_dispatch as td
        called = []
        monkeypatch.setattr(tx, "deliver_report_as_xlsx",
                            lambda *a, **k: called.append(1) or (sfu.OK, ""))
        out = td._maybe_export_xlsx(
            {"_channel_id": "C1", "format": "xlsx"}, "T",
            "I couldn't reach the finance source right now.")
        assert not called and "couldn't reach" in out

    def test_export_failure_never_breaks_the_answer(self, monkeypatch):
        from cora.tools import tool_dispatch as td

        def _boom(*a, **k):
            raise RuntimeError("nope")

        monkeypatch.setattr(tx, "deliver_report_as_xlsx", _boom)
        text = "  • Income: 1"
        assert text in td._maybe_export_xlsx(
            {"_channel_id": "C1", "format": "xlsx"}, "T", text)

    def test_successful_export_still_returns_the_inline_numbers(self, monkeypatch):
        """The sheet is for reuse, the message is for reading -- replacing one
        with the other makes a quick 'what's our AR look like' worse."""
        from cora.tools import tool_dispatch as td
        monkeypatch.setattr(tx, "deliver_report_as_xlsx", lambda *a, **k: (sfu.OK, ""))
        out = td._maybe_export_xlsx({"_channel_id": "C1", "format": "xlsx"},
                                    "T", "  • Income: 125,000.00")
        assert "125,000.00" in out and "xlsx" in out

    def test_export_failure_is_disclosed(self, monkeypatch):
        from cora.tools import tool_dispatch as td
        monkeypatch.setattr(tx, "deliver_report_as_xlsx",
                            lambda *a, **k: (sfu.NO_SCOPE, ""))
        out = td._maybe_export_xlsx({"_channel_id": "C1", "format": "xlsx"},
                                    "T", "  • Income: 1")
        assert "files:write" in out

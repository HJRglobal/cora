"""QBO monthly-report folder populator (cq-96adf03bcda3).

Kills Justin's manual QBO export loop (decision F5, 2026-08-18). The risks worth
pinning are all about writing the RIGHT file to the RIGHT place, because a
misfiled statement is worse than a missing one:

* the folder is the FILING month, the filename carries the REPORT month one month
  earlier -- invert it and a year of statements misfile silently;
* realm -> slug is a money-adjacent mapping (LEX could plausibly have been `llc`,
  `lexcorp` or `lts`), so it is asserted against the live QBO company name before
  any write rather than trusted;
* an existing manual upload is never overwritten;
* the basis is pinned AND verified from the response -- sending a parameter is
  not evidence the API honored it;
* the Slack-facing summary carries no financial figures at all.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
import yaml

_REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO_ROOT / "src"))

from cora import qbo_monthly_reports as qmr  # noqa: E402

_MAP_PATH = _REPO_ROOT / "data" / "maps" / "qbo-monthly-report-slugs.yaml"


def _pl(rows=None, basis="Accrual", no_data=False):
    return {
        "Header": {
            "ReportName": "ProfitAndLoss",
            "ReportBasis": basis,
            "Time": "2026-08-18T17:00:00-07:00",
            # Real responses echo the window back; the fixture must too, or the
            # period verification silently no-ops in every test that uses it.
            "StartPeriod": "2026-07-01",
            "EndPeriod": "2026-07-31",
            "Option": [{"Name": "NoReportData", "Value": "true" if no_data else "false"}],
        },
        "Rows": rows if rows is not None else {
            "Row": [{
                "Header": {"ColData": [{"value": "Income"}, {"value": ""}]},
                "Rows": {"Row": [
                    {"ColData": [{"value": "4275 Management Fees"},
                                 {"value": "79580.19"}]},
                ]},
                "Summary": {"ColData": [{"value": "Total Income"},
                                        {"value": "79580.19"}]},
                "type": "Section",
            }],
        },
    }


def _sources(**over):
    kw = dict(
        provisioned=lambda: ["HJRG"],
        company_name=lambda e: "HJR Global Services LLC",
        profit_loss=lambda e, s, en: _pl(),
        balance_sheet=lambda e, a: _pl(),
    )
    kw.update(over)
    return qmr.Sources(**kw)


def _fake_fs():
    """In-memory Drive: (writer, exists, reader, store)."""
    store: dict[str, bytes] = {}
    return (lambda p, b: store.__setitem__(str(p), b),
            lambda p: str(p) in store,
            lambda p: store[str(p)],
            store)


def _manual_xlsx(total="2000.0", label="Total for Income"):
    """A manual QBO UI export: same layout, numeric cells, and crucially NO
    "pulled by Cora" footer -- that stamp is what makes a re-run idempotent.
    Uses the UI wording "Total for X" and a bare float, both of which the real
    2026-05 archive files use.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["HJR Global Services LLC", ""])
    ws.append(["Profit and Loss", ""])
    ws.append(["July 2026", ""])
    ws.append(["", ""])
    ws.append(["", "Total"])
    ws.append(["4275 Management Fees", float(total)])
    ws.append([label, float(total)])
    ws.append(["", ""])
    ws.append(["Accrual Basis Friday, May 22, 2026 10:02 PM GMTZ", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestMonthConvention:
    """The one inversion that would silently misfile everything."""

    @pytest.mark.parametrize("report_month,folder", [
        ("2026-05", "2026-06"),
        ("2026-07", "2026-08"),
        ("2026-01", "2026-02"),
        ("2026-12", "2027-01"),   # year rollover
        ("2025-12", "2026-01"),
    ])
    def test_filing_folder_is_report_month_plus_one(self, report_month, folder):
        assert qmr.filing_folder_for(report_month) == folder

    def test_matches_the_live_archive_layout(self):
        """Ground truth: the real 2026-06/ folder holds 2026-05_*.xlsx."""
        assert qmr.filing_folder_for("2026-05") == "2026-06"
        assert qmr.filename("2026-05", "hjrg", "pl") == "2026-05_hjrg_pl.xlsx"

    @pytest.mark.parametrize("today,expected", [
        ((2026, 8, 2), "2026-07"),
        ((2026, 1, 2), "2025-12"),
        ((2026, 3, 31), "2026-02"),
    ])
    def test_previous_month(self, today, expected):
        import datetime
        assert qmr.previous_month(datetime.date(*today)) == expected

    @pytest.mark.parametrize("month,bounds", [
        ("2026-02", ("2026-02-01", "2026-02-28")),
        ("2024-02", ("2024-02-01", "2024-02-29")),   # leap year
        ("2026-07", ("2026-07-01", "2026-07-31")),
    ])
    def test_month_bounds(self, month, bounds):
        assert qmr.month_bounds(month) == bounds

    @pytest.mark.parametrize("bad", ["2026", "2026-13", "not-a-month", "2026-00"])
    def test_bad_month_raises(self, bad):
        with pytest.raises(ValueError):
            qmr.month_bounds(bad)

    def test_period_label_matches_qbo_style(self):
        assert qmr.period_label("2026-05") == "May 2026"


class TestSlugMap:
    def test_live_map_covers_every_provisioned_realm(self):
        """Drift guard: OAuth a realm without mapping it and this fails rather
        than the job silently skipping that entity every month."""
        from cora.connectors.qbo_oauth import list_provisioned_entities
        mapped = set(qmr.load_slug_map())
        missing = {e.upper() for e in list_provisioned_entities()} - mapped
        assert not missing, f"unmapped provisioned realms: {sorted(missing)}"

    def test_slugs_are_unique(self):
        slugs = [v["slug"] for v in qmr.load_slug_map().values()]
        assert len(slugs) == len(set(slugs)), "two realms share one slug"

    def test_lex_maps_to_llc_not_lexcorp_or_lts(self):
        """Confirmed twice on 2026-08-18: live companyinfo says 'Lexington LLC',
        and row 1 of the archive's 2026-05_llc_pl.xlsx says the same. lexcorp is
        'LexCorp, LLC' and lts is 'Lexington Therapies, LLC' -- different books."""
        lex = qmr.load_slug_map()["LEX"]
        assert lex["slug"] == "llc"
        assert lex["company_name"] == "Lexington LLC"

    def test_osn_parent_maps_to_core4(self):
        osn = qmr.load_slug_map()["OSN"]
        assert osn["slug"] == "osn-core4"
        assert osn["company_name"] == "CORE 4 OSN LLC"

    def test_every_entry_carries_an_expected_company_name(self):
        for realm, spec in qmr.load_slug_map().items():
            assert spec["company_name"], realm

    def test_unmapped_slugs_are_data_not_prose(self):
        """The 'still manual' line in the summary reads from here, so it cannot
        drift from the map."""
        unmapped = set(qmr.unmapped_slugs())
        assert "ufl" in unmapped and "lts" in unmapped and "lexcorp" in unmapped
        assert not (unmapped & {v["slug"] for v in qmr.load_slug_map().values()})

    def test_map_file_is_parseable_yaml(self):
        assert isinstance(yaml.safe_load(_MAP_PATH.read_text(encoding="utf-8")), dict)

    def test_unreadable_map_degrades_to_empty(self, tmp_path):
        assert qmr.load_slug_map(tmp_path / "nope.yaml") == {}


class TestIdentityAssertion:
    """The control that makes filing OSN's P&L under `llc` impossible."""

    def test_mismatched_company_name_refuses_the_write(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(company_name=lambda e: "Some Other Company LLC"),
            apply=True, writer=w, exists=e, reader=r)
        assert store == {}, "must not write under a mismatched identity"
        assert summary["written"] == []
        assert any("company-name mismatch" in s["reason"]
                   for s in summary["skipped"])

    def test_matching_name_proceeds(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                  writer=w, exists=e, reader=r)
        assert len(summary["written"]) == 2   # pl + bs
        assert len(store) == 2

    def test_name_match_is_case_insensitive(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(company_name=lambda e: "hjr global services llc"),
            apply=True, writer=w, exists=e, reader=r)
        assert len(summary["written"]) == 2

    def test_companyinfo_failure_skips_that_entity_only(self):
        def boom(entity):
            if entity == "HJRG":
                raise RuntimeError("token expired")
            return "Big D Media"
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(provisioned=lambda: ["HJRG", "BDM"],
                             company_name=boom),
            apply=True, writer=w, exists=e, reader=r)
        # BDM still got its two files; only HJRG was skipped.
        assert len(summary["written"]) == 2
        assert all("bdm" in w2["file"] for w2 in summary["written"])
        assert any(s["entity"] == "HJRG" for s in summary["skipped"])

    def test_unmapped_realm_is_reported_not_guessed(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07", sources=_sources(provisioned=lambda: ["NEWCO"]),
            apply=True, writer=w, exists=e, reader=r)
        assert store == {}
        assert any("no slug mapping" in s["reason"] for s in summary["skipped"])


class TestBasisPinning:
    def test_basis_mismatch_refuses(self):
        """Sending accounting_method is not proof QBO honored it."""
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(profit_loss=lambda en, s, e2: _pl(basis="Cash")),
            apply=True, writer=w, exists=e, reader=r)
        assert any("basis mismatch" in s["reason"] for s in summary["skipped"])
        assert all("_pl" not in w2["file"] for w2 in summary["written"])

    def test_no_data_period_is_skipped_not_written_empty(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(profit_loss=lambda en, s, e2: _pl(no_data=True)),
            apply=True, writer=w, exists=e, reader=r)
        assert any("no data" in s["reason"] for s in summary["skipped"])

    def test_report_basis_and_no_data_readers(self):
        assert qmr.report_basis(_pl(basis="Cash")) == "Cash"
        assert qmr.report_has_no_data(_pl(no_data=True)) is True
        assert qmr.report_has_no_data(_pl()) is False


class TestNeverOverwrite:
    def test_existing_manual_file_is_left_untouched(self):
        w, e, r, store = _fake_fs()
        outdir = qmr.target_dir("2026-07")
        manual = outdir / "2026-07_hjrg_pl.xlsx"
        store[str(manual)] = b"MANUAL"
        qmr.build_month("2026-07", sources=_sources(), apply=True,
                        writer=w, exists=e, reader=r)
        assert store[str(manual)] == b"MANUAL", "manual upload was clobbered"
        assert str(outdir / "2026-07_hjrg_pl-cora.xlsx") in store

    def test_collision_with_a_MANUAL_file_is_reported_as_parity(self):
        w, e, r, store = _fake_fs()
        outdir = qmr.target_dir("2026-07")
        # 79580.19 is what _pl() returns, so the figures genuinely agree.
        store[str(outdir / "2026-07_hjrg_pl.xlsx")] = _manual_xlsx(total="79580.19")
        summary = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                  writer=w, exists=e, reader=r)
        parity = [x for x in summary["parity"] if x["kind"] == "pl"]
        assert len(parity) == 1
        assert parity[0]["match"] is True

    def test_parity_compares_NUMERICALLY_not_as_strings(self):
        """The archive stores 2000.0; Cora renders 2000.00. A string compare
        called every collision a DIFFERS and sent a human hunting a discrepancy
        that did not exist (D-051)."""
        assert qmr._values_agree("2000.0", "2000.00") is True
        assert qmr._values_agree("92339.00000000001", "92339.0") is True
        assert qmr._values_agree("2000.0", "2500.0") is False
        assert qmr._values_agree(None, "2000.0") is None
        assert qmr._values_agree("n/a", "2000.0") is False

    def test_a_rerun_over_coras_own_file_is_a_no_op(self):
        """HIGH (D-051): without this, one realm failing mid-run and an operator
        re-running the month made Cora read its OWN earlier output as "the
        manual upload", write a -cora duplicate, and on a third run overwrite
        that -- unbounded duplication of statements of record."""
        w, e, r, store = _fake_fs()
        first = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                writer=w, exists=e, reader=r)
        assert len(first["written"]) == 2
        before = dict(store)

        second = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                 writer=w, exists=e, reader=r)
        assert second["written"] == [], "a re-run must write nothing"
        assert second["parity"] == [], "Cora own file is not a manual upload"
        assert store == before, "a re-run must not add or change any file"
        assert all("already written by Cora" in x["reason"]
                   for x in second["skipped"])

    def test_cora_written_files_carry_a_detectable_stamp(self):
        data = qmr.render_xlsx(_pl(), company_name="HJR Global Services LLC",
                               report_title="Profit and Loss", period="July 2026")
        assert qmr.is_cora_written(data) is True
        assert qmr.is_cora_written(_manual_xlsx()) is False
        assert qmr.is_cora_written(b"not a zip") is False

    def test_refuses_when_both_the_manual_and_the_cora_file_exist(self):
        """The -cora path was previously written with no existence check, so a
        second run overwrote its own sibling (D-051 HIGH)."""
        w, e, r, store = _fake_fs()
        outdir = qmr.target_dir("2026-07")
        store[str(outdir / "2026-07_hjrg_pl.xlsx")] = _manual_xlsx()
        store[str(outdir / "2026-07_hjrg_pl-cora.xlsx")] = b"EARLIER CORA RUN"
        summary = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                  writer=w, exists=e, reader=r)
        # Neither pre-existing P&L file may be touched...
        assert store[str(outdir / "2026-07_hjrg_pl-cora.xlsx")] == b"EARLIER CORA RUN"
        assert qmr.is_cora_written(store[str(outdir / "2026-07_hjrg_pl.xlsx")]) is False
        assert any("refusing to overwrite either" in x["reason"]
                   for x in summary["skipped"])
        assert all(x["kind"] != "pl" for x in summary["written"])
        # ...while the non-colliding balance sheet still lands. One blocked file
        # must not cost the entity its other statement.
        assert str(outdir / "2026-07_hjrg_bs.xlsx") in store

    def test_cora_variant_name(self):
        assert qmr.cora_variant_name("2026-07_hjrg_pl.xlsx") == \
            "2026-07_hjrg_pl-cora.xlsx"

    def test_parity_survives_an_unreadable_manual_file(self):
        w, e, r, store = _fake_fs()
        outdir = qmr.target_dir("2026-07")
        store[str(outdir / "2026-07_hjrg_pl.xlsx")] = b"not a zip"
        summary = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                  writer=w, exists=e, reader=r)
        assert str(outdir / "2026-07_hjrg_pl-cora.xlsx") in store
        assert summary["parity"][0].get("match") is None


class TestDryRunIsDefault:
    def test_dry_run_writes_nothing_but_reports_everything(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month("2026-07", sources=_sources(), apply=False,
                                  writer=w, exists=e, reader=r)
        assert store == {}
        assert len(summary["written"]) == 2   # what WOULD be written
        assert summary["applied"] is False

    def test_script_requires_apply_to_write(self):
        src = (_REPO_ROOT / "scripts" / "run_qbo_monthly_reports.py").read_text(
            encoding="utf-8")
        assert '"--apply", action="store_true"' in src
        assert "apply=args.apply" in src


class TestRendering:
    def test_layout_matches_the_qbo_export_shape(self):
        """Rows 1-5 mirror the archive's manual exports exactly."""
        from openpyxl import load_workbook
        data = qmr.render_xlsx(_pl(), company_name="HJR Global Services LLC",
                               report_title="Profit and Loss", period="May 2026")
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        rows = [tuple("" if c is None else str(c) for c in r)
                for r in ws.iter_rows(values_only=True)]
        assert rows[0][0] == "HJR Global Services LLC"
        assert rows[1][0] == "Profit and Loss"
        assert rows[2][0] == "May 2026"
        assert rows[4][1] == "Total"

    def test_values_are_carried_through_verbatim(self):
        """D-095: no arithmetic, no reformatting -- a float round-trip is a
        transformation and this job performs none."""
        rows = dict(qmr.report_rows(_pl()))
        assert rows["4275 Management Fees"] == "79580.19"
        assert rows["Total Income"] == "79580.19"

    def test_section_header_leaf_and_summary_all_appear_in_order(self):
        labels = [lbl for lbl, _ in qmr.report_rows(_pl())]
        assert labels == ["Income", "4275 Management Fees", "Total Income"]

    def test_nested_sections_are_walked(self):
        nested = {"Row": [{
            "Header": {"ColData": [{"value": "Expenses"}, {"value": ""}]},
            "Rows": {"Row": [{
                "Header": {"ColData": [{"value": "Payroll"}, {"value": ""}]},
                "Rows": {"Row": [{"ColData": [{"value": "5001 Salaries"},
                                              {"value": "51125.01"}]}]},
                "Summary": {"ColData": [{"value": "Total Payroll"},
                                        {"value": "51125.01"}]},
            }]},
            "Summary": {"ColData": [{"value": "Total Expenses"},
                                    {"value": "51125.01"}]},
        }]}
        labels = [lbl for lbl, _ in qmr.report_rows(_pl(rows=nested))]
        assert labels == ["Expenses", "Payroll", "5001 Salaries",
                          "Total Payroll", "Total Expenses"]

    def test_empty_report_renders_headers_without_crashing(self):
        assert qmr.report_rows(_pl(rows={})) == []

    def test_footer_records_basis_and_pull_time(self):
        """The archive's 2026-05 files were exported MID-month (their own footer
        says 2026-05-22), so their figures are an open-month snapshot. Without a
        stamp a reader cannot tell that from a closed-month statement."""
        stamp = qmr.footer_stamp(_pl())
        assert "Accrual Basis" in stamp
        assert "2026-08-18" in stamp

    def test_parity_needle_matches_both_wordings(self):
        """The REST API says 'Total Income'; the UI export says 'Total for
        Income'. One needle has to match a manual file and a Cora file."""
        assert qmr._normalize_label("Total for Income") == \
            qmr._normalize_label("Total Income")


class TestSummaryIsFigureFree:
    def test_no_financial_values_in_the_run_summary(self):
        """This text is safe to post anywhere. HRLLC is personal-expense data and
        the A5 lesson was two figure leaks into shared surfaces -- numbers stay
        in the files."""
        w, e, r, _ = _fake_fs()
        summary = qmr.build_month(
            "2026-07", sources=_sources(provisioned=lambda: ["HJRG"]),
            apply=True, writer=w, exists=e, reader=r)
        text = qmr.format_summary(summary)
        assert "79580" not in text and "79,580" not in text

    def test_summary_names_the_unmapped_entities_every_run(self):
        """No silent caps: a partially populated folder must never read as
        complete."""
        w, e, r, _ = _fake_fs()
        summary = qmr.build_month("2026-07", sources=_sources(), apply=True,
                                  writer=w, exists=e, reader=r)
        text = qmr.format_summary(summary)
        assert "still a manual export" in text
        assert "ufl" in text

    def test_summary_reports_dry_run_mode_explicitly(self):
        w, e, r, _ = _fake_fs()
        s = qmr.build_month("2026-07", sources=_sources(), apply=False,
                            writer=w, exists=e, reader=r)
        assert "DRY-RUN" in qmr.format_summary(s)

    def test_summary_surfaces_skips_and_the_target_folder(self):
        w, e, r, _ = _fake_fs()
        s = qmr.build_month(
            "2026-07", sources=_sources(provisioned=lambda: ["NEWCO"]),
            apply=False, writer=w, exists=e, reader=r)
        text = qmr.format_summary(s)
        assert "no slug mapping" in text
        assert "2026-08" in text

    def test_provisioned_listing_failure_is_reported_not_raised(self):
        def boom():
            raise RuntimeError("no token file")
        w, e, r, _ = _fake_fs()
        s = qmr.build_month("2026-07", sources=_sources(provisioned=boom),
                            apply=True, writer=w, exists=e, reader=r)
        assert "could not list provisioned realms" in s["error"]
        assert "ERROR" in qmr.format_summary(s)


class TestNoLLM:
    def test_module_makes_no_model_calls(self):
        """D-095 is structural here, not a promise: nothing in this module may
        reach a model client."""
        src = (_REPO_ROOT / "src" / "cora" / "qbo_monthly_reports.py").read_text(
            encoding="utf-8")
        for banned in ("anthropic", "claude_client", "openai", "batch_client",
                       "llm_usage"):
            assert banned not in src, f"{banned} must not appear"


class TestMonthInputIsStrict:
    """--month exists only for operator backfill, so a typo IS its main risk
    surface. A loose parse was destructive in three distinct ways (D-051)."""

    @pytest.mark.parametrize("bad", ["2026-7", "26-07", "2026-007", "2026/07",
                                     "202607", "1999-07"])
    def test_malformed_months_are_refused(self, bad):
        with pytest.raises(ValueError):
            qmr.month_bounds(bad)

    def test_unpadded_month_would_have_bypassed_the_collision_check(self):
        """2026-7 filed 2026-7_llc_pl.xlsx, which can never match the real
        2026-07_llc_pl.xlsx, so never-overwrite and parity were both bypassed."""
        with pytest.raises(ValueError):
            qmr.filename("2026-7", "llc", "pl")

    def test_two_digit_year_would_have_created_a_bogus_drive_folder(self):
        """26-07 resolved to a filing folder of 0026-08 and CREATED it on the
        shared Drive."""
        with pytest.raises(ValueError):
            qmr.filing_folder_for("26-07")

    def test_padded_month_still_works(self):
        assert qmr.month_bounds("2026-07") == ("2026-07-01", "2026-07-31")


class TestFutureMonthGuard:
    """QBO answers a future as-of date with TODAY balances, so a future --month
    produced a Balance Sheet labelled July 2027 holding current figures, filed
    under a fabricated 2027-08/ (D-051)."""

    def test_future_month_refused(self):
        import datetime
        with pytest.raises(ValueError):
            qmr.assert_month_is_complete("2027-07",
                                         today=datetime.date(2026, 8, 18))

    def test_current_incomplete_month_refused(self):
        import datetime
        with pytest.raises(ValueError):
            qmr.assert_month_is_complete("2026-08",
                                         today=datetime.date(2026, 8, 18))

    def test_last_completed_month_allowed(self):
        import datetime
        qmr.assert_month_is_complete("2026-07", today=datetime.date(2026, 8, 18))

    def test_apply_is_gated_on_the_guard_but_dry_run_is_not(self):
        src = (_REPO_ROOT / "scripts" / "run_qbo_monthly_reports.py").read_text(
            encoding="utf-8")
        assert "if args.apply:" in src
        assert "assert_month_is_complete(report_month)" in src


class TestPeriodIsVerifiedNotAssumed:
    def test_period_mismatch_refuses(self):
        """Basis was verified from the response and the period was not -- yet the
        period is exactly what the filename and the row-3 label assert."""
        def wrong_period(entity, start, end):
            rep = _pl()
            rep["Header"]["StartPeriod"] = "2026-06-01"
            rep["Header"]["EndPeriod"] = "2026-06-30"
            return rep

        w, e, r, store = _fake_fs()
        summary = qmr.build_month("2026-07",
                                  sources=_sources(profit_loss=wrong_period),
                                  apply=True, writer=w, exists=e, reader=r)
        assert any("period mismatch" in x["reason"] for x in summary["skipped"])
        assert all("_pl" not in x["file"] for x in summary["written"])

    def test_report_period_reader(self):
        assert qmr.report_period(_pl()) == ("2026-07-01", "2026-07-31")


class TestAmountsAreNumericCells:
    def test_amounts_land_as_numbers_so_the_sheet_can_sum(self):
        """Every one of the ~350 existing archive files stores its amount cells
        as floats. A text amount is a statement that looks right and is
        arithmetically inert (D-051)."""
        from openpyxl import load_workbook
        data = qmr.render_xlsx(_pl(), company_name="HJR Global Services LLC",
                               report_title="Profit and Loss", period="July 2026")
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        vals = {str(r[0]): r[1] for r in ws.iter_rows(values_only=True) if r and r[0]}
        assert isinstance(vals["4275 Management Fees"], float)
        assert vals["4275 Management Fees"] == pytest.approx(79580.19)

    def test_non_numeric_values_pass_through_unchanged(self):
        from openpyxl import load_workbook
        rep = _pl(rows={"Row": [{"ColData": [{"value": "Note"},
                                             {"value": "see attachment"}]}]})
        data = qmr.render_xlsx(rep, company_name="X", report_title="P",
                               period="p")
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        vals = {str(r[0]): r[1] for r in ws.iter_rows(values_only=True) if r and r[0]}
        assert vals["Note"] == "see attachment"

    def test_footer_separates_qbo_clock_from_coras_pull_time(self):
        """The footer used to print QBO time and label it Cora time -- which
        defeats its only purpose, telling a pre-close snapshot from a post-close
        statement (D-051)."""
        import datetime
        stamp = qmr.footer_stamp(_pl(),
                                 now=datetime.datetime(2026, 8, 18, 17, 42))
        assert "QBO report time 2026-08-18T17:00:00-07:00" in stamp
        assert "pulled by Cora 2026-08-18 17:42" in stamp
        assert "Accrual Basis" in stamp


class TestSensitiveBooksAreOptIn:
    """HRLLC is Harrison's PERSONAL expense book. The accounting archive it would
    land in IS swept into the KB (justin@lexingtonservices.com, drive_sweep: true,
    entity_default LEX) and the existing hjrllc files are ALREADY ingested
    mis-tagged as LEX / LEX-LLC -- so a #llc-* asker can retrieve a personal
    balance sheet. That exposure predates this job; writing 2 more files a month
    forever would industrialize it (D-051 HIGH).
    """

    def test_hrllc_ships_disabled(self):
        assert qmr.load_slug_map()["HRLLC"]["enabled"] is False

    def test_every_other_realm_is_enabled(self):
        for realm, spec in qmr.load_slug_map().items():
            if realm != "HRLLC":
                assert spec["enabled"] is True, realm

    def test_a_disabled_realm_is_never_written_but_IS_reported(self):
        w, e, r, store = _fake_fs()
        summary = qmr.build_month(
            "2026-07",
            sources=_sources(provisioned=lambda: ["HRLLC"],
                             company_name=lambda x: "Harrison Rogers, LLC"),
            apply=True, writer=w, exists=e, reader=r)
        assert store == {}, "a disabled realm must never be written"
        assert summary["written"] == []
        assert any("opt-in" in x["reason"] for x in summary["skipped"])
        # Silence would read as "this entity simply had no data".
        assert "disabled in the slug map" in qmr.format_summary(summary)

    def test_populator_slugs_all_have_a_deterministic_entity_mapping(self):
        """Any slug this job emits must resolve WITHOUT Haiku, or the file is
        guessed into whatever entity the sweeping mailbox defaults to (LEX).
        osn-core4 was the gap that made this test necessary."""
        from cora.connectors.drive_entity_detect import detect_entity_from_filename
        for realm, spec in qmr.load_slug_map().items():
            if not spec["enabled"]:
                continue   # HRLLC is deliberately unmapped AND unwritten
            name = qmr.filename("2026-07", spec["slug"], "pl")
            assert detect_entity_from_filename(name), (
                f"{realm} emits {spec['slug']!r} with no deterministic entity "
                f"mapping -- it would be Haiku-guessed on ingest")

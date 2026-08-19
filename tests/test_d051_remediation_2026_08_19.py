"""D-051 remediation pins for session #3 (2026-08-19).

Three adversarial lenses (correctness / leak-PHI-egress / broken-promises) ran
over `14ef69c..HEAD`. These pin the fixes whose defects a green suite did not
see. The two HIGHs each lens found independently are covered in their own files
(vendor-spend column read + LEX name opacity in test_qbo_transaction_detail;
the artifact-path leak below).
"""

from __future__ import annotations

import importlib.util
import json
import sys
from datetime import timedelta
from pathlib import Path
from types import SimpleNamespace

import pytest

from cora import delegated_work as dw
from cora import long_message as lm
from cora.tools import table_export as tx

_REPO_ROOT = Path(__file__).resolve().parents[1]
USER = "U0BREMED01"


def _code_only(path: Path) -> str:
    return "\n".join(
        line for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    )


# ---------------------------------------------------------------------------
# HIGH — the delivered artifact path re-emitted a suppressed title
# ---------------------------------------------------------------------------

@pytest.fixture
def ledgers(tmp_path, monkeypatch):
    bot = tmp_path / "b.jsonl"
    runner = tmp_path / "r.jsonl"
    bot.touch()
    runner.touch()
    monkeypatch.setattr(dw, "_BOT_LEDGER", bot)
    monkeypatch.setattr(dw, "_RUNNER_LEDGER", runner)
    return bot, runner


def _append(path: Path, row: dict) -> None:
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(row) + "\n")


def _delivered_job(bot: Path, runner: Path, *, channel="C_PRIVATE",
                   title="LBHS COPA settlement exposure memo"):
    ts = dw._now_iso()
    _append(bot, {"event": "requested", "ts": ts, "job_id": "dw-1",
                  "requester": USER, "archetype": "research_brief",
                  "entity": "LEX-LLC", "title": title, "brief": "b",
                  "channel_id": channel, "fingerprint": "fp"})
    _append(bot, {"event": "queued", "ts": ts, "job_id": "dw-1"})
    _append(runner, {"event": "delivered", "ts": dw._now_iso(), "job_id": "dw-1",
                     "artifact": {"target_path": (
                         r"G:\My Drive\HJR-Founder-OS\04-Lexington\_delegated-work"
                         r"\2026-08\2026-08-19_lex-llc_research-brief-lbhs-copa-"
                         r"settlement-exposure-me-a1b2c3.md")}})


class TestArtifactPathLeak:
    def test_the_path_is_suppressed_outside_the_originating_channel(self, ledgers):
        """The artifact FILENAME embeds a 40-char slug of the job title, so
        rendering the path in the always-shown detail clause re-emitted exactly
        what the same-channel gate exists to suppress. Two lenses found it
        independently; the original test could not, because its fixture seeded a
        QUEUED job and never reached the DELIVERED branch at all.
        """
        bot, runner = ledgers
        _delivered_job(bot, runner)
        out = dw.render_job_list(USER, "C_OTHER")
        assert "dw-1" in out
        assert "copa" not in out.lower()
        assert "settlement" not in out.lower()
        assert "_delegated-work" not in out

    def test_the_path_still_renders_in_the_originating_channel(self, ledgers):
        bot, runner = ledgers
        _delivered_job(bot, runner, channel="C_HOME")
        out = dw.render_job_list(USER, "C_HOME")
        assert "settlement-exposure" in out

    def test_an_empty_channel_id_suppresses_the_path(self, ledgers):
        bot, runner = ledgers
        _delivered_job(bot, runner)
        assert "_delegated-work" not in dw.render_job_list(USER, "")

    def test_state_detail_never_returns_a_path(self, ledgers):
        """Belt: the helper itself must not be able to produce one, so a future
        caller outside the gate cannot re-open this."""
        detail = dw._state_detail({
            "state": dw.STATE_DELIVERED,
            "artifact": {"target_path": r"G:\x\secret-title.md"},
        })
        assert "secret-title" not in detail
        assert detail == "delivered"


# ---------------------------------------------------------------------------
# The quota refund is capped, so a reproducible trip cannot loop it
# ---------------------------------------------------------------------------

class TestRefundCap:
    def test_refunds_are_capped_per_user_per_day(self, ledgers, monkeypatch):
        """content_guard is CHANNEL-caused by design and therefore reliably
        reproducible by the requester -- an uncapped refund turns "submit a
        financials-shaped brief in a TIER_3 channel" into unlimited jobs/day,
        with only the monthly envelope left as a backstop."""
        monkeypatch.setenv("CORA_DELEGATED_MAX_REFUNDS", "2")
        bot, runner = ledgers
        for n in range(5):
            ts = dw._now_iso()
            _append(bot, {"event": "requested", "ts": ts, "job_id": f"dw-{n}",
                          "requester": USER, "archetype": "research_brief",
                          "entity": "F3E", "title": "t", "brief": "b",
                          "channel_id": "C1", "fingerprint": f"fp{n}"})
            _append(bot, {"event": "queued", "ts": ts, "job_id": f"dw-{n}"})
            _append(runner, {"event": "failed", "ts": dw._now_iso(),
                             "job_id": f"dw-{n}",
                             "failure_class": dw.FAILURE_CONTENT_GUARD,
                             "message": "m"})
        assert dw.requested_today(USER) == 5
        assert dw.refunded_today(USER) == 2          # capped
        assert dw.quota_used_today(USER) == 3

    def test_the_cap_is_configurable_and_defaults_sanely(self):
        assert dw.max_refunds_per_day() >= 1


# ---------------------------------------------------------------------------
# long_message hardening
# ---------------------------------------------------------------------------

class TestLongMessageHardening:
    @pytest.mark.parametrize("limit", [0, -5])
    def test_a_non_positive_limit_terminates(self, limit):
        """limit=0 made every slice empty, so `rest` never shrank and the loop
        ran forever while the output list grew. Not reachable from today's
        callers, but `limit` is a public parameter."""
        out = lm.split_for_slack("some text that is longer than nothing", limit=limit)
        assert out and all(out)

    def test_the_weld_branch_is_gone(self):
        """Both branches were byte-identical, i.e. the comment promised a guard
        that did not exist in either direction."""
        src = _code_only(_REPO_ROOT / "src" / "cora" / "long_message.py")
        assert src.count('text = f"{text}{chunk}"') == 1

    def test_continuations_reach_the_spend_ledger(self):
        """A truncated weekly memo costs up to 3x what the ledger records if the
        continuation calls are not logged."""
        src = _code_only(_REPO_ROOT / "src" / "cora" / "long_message.py")
        assert "log_usage(" in src


class TestMultiPartSendsAreFailSoft:
    def test_memo_and_synthesis_do_not_abort_on_a_failed_part(self):
        """Both looped chat_postMessage inside ONE try, so a failure on part 2
        dropped parts 3..N after part 1 had already been posted."""
        for rel in ("src/cora/strategy_memo.py", "src/cora/channel_synthesis.py"):
            src = _code_only(_REPO_ROOT / rel)
            body = src[src.index("for i, part in enumerate(parts"):]
            body = body[:body.index("return ")]
            assert "except Exception" in body, rel


# ---------------------------------------------------------------------------
# table_export
# ---------------------------------------------------------------------------

class TestTableExportCap:
    def test_a_bullet_only_report_still_trips_the_cap_and_says_so(self):
        """The cap was checked only in the non-matching-line branch, so a report
        of pure bullets never tripped it -- build_xlsx just sliced at MAX_ROWS
        with no marker. And the marker it did append pushed the list to
        MAX_ROWS+1, exactly where that slice discarded it."""
        big = "\n".join(f"  \u2022 Acct {i}: {i}" for i in range(tx.MAX_ROWS + 500))
        rows = tx.rows_from_report(big)
        assert len(rows) <= tx.MAX_ROWS
        assert "truncated" in rows[-1][0]

    def test_the_marker_survives_into_the_workbook(self, tmp_path):
        from openpyxl import load_workbook
        big = "\n".join(f"  \u2022 Acct {i}: {i}" for i in range(tx.MAX_ROWS + 500))
        f = tmp_path / "o.xlsx"
        f.write_bytes(tx.build_xlsx(tx.rows_from_report(big), "R"))
        ws = load_workbook(f).active
        assert "truncated" in str(ws.cell(row=ws.max_row, column=1).value)

    def test_the_report_header_is_not_captured_as_a_section(self):
        """"Expense detail for HJRG (...):" ends in a colon and was being taken
        as the first Section, stamping the report title onto every row above the
        first real QBO section."""
        text = ("Expense detail for HJRG (2026-01-01 to 2026-03-31):\n"
                "Expenses:\n"
                "  \u2022 Rent: 1,000.00\n")
        rows = tx.rows_from_report(text)
        rent = next(r for r in rows if r[1] == "Rent")
        assert rent[0] == "Expenses"

    def test_a_leading_equals_is_stored_as_text_not_a_formula(self, tmp_path):
        from openpyxl import load_workbook
        f = tmp_path / "o.xlsx"
        f.write_bytes(tx.build_xlsx([["A"], ["=1+1"]], "R"))
        ws = load_workbook(f).active
        assert ws.cell(row=2, column=1).data_type == "s"


# ---------------------------------------------------------------------------
# Egress hardening in the shared uploader
# ---------------------------------------------------------------------------

class TestUploaderEgress:
    def test_title_and_filename_are_sanitized_structurally(self):
        """The bytes are PUT straight to Slack, bypassing the WebClient egress
        patch, and `title` is shared via completeUploadExternal -- so the
        guarantee has to live in the shared function, not in each caller."""
        src = _code_only(_REPO_ROOT / "src" / "cora" / "tools" / "slack_file_upload.py")
        block = src[src.index("def upload_bytes("):]
        block = block[:block.index("def upload_text(")]
        assert "sanitize_text(title)" in block
        assert "sanitize_text(filename)" in block

    def test_a_signed_upload_url_is_scrubbed_before_logging(self):
        """These lines are archived by compact_logs and bundled into the offsite
        DR backup; an httpx transport error embeds the signed upload URL."""
        src = _code_only(_REPO_ROOT / "src" / "cora" / "tools" / "slack_file_upload.py")
        assert "upload-url redacted" in src


# ---------------------------------------------------------------------------
# Archive-only slugs must not match corpus-wide
# ---------------------------------------------------------------------------

class TestArchiveOnlySlugScoping:
    @pytest.mark.parametrize("title", [
        "MV.xlsx", "LexCorp_Balance+Sheet.xlsx", "Lexcorp.xlsx",
        "LexCorp_Profit+and+Loss+by+Month.xlsx",
    ])
    def test_real_founder_os_files_are_not_pulled_into_lex(self, title):
        """The detector runs over EVERY swept mailbox and overrides Haiku on a
        bare token match. `mv` is two characters. These are real Founder-OS files
        currently tagged HJRG/FNDR -- verified live against 53,350 distinct Drive
        titles -- and a corpus-wide match would move them into LEX, the
        firewalled entity, on their next sweep.
        """
        from cora.connectors.drive_entity_detect import detect_entity_from_filename
        assert detect_entity_from_filename(title) is None

    @pytest.mark.parametrize("title,expected", [
        ("2026-05_mv_pl.xlsx", "LEX-LLA"),
        ("2026-05_lexcorp_bs.xlsx", "LEX"),
        ("2026-05_f3comm_cf.xlsx", "F3C"),
        ("2026-05_hjrpod_pl.xlsx", "HJRPROD"),
        ("2026-05_osn-core4_bs.xlsx", "OSN"),
    ])
    def test_the_dated_archive_convention_still_resolves(self, title, expected):
        from cora.connectors.drive_entity_detect import detect_entity_from_filename
        assert detect_entity_from_filename(title) == expected

    def test_globally_safe_codes_are_unaffected(self):
        """Only the archive-only additions are date-scoped; the long-standing
        codes keep matching undated files."""
        from cora.connectors.drive_entity_detect import detect_entity_from_filename
        assert detect_entity_from_filename("OSN_Master Workbook.xlsx") == "OSN"
        assert detect_entity_from_filename("hjrp_lease.pdf") == "HJRP"

    def test_the_sweep_and_the_purge_share_one_date_predicate(self):
        """"Is this an archive file?" must have exactly one answer, or the two
        disagree about which rows are in scope."""
        from cora.connectors import drive_entity_detect as ded
        src = _code_only(
            _REPO_ROOT / "scripts" / "purge_kb_personal_books_2026-08-19.py")
        assert "has_date_token" in src
        assert ded.has_date_token("2026-05_mv_pl.xlsx") is True
        assert ded.has_date_token("MV.xlsx") is False


# ---------------------------------------------------------------------------
# The re-tag must move the entity in BOTH stores that hold it
# ---------------------------------------------------------------------------

class TestRetagTouchesTheVectorIndex:
    def test_apply_retag_rewrites_the_bin_row(self):
        """`entity` is not metadata: it is a vec0 PARTITION KEY on the coarse bin
        table as well as a column on knowledge_chunks. Updating only the latter
        makes a chunk unreachable under BOTH entities -- never a coarse candidate
        under the new one, filtered out at re-rank under the old. The first cut
        called itself "metadata-only", which is true of sub_entity and false of
        entity, and would have darkened all 52 rows it claimed to be fixing.
        """
        path = _REPO_ROOT / "scripts" / "purge_kb_personal_books_2026-08-19.py"
        src = _code_only(path)
        block = src[src.index("def apply_retag("):src.index("def main(")]
        assert "bin_tables_present" in block
        assert "vec_quantize_binary" in block
        assert "knowledge_vec_f32" in block
        assert "UPDATE knowledge_chunks" in block

    def test_the_docstring_no_longer_claims_no_vec_table_work(self):
        src = (_REPO_ROOT / "scripts" / "purge_kb_personal_books_2026-08-19.py"
               ).read_text(encoding="utf-8")
        assert "no vec-table work" not in src


# ---------------------------------------------------------------------------
# The hot Slack path marks a truncated reply
# ---------------------------------------------------------------------------

class TestHotPathTruncationMarker:
    def test_a_max_tokens_stop_is_marked_for_the_reader(self):
        """The script surfaces CONTINUE; the interactive path is
        latency-sensitive and instead says so. Either way the reader can tell a
        short answer from a severed one -- which the log line alone did not give
        them, and the commit had claimed the family was closed."""
        from cora import claude_client as cc
        out = cc._mark_if_truncated(SimpleNamespace(stop_reason="max_tokens"), "half")
        assert "cut short" in out
        assert cc._mark_if_truncated(
            SimpleNamespace(stop_reason="end_turn"), "whole") == "whole"

    def test_it_is_idempotent_and_skips_an_empty_reply(self):
        from cora import claude_client as cc
        once = cc._mark_if_truncated(SimpleNamespace(stop_reason="max_tokens"), "x")
        assert cc._mark_if_truncated(
            SimpleNamespace(stop_reason="max_tokens"), once) == once
        assert cc._mark_if_truncated(
            SimpleNamespace(stop_reason="max_tokens"), "") == ""

    def test_both_return_paths_are_marked(self):
        from cora import claude_client as cc
        src = _code_only(_REPO_ROOT / "src" / "cora" / "claude_client.py")
        assert src.count("_mark_if_truncated(") >= 3   # def + both returns


# ---------------------------------------------------------------------------
# autowrite_excluded covers BOTH auto-write branches
# ---------------------------------------------------------------------------

class TestAutowriteExclusionCoversTierOne:
    def test_an_excluded_domain_owner_is_not_an_authorized_owner(self, monkeypatch):
        """The flag was applied only in contributor_recognized, which feeds
        TIER_0. TIER_1 / DECISION_OWNER is reached through authorized_owner, and
        _autowrite_eligible auto-writes tier 1 at CORA_AUTOWRITE_LIVE=all -- so
        an excluded person who is also a domain owner kept half the authority the
        flag documents itself as removing."""
        from cora import graduated_trust_shadow as gts

        rec = SimpleNamespace(external=False, autowrite_excluded=True,
                              entity="F3E", entities=["F3E"],
                              all_entities=["F3E"])
        monkeypatch.setattr(gts, "_role_for", lambda sid: rec)
        monkeypatch.setattr(
            "cora.gap_autofill.resolve_owner", lambda entity: "U_OWNER")
        assert gts.authorized_owner("U_OWNER", "F3E") is False

    def test_a_non_excluded_owner_is_still_an_owner(self, monkeypatch):
        """Positive control -- the guard must exclude one person, not break the
        mechanism."""
        from cora import graduated_trust_shadow as gts

        rec = SimpleNamespace(external=False, autowrite_excluded=False,
                              entity="F3E", entities=["F3E"],
                              all_entities=["F3E"])
        monkeypatch.setattr(gts, "_role_for", lambda sid: rec)
        monkeypatch.setattr(
            "cora.gap_autofill.resolve_owner", lambda entity: "U_OWNER")
        assert gts.authorized_owner("U_OWNER", "F3E") is True


# ---------------------------------------------------------------------------
# Approval recon classification
# ---------------------------------------------------------------------------

@pytest.fixture
def recon(tmp_path):
    path = _REPO_ROOT / "scripts" / "run_approval_recon.py"
    spec = importlib.util.spec_from_file_location("approval_recon_rem", path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["approval_recon_rem"] = mod
    spec.loader.exec_module(mod)
    for key in list(mod.LEDGERS):
        mod.LEDGERS[key] = tmp_path / f"{key}.jsonl"
        mod.LEDGERS[key].touch()
    return mod


class TestApprovalReconClassification:
    def test_a_withdrawn_job_is_not_still_waiting_on_an_approver(self, recon):
        """Only `harrison_dismiss` is the approver saying no. Every other cancel
        is terminal but nobody is waiting -- counting them `open` accrued
        unbounded "wait" on jobs nobody was waiting for, and that number is
        exactly what the second-approver recommendation turns on."""
        from datetime import datetime, timezone
        now = datetime(2026, 8, 19, 12, 0, tzinfo=timezone.utc)

        def iso(h):
            return (now - timedelta(hours=h)).isoformat()

        with (recon.LEDGERS["delegated_work"]).open("w", encoding="utf-8") as fh:
            for row in [
                {"event": "held", "ts": iso(50), "job_id": "a"},
                {"event": "cancelled", "ts": iso(40), "job_id": "a",
                 "reason": "requester_cancel"},
                {"event": "held", "ts": iso(50), "job_id": "b"},
                {"event": "cancelled", "ts": iso(40), "job_id": "b",
                 "reason": "harrison_cancel"},
                {"event": "held", "ts": iso(50), "job_id": "c"},
                {"event": "cancelled", "ts": iso(40), "job_id": "c",
                 "reason": "harrison_dismiss"},
            ]:
                fh.write(json.dumps(row) + "\n")

        lane = next(l for l in recon.analyze(now=now, days=30)["lanes"]
                    if l["lane"] == "delegated_work")
        assert lane["withdrawn"] == 2
        assert lane["denied"] == 1
        assert lane["open"] == 0
        assert lane["open_wait_total_h"] == 0

    def test_a_recurrence_does_not_reopen_a_shipped_code_item(self, recon):
        """`status or event` treated every event name as a status, so a teammate
        re-mentioning an already-SHIPPED item appended `recurrence` and the item
        silently reverted to open with its full age added to the accrued wait.
        Verified live: 19 dm_sent / 8 recurrence / 2 evidence rows already
        present in the ledger."""
        from datetime import datetime, timezone
        now = datetime(2026, 8, 19, 12, 0, tzinfo=timezone.utc)

        def iso(h):
            return (now - timedelta(hours=h)).isoformat()

        with (recon.LEDGERS["code"]).open("w", encoding="utf-8") as fh:
            for row in [
                {"event": "captured", "ts": iso(100), "id": "cq-1",
                 "status": "PROPOSED"},
                {"event": "shipped", "ts": iso(80), "id": "cq-1"},
                {"event": "recurrence", "ts": iso(2), "id": "cq-1"},
                {"event": "dm_sent", "ts": iso(1), "id": "cq-1"},
            ]:
                fh.write(json.dumps(row) + "\n")

        lane = next(l for l in recon.analyze(now=now, days=30)["lanes"]
                    if l["lane"] == "code")
        assert lane["approved"] == 1
        assert lane["open"] == 0

    def test_transitions_are_read_from_the_event_name(self, recon):
        """Verified against the live ledger: `status` appears ONLY on the
        `captured` seed row; every transition rides the event name."""
        src = _code_only(_REPO_ROOT / "scripts" / "run_approval_recon.py")
        assert "_EVENT_OUTCOME" in src
        assert '"status") or ev.get("event")' not in src

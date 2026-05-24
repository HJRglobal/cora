"""Tests for the f3e_inventory_pulse tool.

Coverage:
  - inventory_client: _parse_unis, _parse_nimbl, _parse_office, format_inventory_pulse
  - Safety-stock flag thresholds (CRITICAL / LOW / healthy)
  - Merged row logic (DAMAGE aggregation, Nimbl-only SKUs)
  - Empty report / missing sheets
  - Drive error → UNKNOWN_RESPONSE passthrough
  - Tool handler _tool_f3e_inventory_pulse
  - TOOL_DEFINITIONS entry present with required fields
  - _TOOL_FUNCTIONS registration
  - dispatch() integration
  - Source opacity (no file IDs, Drive URLs, sheet names in output)
"""
from __future__ import annotations

import sys
from pathlib import Path

# Ensure src/ is on sys.path (matches pattern used by all other F3E test files)
_REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO_ROOT / "src"))

import io
import types
from unittest.mock import MagicMock, patch

import openpyxl
import pytest


# ── Helpers ────────────────────────────────────────────────────────────────────

def _make_workbook(
    unis_rows: list[list],
    nimbl_rows: list[list],
    office_rows: list[list],
) -> bytes:
    """Build an in-memory xlsx with UNIS / NIMBL / 117 office sheets."""
    wb = openpyxl.Workbook()

    # UNIS sheet
    ws_unis = wb.active
    ws_unis.title = "UNIS"
    # Header row (col 0=Facility, 1=Item ID, 9=Available, 11=Allocated, 16=On Hand, 18=Goods Type)
    header = ["Facility", "Item ID", "UPC Code", "Short Desc", "Description",
              "Units/pkg", "UnitCuFt", "Customer", "Title",
              "Available", "Receiving", "Allocated", "Damaged", "Hold",
              "Incoming", "Open Order", "On Hand", "UOM", "Goods Type"]
    ws_unis.append(header)
    for row in unis_rows:
        ws_unis.append(row)

    # NIMBL sheet
    ws_nimbl = wb.create_sheet("NIMBL")
    nimbl_header = ["Business Unit - Code", "Location - Number", "Item - Number",
                    "Item - Short Description", "Quantity", "Location - Storage Type",
                    "Lot Number", "Expiration Date"]
    ws_nimbl.append(nimbl_header)
    for row in nimbl_rows:
        ws_nimbl.append(row)

    # 117 office sheet
    ws_office = wb.create_sheet("117 office")
    office_header = ["Item ID", "Short Description", "Available", "Damaged"]
    ws_office.append(office_header)
    for row in office_rows:
        ws_office.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _unis_row(item_id: str, available: float, allocated: float = 0,
              on_hand: float = None, goods_type: str = "GOOD") -> list:
    """Build a 19-column UNIS data row."""
    if on_hand is None:
        on_hand = available
    row = ["Cotton", item_id, "000000000000", "Short Desc", "Description",
           1.0, 0.25, "F3 ENERGY LLC", "F3 ENERGY LLC",
           available, 0,         # Available, Receiving
           allocated, 0, 0,     # Allocated, Damaged, Hold
           0, 0,                # Incoming, Open Order
           on_hand,             # On Hand
           "CS", goods_type]    # UOM, Goods Type
    return row


def _nimbl_row(item_id: str, qty: int) -> list:
    return ["F3E", "K100", item_id, "Some description", qty, "Reserve Storage",
            "LOT123", "2027-01-01"]


def _office_row(item_id: str, available: float, damaged: float = 0) -> list:
    return [item_id, "Short Description", available, damaged or None]


# ── Test fixtures ──────────────────────────────────────────────────────────────

HEALTHY_UNIS = [
    _unis_row("F3-Original",  3000),
    _unis_row("F3-Citrus",    2500),
    _unis_row("F3-Tropical",  2800),
    _unis_row("F3SL",         3200),
    _unis_row("F3-Orange",    2200),
    _unis_row("F3-Peach",     3000),
    _unis_row("F3SC",         4000),
    _unis_row("F3PC",         4200),
    _unis_row("F3VPM4",       2000),
    _unis_row("F3VPE4",       500),
    _unis_row("F3-Original24pk", 700),
]

MINIMAL_UNIS = [_unis_row("F3-Original", 1500)]
MINIMAL_NIMBL = [_nimbl_row("F3-Original", 300)]
MINIMAL_OFFICE = [_office_row("F3-Original", 50)]


# ────────────────────────────────────────────────────────────────────────────────
# Class 1: UNIS sheet parsing
# ────────────────────────────────────────────────────────────────────────────────

class TestUnisSheetParsing:
    def _parse(self, rows):
        from cora.tools.inventory_client import _parse_xlsx
        data = _make_workbook(rows, [], [])
        unis, _, _ = _parse_xlsx(data)
        return unis

    def test_basic_row_parsed(self):
        unis = self._parse([_unis_row("F3-Original", 3149)])
        assert "F3-Original" in unis
        assert unis["F3-Original"]["available"] == 3149

    def test_allocated_parsed(self):
        unis = self._parse([_unis_row("F3-Citrus", 2933, allocated=1)])
        assert unis["F3-Citrus"]["allocated"] == 1

    def test_damage_row_goes_to_damaged_not_available(self):
        unis = self._parse([
            _unis_row("F3-Peach", 0, on_hand=2, goods_type="DAMAGE"),
            _unis_row("F3-Peach", 3966),
        ])
        assert unis["F3-Peach"]["available"] == 3966
        assert unis["F3-Peach"]["damaged"] == 2

    def test_empty_item_id_skipped(self):
        empty = ["Cotton", "", "850045501167", "desc", "desc", 1, 0.25,
                 "F3 ENERGY LLC", "F3 ENERGY LLC", 0, 0, 0, 0, 0, 0, 0, 0, "CS", "GOOD"]
        unis = self._parse([empty])
        assert "" not in unis

    def test_multiple_good_rows_same_sku_aggregated(self):
        unis = self._parse([
            _unis_row("F3-Original", 1000),
            _unis_row("F3-Original", 500),
        ])
        assert unis["F3-Original"]["available"] == 1500

    def test_missing_unis_sheet_returns_empty(self):
        from cora.tools.inventory_client import _parse_xlsx
        # Make workbook without UNIS sheet
        wb = openpyxl.Workbook()
        wb.active.title = "OTHER"
        buf = io.BytesIO(); wb.save(buf)
        unis, _, _ = _parse_xlsx(buf.getvalue())
        assert unis == {}


# ────────────────────────────────────────────────────────────────────────────────
# Class 2: NIMBL sheet parsing
# ────────────────────────────────────────────────────────────────────────────────

class TestNimblSheetParsing:
    def _parse_nimbl(self, rows):
        from cora.tools.inventory_client import _parse_xlsx
        data = _make_workbook([], rows, [])
        _, nimbl, _ = _parse_xlsx(data)
        return nimbl

    def test_single_row(self):
        nimbl = self._parse_nimbl([_nimbl_row("F3SL", 208)])
        assert nimbl["F3SL"] == 208

    def test_multiple_lots_aggregated(self):
        nimbl = self._parse_nimbl([
            _nimbl_row("F3-Original", 211),
            _nimbl_row("F3-Original", 200),
            _nimbl_row("F3-Original", 208),
        ])
        assert nimbl["F3-Original"] == 619

    def test_pure_sku_aggregated(self):
        nimbl = self._parse_nimbl([
            _nimbl_row("PURE-Original", 208),
            _nimbl_row("PURE-Original", 208),
            _nimbl_row("PURE-Original", 127),
        ])
        assert nimbl["PURE-Original"] == 543

    def test_different_skus_independent(self):
        nimbl = self._parse_nimbl([
            _nimbl_row("F3-Original", 600),
            _nimbl_row("F3SL", 531),
        ])
        assert nimbl["F3-Original"] == 600
        assert nimbl["F3SL"] == 531

    def test_missing_nimbl_sheet_returns_empty(self):
        from cora.tools.inventory_client import _parse_xlsx
        wb = openpyxl.Workbook(); wb.active.title = "UNIS"
        buf = io.BytesIO(); wb.save(buf)
        _, nimbl, _ = _parse_xlsx(buf.getvalue())
        assert nimbl == {}


# ────────────────────────────────────────────────────────────────────────────────
# Class 3: 117 office sheet parsing
# ────────────────────────────────────────────────────────────────────────────────

class TestOfficeSheetParsing:
    def _parse_office(self, rows):
        from cora.tools.inventory_client import _parse_xlsx
        data = _make_workbook([], [], rows)
        _, _, office = _parse_xlsx(data)
        return office

    def test_basic_row(self):
        office = self._parse_office([_office_row("F3VPE4", 69)])
        assert office["F3VPE4"]["available"] == 69

    def test_damaged_column(self):
        office = self._parse_office([_office_row("F3-Peach", 236, damaged=8)])
        assert office["F3-Peach"]["damaged"] == 8

    def test_missing_office_sheet_returns_empty(self):
        from cora.tools.inventory_client import _parse_xlsx
        wb = openpyxl.Workbook(); wb.active.title = "UNIS"
        buf = io.BytesIO(); wb.save(buf)
        _, _, office = _parse_xlsx(buf.getvalue())
        assert office == {}


# ────────────────────────────────────────────────────────────────────────────────
# Class 4: Safety stock flag thresholds
# ────────────────────────────────────────────────────────────────────────────────

class TestSafetyStockFlags:
    def test_critical_at_zero(self):
        from cora.tools.inventory_client import _flag
        assert _flag(0) == "🚨"

    def test_critical_at_fifty(self):
        from cora.tools.inventory_client import _flag
        assert _flag(50) == "🚨"

    def test_low_at_51(self):
        from cora.tools.inventory_client import _flag
        assert "⚠️" in _flag(51)

    def test_low_at_200(self):
        from cora.tools.inventory_client import _flag
        assert "⚠️" in _flag(200)

    def test_healthy_at_201(self):
        from cora.tools.inventory_client import _flag
        assert _flag(201) == "✅"

    def test_healthy_large_number(self):
        from cora.tools.inventory_client import _flag
        assert _flag(5000) == "✅"


# ────────────────────────────────────────────────────────────────────────────────
# Class 5: format_inventory_pulse output structure
# ────────────────────────────────────────────────────────────────────────────────

class TestFormatInventoryPulse:
    def _format(self, unis=None, nimbl=None, office=None, modified="2026-05-08T00:52:12Z"):
        from cora.tools.inventory_client import format_inventory_pulse
        return format_inventory_pulse(
            unis   or {"F3-Original": {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}},
            nimbl  or {},
            office or {},
            modified,
        )

    def test_header_present(self):
        out = self._format()
        assert "F3 Inventory Pulse" in out

    def test_report_date_in_output(self):
        out = self._format()
        assert "2026-05-08" in out

    def test_brand_sections_present(self):
        unis = {k: {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}
                for k in ("F3-Original", "F3-Orange", "PURE-Original")}
        out = self._format(unis=unis)
        assert "F3 Energy" in out
        assert "F3 Mood" in out
        assert "F3 Pure" in out

    def test_critical_sku_in_alerts(self):
        unis = {"F3VPE4": {"available": 0, "allocated": 0, "on_hand": 1, "damaged": 0}}
        out = self._format(unis=unis)
        assert "Alerts" in out
        assert "Energy Variety Pack" in out
        assert "🚨" in out

    def test_low_sku_in_alerts(self):
        unis = {"F3-Citrus": {"available": 150, "allocated": 0, "on_hand": 150, "damaged": 0}}
        out = self._format(unis=unis)
        assert "Alerts" in out
        assert "⚠️" in out

    def test_healthy_sku_not_in_alerts(self):
        unis = {"F3-Original": {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}}
        out = self._format(unis=unis)
        assert "Alerts" not in out

    def test_no_alerts_section_when_all_healthy(self):
        unis = {k: {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}
                for k in ("F3-Original", "F3SL")}
        out = self._format(unis=unis)
        assert "⚠️ Alerts" not in out

    def test_nimbl_qty_shown_in_detail(self):
        out = self._format(
            unis={"F3-Original": {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}},
            nimbl={"F3-Original": 619},
        )
        assert "Nimbl" in out
        assert "619" in out

    def test_damaged_shown_when_nonzero(self):
        out = self._format(
            unis={"F3-Peach": {"available": 3966, "allocated": 0, "on_hand": 3967, "damaged": 2}},
        )
        assert "damaged" in out.lower()

    def test_legend_present(self):
        out = self._format()
        assert "Critical" in out
        assert "Low" in out

    def test_empty_rows_returns_no_data_message(self):
        from cora.tools.inventory_client import format_inventory_pulse
        out = format_inventory_pulse({}, {}, {}, "2026-05-08T00:52:12Z")
        assert "No inventory data" in out

    def test_unknown_sku_skipped_gracefully(self):
        unis = {"UNKNOWN-SKU-XYZ": {"available": 100, "allocated": 0, "on_hand": 100, "damaged": 0}}
        # Should not raise; unknown SKU just won't appear in output
        out = self._format(unis=unis)
        assert isinstance(out, str)

    def test_nimbl_only_sku_appears(self):
        """A Pure SKU that exists only at Nimbl (not yet in Cotton) should appear."""
        out = self._format(
            unis={},
            nimbl={"PURE-Original": 2415},
        )
        assert "Pure Original" in out


# ────────────────────────────────────────────────────────────────────────────────
# Class 6: Merged row logic
# ────────────────────────────────────────────────────────────────────────────────

class TestMergedRowLogic:
    def test_total_avail_sums_all_locations(self):
        from cora.tools.inventory_client import _build_inventory_rows, _total_avail
        unis   = {"F3-Original": {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}}
        nimbl  = {"F3-Original": 619}
        office = {"F3-Original": {"available": 116, "damaged": 0}}
        rows = _build_inventory_rows(unis, nimbl, office)
        assert _total_avail(rows["F3-Original"]) == 3000 + 619 + 116

    def test_cotton_damage_excluded_from_avail(self):
        from cora.tools.inventory_client import _build_inventory_rows, _total_avail
        unis = {"F3-Peach": {"available": 3966, "allocated": 0, "on_hand": 3967, "damaged": 2}}
        rows = _build_inventory_rows(unis, {}, {})
        assert _total_avail(rows["F3-Peach"]) == 3966   # not 3968

    def test_nimbl_only_sku_included(self):
        from cora.tools.inventory_client import _build_inventory_rows
        rows = _build_inventory_rows({}, {"PURE-Original": 2415}, {})
        assert "PURE-Original" in rows
        assert rows["PURE-Original"]["nimbl_qty"] == 2415
        assert rows["PURE-Original"]["unis_avail"] == 0


# ────────────────────────────────────────────────────────────────────────────────
# Class 7: Drive error → UNKNOWN_RESPONSE
# ────────────────────────────────────────────────────────────────────────────────

class TestDriveErrors:
    def _call(self, service_side_effect=None, find_side_effect=None,
              download_side_effect=None):
        from cora.tools.inventory_client import get_f3e_inventory_pulse_text
        with patch("cora.tools.inventory_client._build_service") as mock_svc, \
             patch("cora.tools.inventory_client._find_latest_file") as mock_find, \
             patch("cora.tools.inventory_client._download_file") as mock_dl:

            if service_side_effect:
                mock_svc.side_effect = service_side_effect
            else:
                mock_svc.return_value = MagicMock()

            if find_side_effect:
                mock_find.side_effect = find_side_effect
            elif not service_side_effect:
                mock_find.return_value = ("fake_id", "2026-05-08T00:52:12Z")

            if download_side_effect:
                mock_dl.side_effect = download_side_effect
            elif not service_side_effect and not find_side_effect:
                # Return a valid workbook with no data
                data = _make_workbook(MINIMAL_UNIS, MINIMAL_NIMBL, MINIMAL_OFFICE)
                mock_dl.return_value = data

            return get_f3e_inventory_pulse_text()

    def test_service_build_failure_returns_unknown_response(self):
        from cora.tools.inventory_client import InventoryClientError
        out = self._call(service_side_effect=Exception("auth failed"))
        assert "I don't have that right now" in out

    def test_drive_find_failure_returns_unknown_response(self):
        from cora.tools.inventory_client import InventoryClientError
        out = self._call(find_side_effect=InventoryClientError("not found"))
        assert "I don't have that right now" in out

    def test_drive_download_failure_returns_unknown_response(self):
        from cora.tools.inventory_client import InventoryClientError
        out = self._call(download_side_effect=InventoryClientError("download failed"))
        assert "I don't have that right now" in out

    def test_happy_path_returns_pulse(self):
        out = self._call()
        assert "Inventory Pulse" in out


# ────────────────────────────────────────────────────────────────────────────────
# Class 8: Tool handler
# ────────────────────────────────────────────────────────────────────────────────

class TestToolHandler:
    def test_handler_calls_get_text(self):
        from cora.tools.tool_dispatch import _tool_f3e_inventory_pulse
        with patch("cora.tools.inventory_client.get_f3e_inventory_pulse_text",
                   return_value="*📦 F3 Inventory Pulse* _as of 2026-05-08_\n...") as mock_fn:
            result = _tool_f3e_inventory_pulse("U123", "F3E", {})
        mock_fn.assert_called_once()
        assert "Inventory Pulse" in result

    def test_handler_ignores_input_dict(self):
        from cora.tools.tool_dispatch import _tool_f3e_inventory_pulse
        with patch("cora.tools.inventory_client.get_f3e_inventory_pulse_text",
                   return_value="ok"):
            # extra keys in _input should not cause errors
            result = _tool_f3e_inventory_pulse("U123", "F3E", {"unexpected_key": "val"})
        assert result == "ok"


# ────────────────────────────────────────────────────────────────────────────────
# Class 9: TOOL_DEFINITIONS entry
# ────────────────────────────────────────────────────────────────────────────────

class TestToolDefinitionsEntry:
    def _entry(self):
        from cora.tools.tool_dispatch import TOOL_DEFINITIONS
        return next((t for t in TOOL_DEFINITIONS if t["name"] == "f3e_inventory_pulse"), None)

    def test_entry_exists(self):
        assert self._entry() is not None

    def test_has_description(self):
        entry = self._entry()
        assert "description" in entry
        assert len(entry["description"]) > 50

    def test_has_input_schema(self):
        entry = self._entry()
        assert "input_schema" in entry
        assert entry["input_schema"]["type"] == "object"

    def test_description_mentions_inventory(self):
        entry = self._entry()
        desc = entry["description"].lower()
        assert "inventory" in desc

    def test_description_mentions_safety_flags(self):
        entry = self._entry()
        desc = entry["description"]
        assert "Critical" in desc or "critical" in desc

    def test_description_channel_scoped(self):
        entry = self._entry()
        desc = entry["description"]
        assert "#f3e-ops" in desc or "f3e-ops" in desc

    def test_source_opacity_in_description(self):
        entry = self._entry()
        desc = entry["description"]
        assert "source-opaque" in desc or "Drive" not in desc


# ────────────────────────────────────────────────────────────────────────────────
# Class 10: _TOOL_FUNCTIONS registration
# ────────────────────────────────────────────────────────────────────────────────

class TestToolFunctionsRegistration:
    def test_registered(self):
        from cora.tools.tool_dispatch import _TOOL_FUNCTIONS
        assert "f3e_inventory_pulse" in _TOOL_FUNCTIONS

    def test_callable(self):
        from cora.tools.tool_dispatch import _TOOL_FUNCTIONS
        assert callable(_TOOL_FUNCTIONS["f3e_inventory_pulse"])


# ────────────────────────────────────────────────────────────────────────────────
# Class 11: dispatch() integration
# ────────────────────────────────────────────────────────────────────────────────

class TestDispatchIntegration:
    def test_dispatch_routes_to_handler(self):
        from cora.tools.tool_dispatch import dispatch
        with patch("cora.tools.inventory_client.get_f3e_inventory_pulse_text",
                   return_value="*📦 F3 Inventory Pulse*"):
            result = dispatch("f3e_inventory_pulse", "U123", "F3E", {})
        assert "Inventory Pulse" in result

    def test_dispatch_unknown_tool_still_works(self):
        from cora.tools.tool_dispatch import dispatch
        result = dispatch("nonexistent_tool_xyz", "U123", "F3E", {})
        assert "Unknown tool" in result


# ────────────────────────────────────────────────────────────────────────────────
# Class 12: Source opacity
# ────────────────────────────────────────────────────────────────────────────────

class TestSourceOpacity:
    def _run(self, unis=None, nimbl=None, office=None):
        from cora.tools.inventory_client import format_inventory_pulse
        return format_inventory_pulse(
            unis   or {"F3-Original": {"available": 3000, "allocated": 0, "on_hand": 3000, "damaged": 0}},
            nimbl  or {},
            office or {},
            "2026-05-08T00:52:12Z",
        )

    def test_no_drive_url_in_output(self):
        out = self._run()
        assert "drive.google.com" not in out
        assert "docs.google.com" not in out

    def test_no_file_id_in_output(self):
        out = self._run()
        assert "1lXaayKKm" not in out  # Real file ID from Drive search

    def test_no_sheet_names_in_output(self):
        out = self._run()
        assert "UNIS" not in out
        assert "NIMBL" not in out
        assert "117 office" not in out

    def test_no_xlsx_filename_in_output(self):
        out = self._run()
        assert "Weekly Inventory Report" not in out
        assert ".xlsx" not in out

    def test_tool_description_does_not_mention_drive_file(self):
        from cora.tools.tool_dispatch import TOOL_DEFINITIONS
        entry = next(t for t in TOOL_DEFINITIONS if t["name"] == "f3e_inventory_pulse")
        assert "1lXaayKKm" not in entry["description"]
        assert "Weekly Inventory Report.xlsx" not in entry["description"]

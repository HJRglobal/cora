"""Tests for the delegated-work worker + archetypes + runner (S2/S3).

Covers: two-phase loop caps (turns / searches incl. pause_turn clamp / fetches /
kb_search 8 / $2 partial delivery), kb_search pinning (entity computed in code,
owned_emails(requester), non-custodian scrub), allowlist hard-refusal + the
zero-write-tool structural invariant, artifact guard trip -> content_guard,
xlsx spec validation + assembly, artifact pathing (parent collapse + dwid6),
lockfile stale/pid/fail-closed, crash recovery both arms, expiry, mis-homed
retry idempotency, runner flag semantics, envelope re-check at claim, and the
D-047 no-bot-import guard.
"""
from __future__ import annotations

import importlib.util
import io
import json
import subprocess
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest

import cora.delegated_archetypes as arch
import cora.delegated_work as dw
import cora.delegated_worker as worker
import cora.llm_rates as llm_rates

_REPO_ROOT = Path(__file__).resolve().parents[1]

# ---------------------------------------------------------------------------
# Runner module loader (scripts/ is not a package)
# ---------------------------------------------------------------------------
_spec = importlib.util.spec_from_file_location(
    "run_delegated_work_runner",
    _REPO_ROOT / "scripts" / "run_delegated_work_runner.py")
runner = importlib.util.module_from_spec(_spec)
# The runner's top-level load_dotenv(.env, override=True) must NOT leak the
# REAL .env (live tokens, live CORA_DELEGATED_WORK) into the pytest process
# for the rest of the session (D-051 test lens) -- snapshot + restore.
import os as _os  # noqa: E402

_env_before = dict(_os.environ)
_spec.loader.exec_module(runner)
_os.environ.clear()
_os.environ.update(_env_before)


# ---------------------------------------------------------------------------
# Fakes for the Anthropic client
# ---------------------------------------------------------------------------
def _usage(inp=100, out=50, cc=0, cr=0, searches=0, fetches=0):
    stu = SimpleNamespace(web_search_requests=searches, web_fetch_requests=fetches) \
        if (searches or fetches) else None
    return SimpleNamespace(input_tokens=inp, output_tokens=out,
                           cache_creation_input_tokens=cc,
                           cache_read_input_tokens=cr, server_tool_use=stu)


class _Block(SimpleNamespace):
    """SDK-content-block stand-in WITH model_dump, so the pause_turn
    serialization path (_serialize_blocks) is exercised for real."""

    def model_dump(self, exclude_none=False):
        d = dict(vars(self))
        if exclude_none:
            d = {k: v for k, v in d.items() if v is not None}
        return d


def _text_block(text):
    return _Block(type="text", text=text, citations=[])


def _tool_block(name, tool_input=None, tid="toolu_1"):
    return _Block(type="tool_use", name=name, input=tool_input or {}, id=tid)


def _resp(blocks, stop_reason="end_turn", usage=None):
    return SimpleNamespace(content=blocks, stop_reason=stop_reason,
                           usage=usage or _usage())


class FakeClient:
    def __init__(self, responses):
        self._responses = list(responses)
        self.calls: list[dict] = []
        self.messages = SimpleNamespace(create=self._create)

    def _create(self, **kwargs):
        self.calls.append(kwargs)
        if not self._responses:
            raise AssertionError("FakeClient ran out of scripted responses")
        return self._responses.pop(0)


def _job(entity="F3E", archetype="doc_draft", deliverable="md",
         requester="U_REQ", job_id="dw-abc123def456"):
    return {
        "job_id": job_id, "archetype": archetype,
        "title": "Test Job Title", "brief": "write me a test document brief",
        "requester": requester, "requester_name": "Testy",
        "entity": entity, "channel_id": "C_SRC", "channel_name": "f3e-leadership",
        "thread_ts": "111.222", "deliverable": deliverable,
    }


@pytest.fixture(autouse=True)
def _isolated(monkeypatch, tmp_path):
    monkeypatch.setattr(dw, "_BOT_LEDGER", tmp_path / "dw-bot.jsonl")
    monkeypatch.setattr(dw, "_RUNNER_LEDGER", tmp_path / "dw-runner.jsonl")
    monkeypatch.setattr(dw, "_STAGING_ROOT", tmp_path / "staging")
    monkeypatch.setattr(runner, "LOCK_PATH", tmp_path / "dw.lock")
    monkeypatch.setenv("CORA_DELEGATED_WORK", "live")
    monkeypatch.setenv("FOUNDER_OS_ROOT", str(tmp_path / "fos"))
    yield


def _seed_job(state_events=("queued",), job_id="dw-abc123def456", age_hours=0.0,
              **overrides):
    job = _job(job_id=job_id)
    job.update(overrides)
    ts = (datetime.now(timezone.utc) - timedelta(hours=age_hours)).isoformat()
    dw.append_bot_event({"event": "requested", "ts": ts, **job,
                         "fingerprint": f"fp-{job_id}"})
    for ev in state_events:
        if ev in dw.BOT_EVENTS:
            dw.append_bot_event({"event": ev, "ts": ts, "job_id": job_id})
        else:
            dw.append_runner_event({"event": ev, "ts": ts, "job_id": job_id})
    return job


# ---------------------------------------------------------------------------
# llm_rates
# ---------------------------------------------------------------------------

def test_estimate_usd_four_term_formula_plus_search():
    # 1M input + 1M cache_create + 1M cache_read + 1M output + 12 searches
    usd = llm_rates.estimate_usd(1_000_000, 1_000_000, 1_000_000, 1_000_000, 12)
    assert usd == pytest.approx(3.0 + 3.75 + 0.30 + 15.0 + 0.12)


def test_estimate_usd_search_surcharge_alone():
    assert llm_rates.estimate_usd(0, 0, 0, 0, 5) == pytest.approx(0.05)


# ---------------------------------------------------------------------------
# Archetypes: allowlists + xlsx spec
# ---------------------------------------------------------------------------

def test_allowlists_carry_zero_write_tools():
    """Structural invariant (design 8.2): no staged-write tool (anything with a
    `confirmed` schema property) and no known write/side-effect tool may appear
    in any archetype allowlist."""
    from cora.tools import tool_dispatch as td
    staged_write = {
        t["name"] for t in td.TOOL_DEFINITIONS
        if "confirmed" in (t.get("input_schema", {}).get("properties") or {})
    }
    assert staged_write  # sanity: the probe itself works
    known_writes = {
        "slack_send_dm", "gmail_create_draft", "hubspot_update_deal_stage",
        "hubspot_add_note", "influencer_add_handle", "influencer_log_deliverable",
        "cora_remember", "cora_forget_note", "cora_queue_code_session",
        "cora_delegate_work", "f3e_shopify_set_inventory",
    }
    for name, spec in arch.ARCHETYPE_SPECS.items():
        bad = spec["allowlist"] & (staged_write | known_writes)
        assert not bad, f"{name} allowlist carries write tools: {bad}"


def test_kb_search_never_in_global_catalog():
    from cora.tools import tool_dispatch as td
    assert "kb_search" not in [t["name"] for t in td.TOOL_DEFINITIONS]
    assert "kb_search" not in td._TOOL_FUNCTIONS


def test_validate_table_spec_accepts_valid():
    spec = {"sheets": [{"name": "Data", "columns": ["a", "b"],
                        "rows": [["x", 1], ["y", 2.5]]}]}
    assert arch.validate_table_spec(spec) is None


@pytest.mark.parametrize("spec,fragment", [
    ({"sheets": []}, "non-empty"),
    ({"sheets": [{"name": "S", "columns": ["a"], "rows": [["x", "extra"]]}]}, "cells"),
    ({"sheets": [{"name": "S", "columns": ["a"],
                  "rows": [[{"nested": 1}]]}]}, "non-scalar"),
    ({"sheets": [{"name": "S", "columns": ["a"], "rows": []},
                 {"name": "s", "columns": ["a"], "rows": []}]}, "duplicate"),
    ({"sheets": [{"name": "bad[name]", "columns": ["a"], "rows": []}]}, "invalid"),
    ("not a dict", "not an object"),
])
def test_validate_table_spec_rejects(spec, fragment):
    err = arch.validate_table_spec(spec)
    assert err and fragment in err


def test_validate_table_spec_row_cap():
    spec = {"sheets": [{"name": "Big", "columns": ["a"],
                        "rows": [["x"]] * (arch.MAX_TOTAL_ROWS + 1)}]}
    err = arch.validate_table_spec(spec)
    assert err and "too many rows" in err


def test_build_xlsx_roundtrip():
    from openpyxl import load_workbook
    spec = {"sheets": [{"name": "Q2", "columns": ["Vendor", "USD"],
                        "rows": [["Acme", 1200], ["Globex", 99.5]]}]}
    data = arch.build_xlsx_bytes(spec)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Q2"]
    assert [c.value for c in ws[1]] == ["Vendor", "USD"]
    assert [c.value for c in ws[2]] == ["Acme", 1200]
    assert [c.value for c in ws[3]] == ["Globex", 99.5]


def test_extract_table_spec_takes_last_json_fence():
    text = ('junk ```json\n{"not": "spec"}\n``` more '
            '```json\n{"sheets": [{"name": "S", "columns": ["a"], "rows": []}]}\n```')
    spec = arch.extract_table_spec(text)
    assert spec and spec["sheets"][0]["name"] == "S"


def test_split_summary_artifact_marker_and_fallback():
    s, a = arch.split_summary_artifact(
        "## Summary\nshort\n" + arch.ARTIFACT_MARKER + "\nbody here")
    assert s == "short" and a == "body here"
    s2, a2 = arch.split_summary_artifact("no marker at all")
    assert a2 == "no marker at all" and s2 == "no marker at all"


def test_markdown_artifact_carries_provenance_and_banner():
    out = arch.assemble_markdown(_job(), "The body.")
    assert "AI-GENERATED delegated-work output" in out
    assert "NOT canon" in out
    assert "dw-abc123def456" in out
    assert out.rstrip().endswith("The body.")


# ---------------------------------------------------------------------------
# Worker: run_job two-phase loop
# ---------------------------------------------------------------------------

def _run(job, responses, **kwargs):
    client = FakeClient(responses)
    kwargs.setdefault("kb_search_fn", lambda q, l=None: "no results")
    kwargs.setdefault("dispatch_fn", MagicMock(return_value="dispatched"))
    with patch.object(worker, "web_withheld_reason", return_value="screen:test"):
        out = worker.run_job(job, anthropic_client=client, **kwargs)
    return out, client


def test_run_job_md_happy_path():
    text = "## Summary\nAll done.\n" + arch.ARTIFACT_MARKER + "\n# Draft\ncontent"
    out, client = _run(_job(), [_resp([_text_block(text)])])
    assert out["ok"] is True
    assert out["summary"] == "All done."
    assert "# Draft" in out["artifact_text"]
    assert "AI-GENERATED" in out["artifact_text"]
    assert out["artifact_ext"] == "md"
    assert out["partial"] is False
    assert out["cost"]["turns"] == 1


def test_run_job_prompt_caching_applied_every_create():
    """cache_control is MANDATORY (design section 5): system block, tools last
    block, and the moving conversation prefix."""
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    out, client = _run(_job(), [
        _resp([_tool_block("kb_search", {"query": "q"})], stop_reason="tool_use"),
        _resp([_text_block(text)]),
    ])
    assert out["ok"]
    for call in client.calls:
        assert call["system"][0]["cache_control"] == {"type": "ephemeral"}
        assert call["tools"][-1].get("cache_control") == {"type": "ephemeral"}
        # The LAST user-role message carries the moving conversation breakpoint.
        last_user = [m for m in call["messages"] if m.get("role") == "user"][-1]
        content = last_user["content"]
        assert isinstance(content, list)
        assert content[-1].get("cache_control") == {"type": "ephemeral"}
        assert call["thinking"] == {"type": "disabled"}


def test_run_job_kb_search_routed_locally_and_dispatch_refused_offlist():
    kb_fn = MagicMock(return_value="kb says hi")
    dispatch_fn = MagicMock(return_value="dispatched")
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    out, client = _run(
        _job(archetype="doc_draft"),  # allowlist = {kb_search} only
        [
            _resp([_tool_block("kb_search", {"query": "vendors"}, "t1"),
                   _tool_block("hubspot_get_my_deals", {}, "t2")],
                  stop_reason="tool_use"),
            _resp([_text_block(text)]),
        ],
        kb_search_fn=kb_fn, dispatch_fn=dispatch_fn)
    assert out["ok"]
    kb_fn.assert_called_once_with("vendors", None)
    dispatch_fn.assert_not_called()  # off-list name hard-refused
    tool_results = client.calls[1]["messages"][-1]["content"]
    contents = {r["tool_use_id"]: r["content"] for r in tool_results
                if isinstance(r, dict) and r.get("type") == "tool_result"}
    assert contents["t1"] == "kb says hi"
    assert "not permitted" in contents["t2"]


def test_run_job_allowed_dispatch_tool_runs_with_requester_identity():
    dispatch_fn = MagicMock(return_value="deal data")
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    job = _job(archetype="research_brief")  # allowlist includes hubspot reads
    out, client = _run(
        job,
        [
            _resp([_tool_block("hubspot_get_my_deals", {"x": 1}, "t1")],
                  stop_reason="tool_use"),
            _resp([_text_block(text)]),
        ],
        dispatch_fn=dispatch_fn)
    assert out["ok"]
    dispatch_fn.assert_called_once_with(
        "hubspot_get_my_deals", {"x": 1}, "U_REQ", "F3E", "f3e-leadership", "C_SRC")


def test_run_job_cost_cap_aborts_with_partial_delivery(monkeypatch):
    monkeypatch.setenv("CORA_DELEGATED_JOB_USD", "0.01")
    # One tool-use turn with narration text + big usage -> cost cap trips at
    # the loop top; the narration is still delivered as a partial artifact.
    out, client = _run(_job(), [
        _resp([_text_block("Partial narration so far."),
               _tool_block("kb_search", {"query": "q"})],
              stop_reason="tool_use", usage=_usage(inp=10_000, out=2_000)),
    ])
    assert out["ok"] is True
    assert out["partial"] is True
    assert out["partial_reason"] == "cost_cap"
    assert "Partial narration" in out["artifact_text"]
    assert len(client.calls) == 1  # no second create after the cap


def test_run_job_turn_cap_bounds_the_loop():
    responses = [
        _resp([_tool_block("kb_search", {"query": f"q{i}"})], stop_reason="tool_use")
        for i in range(worker.MAX_JOB_TURNS + 5)
    ]
    out, client = _run(_job(), responses)
    assert len(client.calls) == worker.MAX_JOB_TURNS
    assert out["cost"]["turns"] == worker.MAX_JOB_TURNS


def test_run_job_no_output_fails_honestly():
    out, _client = _run(_job(), [_resp([_text_block("")])])
    assert out["ok"] is False
    assert out["failure_class"] == "no_output"


def test_run_job_xlsx_valid_spec_builds_workbook():
    spec_json = json.dumps({"sheets": [{"name": "S", "columns": ["a"],
                                        "rows": [["v"]]}]})
    text = ("## Summary\nsheet done\n" + arch.ARTIFACT_MARKER
            + "\n```json\n" + spec_json + "\n```")
    out, _client = _run(_job(archetype="spreadsheet_build", deliverable="xlsx"),
                        [_resp([_text_block(text)])])
    assert out["ok"] and out["artifact_ext"] == "xlsx"
    assert out["artifact_bytes"]
    assert "v" in out["xlsx_cell_text"]


def test_run_job_xlsx_malformed_spec_falls_back_to_md_honestly():
    text = ("## Summary\nsheet done\n" + arch.ARTIFACT_MARKER
            + "\n```json\n{\"sheets\": [{\"name\": \"S\"}]}\n```")
    out, _client = _run(_job(archetype="spreadsheet_build", deliverable="xlsx"),
                        [_resp([_text_block(text)])])
    assert out["ok"]
    assert out["artifact_ext"] == "md"  # honest fallback, never a broken file
    assert "invalid" in out["summary"]


def test_run_job_lex_belt():
    # Bind the BELT, not just "failed somehow" (D-051 test lens): the refusal
    # must fire BEFORE any model call, with the LEX-specific classification.
    client = FakeClient([])
    out = worker.run_job(_job(entity="LEX-LLC"), anthropic_client=client)
    assert out["ok"] is False
    assert out["failure_class"] == "error"
    assert "LEX" in out["message"]
    assert client.calls == []  # no model call ever fired for a LEX job


# ---------------------------------------------------------------------------
# Phase A: web budget clamp (incl. pause_turn) + withheld reasons
# ---------------------------------------------------------------------------

def test_phase_a_pause_turn_reclamps_max_uses():
    """max_uses is per REQUEST: after 11 searches the continuation create must
    clamp web_search max_uses to 1 (min(env, remaining)); at 0 the tool drops."""
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    client = FakeClient([
        _resp([_text_block("searching...")], stop_reason="pause_turn",
              usage=_usage(searches=worker.MAX_WEB_SEARCHES - 1, fetches=0)),
        _resp([_text_block("found stuff")], stop_reason="end_turn"),
        _resp([_text_block(text)]),  # phase B
    ])
    with patch.object(worker, "web_withheld_reason", return_value=None), \
         patch.object(worker.web_guard, "record_usage") as rec:
        out = worker.run_job(_job(archetype="research_brief"),
                             anthropic_client=client,
                             kb_search_fn=lambda q, l=None: "x",
                             dispatch_fn=MagicMock(return_value="y"))
    assert out["ok"]
    # Call 2 is the pause_turn continuation: web_search clamped to remaining=1.
    tools_second = client.calls[1]["tools"]
    ws = [t for t in tools_second if t.get("name") == "web_search"]
    assert ws and ws[0]["max_uses"] == 1
    # Usage recorded into the shared ledger under the jobs-lane channel.
    rec.assert_called_once()
    assert rec.call_args.kwargs.get("channel_name") == worker.JOBS_LANE_CHANNEL


def test_phase_a_search_budget_exhausted_drops_tool():
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    client = FakeClient([
        _resp([_text_block("burned the budget")], stop_reason="pause_turn",
              usage=_usage(searches=worker.MAX_WEB_SEARCHES,
                           fetches=worker.MAX_WEB_FETCHES)),
        _resp([_text_block(text)]),  # phase B (phase A loop exits: no tools left)
    ])
    with patch.object(worker, "web_withheld_reason", return_value=None), \
         patch.object(worker.web_guard, "record_usage"):
        out = worker.run_job(_job(archetype="research_brief"),
                             anthropic_client=client,
                             kb_search_fn=lambda q, l=None: "x",
                             dispatch_fn=MagicMock(return_value="y"))
    assert out["ok"]
    assert len(client.calls) == 2  # NO third create with empty/invalid web tools


def test_web_withheld_screen_trip(monkeypatch):
    monkeypatch.setattr(worker.web_guard, "_screen_query", lambda q, **kw: "phi")
    assert worker.web_withheld_reason("brief") == "screen:phi"


def test_web_withheld_org_cap(monkeypatch):
    monkeypatch.setattr(worker.web_guard, "_screen_query", lambda q, **kw: None)
    monkeypatch.setattr(worker.web_guard, "searches_today", lambda: 40)
    monkeypatch.setattr(worker.web_guard, "daily_cap", lambda: 40)
    assert worker.web_withheld_reason("brief") == "org_daily_cap"


def test_web_withheld_jobs_lane_cap(monkeypatch):
    monkeypatch.setattr(worker.web_guard, "_screen_query", lambda q, **kw: None)
    monkeypatch.setattr(worker.web_guard, "searches_today", lambda: 0)
    monkeypatch.setattr(worker, "jobs_lane_searches_today",
                        lambda: worker.JOBS_LANE_DAILY_SEARCH_CAP)
    assert worker.web_withheld_reason("brief") == "jobs_lane_cap"


def test_jobs_lane_counter_reads_only_jobs_rows(tmp_path, monkeypatch):
    import cora.web_guard as wg
    monkeypatch.setattr(wg, "_USAGE_LEDGER", tmp_path / "web-usage.jsonl")
    wg.record_usage(3, 1, "F3E", channel_name=worker.JOBS_LANE_CHANNEL)
    wg.record_usage(5, 0, "F3E", channel_name="f3e-leadership")  # interactive
    assert worker.jobs_lane_searches_today() == 3
    assert wg.searches_today() == 8  # org-wide stays ONE number


# ---------------------------------------------------------------------------
# kb_search pinning (structural)
# ---------------------------------------------------------------------------

def test_kb_search_pins_entity_and_requester_scope(monkeypatch):
    import cora.context_loader as cl
    import cora.historical_access as ha

    fake_kb = MagicMock()
    chunk = SimpleNamespace(distance=0.5, source="static_md", content="c",
                            title="t", deep_link="", chunk_id="1", entity="OSN")
    fake_kb.search.return_value = [chunk]
    monkeypatch.setattr(worker, "_get_ro_kb", lambda: fake_kb)
    monkeypatch.setattr(worker, "_RO_KB", None)
    from cora.knowledge_base import embeddings
    monkeypatch.setattr(embeddings, "embed_query", lambda q: [0.0] * 8)

    seen = {}

    def fake_tier1(results, emails, unrestricted):
        seen["emails"] = emails
        seen["unrestricted"] = unrestricted
        return results, False

    monkeypatch.setattr(ha, "apply_tier1", fake_tier1)
    monkeypatch.setattr(ha, "owned_emails", lambda uid: frozenset({f"{uid}@x.com"}))
    lex_scrub = MagicMock(side_effect=lambda r: r)
    nonlex_scrub = MagicMock(side_effect=lambda r: r)
    monkeypatch.setattr(cl, "_apply_lex_phi_scrub", lex_scrub)
    monkeypatch.setattr(cl, "_withhold_non_lex_phi", nonlex_scrub)
    monkeypatch.setattr(cl, "_format_kb_chunks", lambda r: "formatted chunks")

    # OSNGW: a store code -- must collapse to OSN for the search.
    fn = worker.make_kb_search(_job(entity="OSNGW", requester="U_REQ"))
    out = fn("test query")
    assert "formatted chunks" in out
    kwargs = fake_kb.search.call_args.kwargs
    assert kwargs["entity"] == "OSN"          # parent-collapsed, computed in code
    assert kwargs["sub_entity"] is None
    assert kwargs["include_fndr"] is True
    assert seen["emails"] == frozenset({"U_REQ@x.com"})  # owned_emails(requester)
    assert seen["unrestricted"] is False       # never the founder surface
    nonlex_scrub.assert_called_once()          # non-custodian scrub ALWAYS runs
    lex_scrub.assert_not_called()


def test_kb_search_capped_at_eight_calls(monkeypatch):
    monkeypatch.setattr(worker, "_get_ro_kb", lambda: None)
    fn = worker.make_kb_search(_job())
    for _ in range(worker.MAX_KB_CALLS):
        fn("query")  # returns "unavailable" (kb None) but consumes the budget
    out = fn("query")
    assert "budget exhausted" in out


def test_kb_search_schema_has_no_entity_parameter():
    props = worker.KB_SEARCH_DEF["input_schema"]["properties"]
    assert set(props) == {"query", "limit"}  # a model-supplied entity cannot exist


# ---------------------------------------------------------------------------
# Artifact guard + pathing
# ---------------------------------------------------------------------------

def test_guard_artifact_trip_returns_content_guard(monkeypatch):
    import cora.channel_content_guard as ccg
    monkeypatch.setattr(ccg, "guard_outbound",
                        lambda text, **k: ("refusal text", True))
    fclass, text = worker.guard_artifact_text(_job(), "company revenue $320,615")
    assert fclass == "content_guard"
    assert text == "refusal text"


def test_guard_artifact_pass_and_phi_backstop(monkeypatch):
    import cora.channel_content_guard as ccg
    import cora.phi_guard as pg
    monkeypatch.setattr(ccg, "guard_outbound", lambda text, **k: (text, False))
    monkeypatch.setattr(pg, "non_lex_phi_backstop_trips_live",
                        lambda text, allowed_names=None: False)
    fclass, text = worker.guard_artifact_text(_job(), "clean body")
    assert fclass is None and text == "clean body"
    monkeypatch.setattr(pg, "non_lex_phi_backstop_trips_live",
                        lambda text, allowed_names=None: True)
    fclass, _ = worker.guard_artifact_text(_job(), "Marcus's service hours...")
    assert fclass == "content_guard"


def test_guard_artifact_error_fails_closed(monkeypatch):
    import cora.channel_content_guard as ccg

    def _boom(text, **k):
        raise RuntimeError("guard exploded")

    monkeypatch.setattr(ccg, "guard_outbound", _boom)
    fclass, _ = worker.guard_artifact_text(_job(), "anything")
    assert fclass == "content_guard"


def test_resolve_delivery_tier_fails_most_restrictive(monkeypatch):
    import cora.channel_classifier as cc

    def _boom(name):
        raise RuntimeError("routing unavailable")

    monkeypatch.setattr(cc, "classify_function", _boom)
    assert worker.resolve_delivery_tier(_job()) == "TIER_3"


def test_artifact_path_parent_collapse_and_dwid6(monkeypatch, tmp_path):
    monkeypatch.setenv("FOUNDER_OS_ROOT", str(tmp_path))
    job = _job(entity="OSNGW", job_id="dw-9f8e7d6c5b4a")
    path = worker.artifact_target_path(job)
    s = str(path)
    assert "09-One-Stop-Nutrition" in s      # OSNGW -> OSN -> its folder,
    assert "00-Founder" not in s             # NEVER the founder catch-all
    assert "_delegated-work" in s
    assert path.name.endswith("-9f8e7d.md")  # dwid6 suffix (collision-proof)
    assert path.name.startswith(datetime.now(worker._AZ_TZ).strftime("%Y-%m-%d"))
    assert "osngw" in path.name              # the job's own entity code in the name


def test_artifact_filename_slug_is_sanitized():
    job = _job(job_id="dw-aaaabbbbcccc")
    job["title"] = 'Weird/Title: with "chars" & spaces!'
    name = worker.artifact_filename(job)
    assert "/" not in name and '"' not in name and ":" not in name.split("_", 1)[1]


# ---------------------------------------------------------------------------
# Runner: lockfile
# ---------------------------------------------------------------------------

def test_lock_fresh_holder_blocks():
    runner.LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    runner.LOCK_PATH.write_text(json.dumps({"pid": 999999, "ts": time.time()}),
                                encoding="utf-8")
    assert runner.acquire_lock() is False


def test_lock_stale_dead_pid_overrides(monkeypatch):
    runner.LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    runner.LOCK_PATH.write_text(
        json.dumps({"pid": 999999, "ts": time.time() - 3600}), encoding="utf-8")
    monkeypatch.setattr(runner, "_pid_alive", lambda pid: False)
    assert runner.acquire_lock() is True


def test_lock_stale_but_alive_pid_fails_closed(monkeypatch):
    runner.LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    runner.LOCK_PATH.write_text(
        json.dumps({"pid": 12345, "ts": time.time() - 3600}), encoding="utf-8")
    monkeypatch.setattr(runner, "_pid_alive", lambda pid: True)
    assert runner.acquire_lock() is False


def test_lock_unreadable_fails_closed():
    runner.LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    runner.LOCK_PATH.write_text("{torn json", encoding="utf-8")
    assert runner.acquire_lock() is False


# ---------------------------------------------------------------------------
# Runner: flag semantics + claim path
# ---------------------------------------------------------------------------

def test_runner_at_off_claims_nothing(monkeypatch):
    monkeypatch.setenv("CORA_DELEGATED_WORK", "off")
    _seed_job()
    n = runner.run_pass(1, dry_run=False)
    assert n == 0
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_QUEUED


def test_runner_log_mode_simulates_without_model_calls(monkeypatch):
    monkeypatch.setenv("CORA_DELEGATED_WORK", "log")
    monkeypatch.setattr(runner, "_client", lambda: None)
    monkeypatch.setattr(worker, "run_job",
                        MagicMock(side_effect=AssertionError("must not run")))
    _seed_job()
    n = runner.run_pass(1, dry_run=False)
    assert n == 1
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_SIMULATED


def test_runner_live_runs_and_delivers(monkeypatch, tmp_path):
    monkeypatch.setattr(runner, "_client", lambda: None)
    outcome = {"ok": True, "summary": "done",
               "artifact_text": "# body", "artifact_bytes": None,
               "artifact_ext": "md", "partial": False, "partial_reason": None,
               "web_withheld_reason": None, "cost": {"est_usd": 0.4}}
    monkeypatch.setattr(runner.worker, "run_job", MagicMock(return_value=outcome))
    monkeypatch.setattr(runner.worker, "guard_artifact_text",
                        lambda job, text: (None, text))
    written = {}
    monkeypatch.setattr(runner.drive_io, "write_text_atomic",
                        lambda path, text, **k: written.setdefault("path", str(path)))
    monkeypatch.setattr(runner, "post_threaded", lambda c, j, t: True)
    _seed_job()
    n = runner.run_pass(1, dry_run=False)
    assert n == 1
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_DELIVERED
    assert rec["artifact"]["mis_homed"] is False
    assert "_delegated-work" in written["path"]


def test_runner_envelope_recheck_skips_unreleased(monkeypatch):
    monkeypatch.setenv("CORA_DELEGATED_WORK", "log")
    monkeypatch.setattr(runner, "_client", lambda: None)
    monkeypatch.setattr(runner.dw, "envelope_headroom", lambda *a, **k: -1.0)
    _seed_job()
    n = runner.run_pass(1, dry_run=False)
    assert n == 0
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_QUEUED


def test_runner_envelope_recheck_honors_release(monkeypatch):
    monkeypatch.setenv("CORA_DELEGATED_WORK", "log")
    monkeypatch.setattr(runner, "_client", lambda: None)
    monkeypatch.setattr(runner.dw, "envelope_headroom", lambda *a, **k: -1.0)
    _seed_job(state_events=("held", "released"))
    n = runner.run_pass(1, dry_run=False)
    assert n == 1  # Harrison's release is an explicit override


# ---------------------------------------------------------------------------
# Runner: delivery guard trip + cancel suppression + mis-homed
# ---------------------------------------------------------------------------

def _ok_outcome(**over):
    out = {"ok": True, "summary": "sum", "artifact_text": "# body",
           "artifact_bytes": None, "artifact_ext": "md", "partial": False,
           "partial_reason": None, "web_withheld_reason": None,
           "cost": {"est_usd": 0.5}}
    out.update(over)
    return out


def test_deliver_content_guard_trip_fails_job(monkeypatch):
    _seed_job(state_events=("queued", "started"))
    monkeypatch.setattr(runner.worker, "guard_artifact_text",
                        lambda job, text: ("content_guard", "guard refusal"))
    posts = []
    monkeypatch.setattr(runner, "post_threaded",
                        lambda c, j, t: posts.append(t) or True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: posts.append(t))
    runner.deliver(None, "dw-abc123def456", _ok_outcome())
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_FAILED
    assert rec["failure"]["class"] == "content_guard"
    # FAIL line carries the CLASS enum only.
    fail_lines = [p for p in posts if p.startswith("DW FAIL")]
    assert fail_lines == ["DW FAIL dw-abc123def456 content_guard"]


def test_deliver_suppressed_when_cancelled_mid_run(monkeypatch):
    _seed_job(state_events=("queued", "started"))
    dw.append_bot_event({"event": "cancelled", "ts": dw._now_iso(),
                         "job_id": "dw-abc123def456", "reason": "requester_cancel"})
    posted = MagicMock()
    monkeypatch.setattr(runner, "post_threaded", posted)
    runner.deliver(None, "dw-abc123def456", _ok_outcome())
    posted.assert_not_called()
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_CANCELLED


def test_deliver_drive_down_marks_mis_homed_and_still_delivers(monkeypatch):
    _seed_job(state_events=("queued", "started"))
    monkeypatch.setattr(runner.worker, "guard_artifact_text",
                        lambda job, text: (None, text))

    def _down(path, text, **k):
        raise runner.drive_io.DriveUnavailable("G: gone")

    monkeypatch.setattr(runner.drive_io, "write_text_atomic", _down)
    posts = []
    monkeypatch.setattr(runner, "post_threaded",
                        lambda c, j, t: posts.append(t) or True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: None)
    runner.deliver(None, "dw-abc123def456", _ok_outcome())
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_DELIVERED
    assert rec["artifact"]["mis_homed"] is True
    assert any("staged locally" in p for p in posts)
    # Staging retained until homed.
    assert dw.staging_dir("dw-abc123def456").exists()


def test_mis_homed_retry_homes_without_reposting(monkeypatch):
    _seed_job(state_events=("queued", "started"))
    monkeypatch.setattr(runner.worker, "guard_artifact_text",
                        lambda job, text: (None, text))
    calls = {"n": 0}

    def _down_then_up(path, text, **k):
        calls["n"] += 1
        if calls["n"] == 1:
            raise runner.drive_io.DriveUnavailable("G: gone")

    monkeypatch.setattr(runner.drive_io, "write_text_atomic", _down_then_up)
    monkeypatch.setattr(runner, "post_threaded", lambda c, j, t: True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: None)
    runner.deliver(None, "dw-abc123def456", _ok_outcome())
    assert dw.get_job("dw-abc123def456")["artifact"]["mis_homed"] is True

    posted = MagicMock()
    monkeypatch.setattr(runner, "post_threaded", posted)
    runner.mis_homed_retry_pass(None)
    rec = dw.get_job("dw-abc123def456")
    assert rec["artifact"]["mis_homed"] is False
    posted.assert_not_called()  # artifact_homed NEVER re-posts
    assert not dw.staging_dir("dw-abc123def456").exists()  # cleaned once homed
    # Idempotency: a second pass finds nothing to do.
    runner.mis_homed_retry_pass(None)
    assert dw.get_job("dw-abc123def456")["artifact"]["mis_homed"] is False


# ---------------------------------------------------------------------------
# Runner: crash recovery (both arms) + expiry
# ---------------------------------------------------------------------------

def _age_event(job_id, event, hours, lane="runner", **extra):
    ts = (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat()
    row = {"event": event, "ts": ts, "job_id": job_id, **extra}
    if lane == "runner":
        dw.append_runner_event(row)
    else:
        dw.append_bot_event(row)


def test_crash_recovery_bare_running_fails_interrupted(monkeypatch):
    _seed_job(age_hours=2)
    _age_event("dw-abc123def456", "started", 1)  # 1h > 2x 10-min wall
    posts = []
    monkeypatch.setattr(runner, "post_threaded",
                        lambda c, j, t: posts.append(t) or True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: posts.append(t))
    runner.crash_recovery_pass(None)
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_FAILED
    assert rec["failure"]["class"] == "interrupted"
    assert any("interrupted" in p for p in posts)


def test_crash_recovery_delivering_marker_deliver_verifies(monkeypatch, tmp_path):
    _seed_job(age_hours=2)
    art_file = tmp_path / "artifact.md"
    art_file.write_text("body", encoding="utf-8")
    _age_event("dw-abc123def456", "started", 1)
    _age_event("dw-abc123def456", "delivering", 1,
               artifact={"local_path": str(art_file), "target_path": "",
                         "mis_homed": False},
               cost={"est_usd": 0.3})
    posts = []
    monkeypatch.setattr(runner, "post_threaded",
                        lambda c, j, t: posts.append(t) or True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: None)
    runner.crash_recovery_pass(None)
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_DELIVERED  # never a bare FAILED after delivery
    assert any("recovered" in p for p in posts)
    assert len(posts) == 1
    # Second pass: terminal now -- no double-post.
    runner.crash_recovery_pass(None)
    assert len(posts) == 1


def test_crash_recovery_ignores_fresh_running():
    _seed_job(state_events=("queued", "started"))
    runner.crash_recovery_pass(None)
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_RUNNING


def test_expiry_pass_expires_old_queued_never_held(monkeypatch):
    jid_q = "dw-queuedold1111"
    job = _job(job_id=jid_q)
    old = (datetime.now(timezone.utc) - timedelta(hours=49)).isoformat()
    dw.append_bot_event({"event": "requested", "ts": old, **job, "fingerprint": "f1"})
    dw.append_bot_event({"event": "queued", "ts": old, "job_id": jid_q})
    jid_h = "dw-heldold22222"
    job2 = _job(job_id=jid_h)
    job2["fingerprint"] = "f2"
    dw.append_bot_event({"event": "requested", "ts": old, **job2})
    dw.append_bot_event({"event": "held", "ts": old, "job_id": jid_h,
                         "reason": "user_quota"})
    monkeypatch.setattr(runner, "post_threaded", lambda c, j, t: True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: None)
    runner.expiry_pass(None)
    assert dw.get_job(jid_q)["state"] == dw.STATE_EXPIRED
    assert dw.get_job(jid_h)["state"] == dw.STATE_HELD  # HELD never expires


# ---------------------------------------------------------------------------
# Egress + import hygiene
# ---------------------------------------------------------------------------

def test_runner_posts_are_sanitized_source_pin():
    src = (_REPO_ROOT / "scripts" / "run_delegated_work_runner.py").read_text(
        encoding="utf-8")
    assert "sanitize_text" in src  # B1 doctrine + test_no_raw_slack_post rule
    assert "from cora.slack_egress import sanitize_text" in src


def test_worker_import_pulls_no_bot_process_modules():
    code = (
        "import sys; sys.path.insert(0, r'%s'); "
        "import cora.delegated_worker, cora.delegated_work, cora.delegated_archetypes; "
        "bad = [m for m in ('cora.app', 'cora.main', 'cora.tools.tool_dispatch', "
        "'cora.claude_client') if m in sys.modules]; "
        "assert not bad, f'bot-process modules imported: {bad}'"
    ) % str(_REPO_ROOT / "src")
    result = subprocess.run([sys.executable, "-c", code],
                            capture_output=True, text=True, timeout=120)
    assert result.returncode == 0, result.stderr


def test_ps1_is_ascii_only():
    raw = (_REPO_ROOT / "deployment" / "setup-delegated-work-task.ps1").read_bytes()
    assert all(b < 128 for b in raw), "PS1 must be ASCII-only (D-016)"
    text = raw.decode("ascii")
    assert ".venv\\Scripts\\python.exe" in text or ".venv\\\\Scripts" in text or ".venv" in text
    assert "cowork-cora-delegated-work" in text


# ---------------------------------------------------------------------------
# D-051 remediation bindings (2026-08-01 review)
# ---------------------------------------------------------------------------

def test_phase_a_continuation_carries_moving_cache_breakpoint():
    """The Phase-A pause_turn continuation must re-read the accumulated web
    content from CACHE: the serialized assistant tail block carries the moving
    breakpoint (D-051 cost lens, HIGH -- annotating only user turns parked the
    breakpoint on the tiny brief forever)."""
    text = "## Summary\ns\n" + arch.ARTIFACT_MARKER + "\nbody"
    client = FakeClient([
        _resp([_text_block("searching...")], stop_reason="pause_turn",
              usage=_usage(searches=2)),
        _resp([_text_block("found")], stop_reason="end_turn"),
        _resp([_text_block(text)]),  # phase B
    ])
    with patch.object(worker, "web_withheld_reason", return_value=None), \
         patch.object(worker.web_guard, "record_usage"):
        out = worker.run_job(_job(archetype="research_brief"),
                             anthropic_client=client,
                             kb_search_fn=lambda q, l=None: "x",
                             dispatch_fn=MagicMock(return_value="y"))
    assert out["ok"]
    cont = client.calls[1]["messages"]
    assert cont[-1]["role"] == "assistant"
    tail = cont[-1]["content"][-1]
    assert isinstance(tail, dict)  # serialized, not an SDK object
    assert tail.get("cache_control") == {"type": "ephemeral"}


def test_apply_conversation_cache_strips_stale_breakpoints():
    msgs = [
        {"role": "user", "content": [
            {"type": "text", "text": "a", "cache_control": {"type": "ephemeral"}}]},
        {"role": "assistant", "content": [
            {"type": "text", "text": "b", "cache_control": {"type": "ephemeral"}}]},
        {"role": "user", "content": [
            {"type": "tool_result", "tool_use_id": "t", "content": "c",
             "cache_control": {"type": "ephemeral"}}]},
    ]
    worker._apply_conversation_cache(msgs)
    marked = [b for m in msgs for b in m["content"] if "cache_control" in b]
    assert len(marked) == 1  # only the tail of the LAST message (max-4 budget)
    assert msgs[-1]["content"][-1]["cache_control"] == {"type": "ephemeral"}


def test_worker_timeout_is_not_retryable():
    import anthropic
    assert worker._is_retryable(anthropic.APITimeoutError(request=MagicMock())) is False


def test_all_allowlist_names_exist_in_tool_definitions():
    """A renamed tool must not silently empty an allowlist via the
    tools_for_entity intersection."""
    from cora.tools import tool_dispatch as td
    catalog = {t["name"] for t in td.TOOL_DEFINITIONS}
    for name, spec in arch.ARCHETYPE_SPECS.items():
        missing = (spec["allowlist"] - {"kb_search"}) - catalog
        assert not missing, f"{name} allowlist names not in TOOL_DEFINITIONS: {missing}"


def test_pid_alive_real_semantics():
    """The Windows probe must be REAL (os.kill(pid,0) is not a liveness check
    on this OS -- D-051 state lens HIGH): our own pid reads alive; a spawned-
    then-exited child's pid reads dead."""
    import os as os_mod
    assert runner._pid_alive(os_mod.getpid()) is True
    child = subprocess.run([sys.executable, "-c", "import os; print(os.getpid())"],
                           capture_output=True, text=True, timeout=60)
    dead_pid = int(child.stdout.strip())
    assert runner._pid_alive(dead_pid) is False
    assert runner._pid_alive(0) is False
    assert runner._pid_alive("garbage") is False


def test_release_lock_never_unlinks_another_pids_lock():
    runner.LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    runner.LOCK_PATH.write_text(json.dumps({"pid": 424242, "ts": time.time()}),
                                encoding="utf-8")
    runner.release_lock()
    assert runner.LOCK_PATH.exists()  # not ours -- left alone
    runner.refresh_lock()  # now ours
    runner.release_lock()
    assert not runner.LOCK_PATH.exists()


def test_requested_orphan_expires_and_frees_reservation():
    """A crash between the `requested` and `queued` appends leaves a REQUESTED
    orphan: never claimable, but it held an envelope reservation + a dedup
    block forever until the reaper (D-051 state lens)."""
    jid = "dw-orphan1111111"
    old = (datetime.now(timezone.utc) - timedelta(hours=49)).isoformat()
    dw.append_bot_event({"event": "requested", "ts": old, "job_id": jid,
                         **{k: v for k, v in _job(job_id=jid).items() if k != "job_id"},
                         "fingerprint": "fp-orphan"})
    assert dw.get_job(jid)["state"] == "REQUESTED"
    before = dw.open_job_count()
    monkey_posts = MagicMock(return_value=True)
    with patch.object(runner, "post_threaded", monkey_posts), \
         patch.object(runner, "post_sessions_line", lambda c, t: None):
        runner.expiry_pass(None)
    assert dw.get_job(jid)["state"] == dw.STATE_EXPIRED
    assert dw.open_job_count() == before - 1


def test_deliver_verify_failed_repost_leaves_marker(monkeypatch, tmp_path):
    """Deliver-verify must NOT append `delivered` when its recovery re-post
    fails -- that would terminally never-notify the requester."""
    _seed_job(age_hours=2)
    art_file = tmp_path / "artifact.md"
    art_file.write_text("body", encoding="utf-8")
    _age_event("dw-abc123def456", "started", 1)
    _age_event("dw-abc123def456", "delivering", 1,
               artifact={"local_path": str(art_file), "target_path": "",
                         "mis_homed": False},
               cost={"est_usd": 0.3})
    monkeypatch.setattr(runner, "post_threaded", lambda c, j, t: False)  # Slack down
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: None)
    runner.crash_recovery_pass(None)
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_RUNNING  # marker left for a later pass
    # Slack recovers -> the next pass completes the delivery.
    monkeypatch.setattr(runner, "post_threaded", lambda c, j, t: True)
    runner.crash_recovery_pass(None)
    assert dw.get_job("dw-abc123def456")["state"] == dw.STATE_DELIVERED


def test_crash_recovery_delivering_but_artifact_missing_fails(monkeypatch):
    _seed_job(age_hours=2)
    _age_event("dw-abc123def456", "started", 1)
    _age_event("dw-abc123def456", "delivering", 1,
               artifact={"local_path": "Z:/nowhere/gone.md", "target_path": "",
                         "mis_homed": False},
               cost={"est_usd": 0.3})
    posts = []
    monkeypatch.setattr(runner, "post_threaded",
                        lambda c, j, t: posts.append(t) or True)
    monkeypatch.setattr(runner, "post_sessions_line", lambda c, t: posts.append(t))
    runner.crash_recovery_pass(None)
    rec = dw.get_job("dw-abc123def456")
    assert rec["state"] == dw.STATE_FAILED
    assert rec["failure"]["class"] == "interrupted"


def test_staging_sweep_never_purges_mis_homed_artifact(monkeypatch):
    _seed_job(age_hours=1000)
    jid = "dw-abc123def456"
    d = dw.staging_dir(jid)
    d.mkdir(parents=True)
    (d / "artifact.md").write_text("the only copy", encoding="utf-8")
    old = time.time() - 40 * 86400
    import os as os_mod
    os_mod.utime(d, (old, old))
    dw.append_runner_event({"event": "started", "ts": dw._now_iso(), "job_id": jid})
    dw.append_runner_event({"event": "delivering", "ts": dw._now_iso(), "job_id": jid,
                            "artifact": {"local_path": str(d / "artifact.md"),
                                         "target_path": "G:/x.md", "mis_homed": True},
                            "cost": {}})
    dw.append_runner_event({"event": "delivered", "ts": dw._now_iso(), "job_id": jid,
                            "cost": {}, "artifact": {"mis_homed": True}})
    runner.staging_sweep()
    assert (d / "artifact.md").exists()  # the only copy survives, however old


def test_overflow_digest_sends_once_with_buttons_and_flushes():
    with patch.object(dw, "user_daily_quota", return_value=0):
        _submit_via_dw(client_factory=lambda: None)  # HELD + card_held (no client)
    held = dw.held_jobs_awaiting_card()
    assert len(held) == 1
    client = MagicMock()
    client.conversations_open.return_value = {"channel": {"id": "D_H"}}
    runner.overflow_digest_pass(client)
    client.chat_postMessage.assert_called_once()
    blocks = client.chat_postMessage.call_args.kwargs["blocks"]
    action_ids = [e["action_id"] for b in blocks if b.get("type") == "actions"
                  for e in b["elements"]]
    assert dw.ACTION_RELEASE in action_ids and dw.ACTION_DISMISS in action_ids
    assert dw.held_jobs_awaiting_card() == []  # card_flushed rode the runner ledger
    client.chat_postMessage.reset_mock()
    runner.overflow_digest_pass(client)  # second pass: nothing left
    client.chat_postMessage.assert_not_called()


def _submit_via_dw(client_factory):
    import cora.org_roles as org_roles
    import cora.user_access as user_access
    import cora.sibling_guard as sibling_guard
    import cora.cross_entity_guard as cross_entity_guard
    import cora.phi_guard as phi_guard
    with patch.object(org_roles, "get_role",
                      lambda uid: SimpleNamespace(entity="F3E", external=False,
                                                  name="T")), \
         patch.object(user_access, "check_access", lambda *a, **k: None), \
         patch.object(sibling_guard, "check_redirect", lambda *a, **k: None), \
         patch.object(cross_entity_guard, "check_cross_entity", lambda *a, **k: None), \
         patch.object(phi_guard, "is_any_phi", lambda t: False):
        return dw.submit_job("U_REQ", "F3E", "C_SRC", "f3e-leadership", "1.2",
                             "doc_draft", "an overflow digest test brief long enough",
                             "md", client_factory=client_factory)


def test_dispatch_end_to_end_persists_channel_and_thread():
    """The channel_id/thread_ts plumbing must hold through the REAL dispatch()
    injection -- not just hand-fed tool inputs (D-051 test lens)."""
    from cora.tools import tool_dispatch as td
    import cora.org_roles as org_roles
    import cora.user_access as user_access
    import cora.sibling_guard as sibling_guard
    import cora.cross_entity_guard as cross_entity_guard
    import cora.phi_guard as phi_guard
    with patch.object(org_roles, "get_role",
                      lambda uid: SimpleNamespace(entity="F3E", external=False,
                                                  name="T")), \
         patch.object(user_access, "check_access", lambda *a, **k: None), \
         patch.object(sibling_guard, "check_redirect", lambda *a, **k: None), \
         patch.object(cross_entity_guard, "check_cross_entity", lambda *a, **k: None), \
         patch.object(phi_guard, "is_any_phi", lambda t: False):
        out = td.dispatch("cora_delegate_work",
                          {"action": "request", "archetype": "doc_draft",
                           "brief": "end to end plumbing brief long enough"},
                          "U_REQ", "F3E", "f3e-leadership", "C_REAL", "111.222")
        assert "WRITE_BLOCKED" in out
        out = td.dispatch("cora_delegate_work", {"confirmed": True},
                          "U_REQ", "F3E", "f3e-leadership", "C_REAL", "111.222")
        assert "WRITE_CONFIRMED" in out
    job = dw.load_jobs()[0]
    assert job["channel_id"] == "C_REAL"
    assert job["thread_ts"] == "111.222"


def test_runner_script_subprocess_pulls_no_bot_process():
    """D-047 for the runner SCRIPT itself (module-level imports incl. its
    load_dotenv) -- not just the worker modules."""
    script = _REPO_ROOT / "scripts" / "run_delegated_work_runner.py"
    code = (
        "import sys, importlib.util; "
        "sys.path.insert(0, r'%s'); "
        "spec = importlib.util.spec_from_file_location('rnr', r'%s'); "
        "m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); "
        "bad = [x for x in ('cora.app', 'cora.main') if x in sys.modules]; "
        "assert not bad, f'bot-process modules imported: {bad}'"
    ) % (str(_REPO_ROOT / "src"), str(script))
    result = subprocess.run([sys.executable, "-c", code],
                            capture_output=True, text=True, timeout=120)
    assert result.returncode == 0, result.stderr


# ---------------------------------------------------------------------------
# D-051 remediation (2026-08-06 multi-lens review) -- LEX lane
# ---------------------------------------------------------------------------

def test_f1_lex_job_refuses_rather_than_deliver_raw_phase_a_output(monkeypatch):
    """Phase-A output is model-summarized text built directly from untrusted
    fetched pages. Delivering it unmediated as a LEX artifact would persist it
    to the Lexington Drive tree and KB-ingest it, where hostile prose becomes
    retrievable LEX 'knowledge'. Phase B is what grounds it -- without B there
    is nothing worth persisting at LEX."""
    monkeypatch.setenv("CORA_DELEGATED_WORK_LEX", "on")
    job = _job(entity="LEX-LLC", archetype="research_brief")
    # Phase A yields findings; Phase B returns empty.
    client = FakeClient([
        _resp([_text_block("web findings from a fetched page")]),
        _resp([_text_block("")]),
    ])
    out = worker.run_job(job, anthropic_client=client)
    assert out["ok"] is False
    assert out["failure_class"] == "no_output"


def test_f1_non_lex_job_still_delivers_partial_phase_a(monkeypatch):
    """Invariant 4: the LEX-only refusal must not change non-LEX behaviour --
    a non-LEX job still gets its partial Phase-A delivery."""
    job = _job(entity="F3E", archetype="research_brief")
    client = FakeClient([
        _resp([_text_block("web findings from a fetched page")]),
        _resp([_text_block("")]),
    ])
    out = worker.run_job(job, anthropic_client=client)
    assert out["ok"] is True
    assert out["partial_reason"] == "phase_b_empty"


def test_f1_delegated_artifacts_are_pinned_gm_level_at_ingest():
    """The artifact carries no sub_entity, so the store would RE-DERIVE one from
    its own CONTENT -- letting model-written text choose which Lexington
    sub-entity channel it surfaces in. lex_gm_level opts it out of detection."""
    src = (_REPO_ROOT / "scripts" / "incremental_sync_static.py").read_text(
        encoding="utf-8")
    i = src.index("if is_delegated_work_path(path):")
    seg = src[i:i + 1400]
    assert 'metadata["bot_authored"] = True' in seg
    assert 'metadata["lex_gm_level"] = True' in seg

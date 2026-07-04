"""Unit tests for src/cora/connectors/drive_sweep.py.

Tests use mocks for Drive API, Anthropic client, and KB -- no real network calls.
"""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import MagicMock, patch, PropertyMock

import pytest

from cora.connectors.drive_sweep import (
    _chunk_text,
    _classify,
    _extract_google_doc,
    _extract_google_sheet,
    _extract_pdf_bytes,
    _extract_xlsx_bytes,
    _extract_content,
    _ingest_file,
    _PHI_PATTERNS,
    run_sweep,
    sweep_user,
)


# ── _chunk_text ────────────────────────────────────────────────────────────────

class TestChunkText:
    def test_short_text_single_chunk(self):
        text = "short text"
        chunks = _chunk_text(text)
        assert chunks == [text]

    def test_long_text_splits_correctly(self):
        text = "x" * 3000
        chunks = _chunk_text(text)
        assert len(chunks) > 1
        # Each chunk no longer than _CHUNK_SIZE
        for chunk in chunks:
            assert len(chunk) <= 1400

    def test_overlap_between_chunks(self):
        text = "a" * 1400 + "b" * 1400
        chunks = _chunk_text(text)
        assert len(chunks) >= 2
        # Overlap: end of chunk 0 should appear in start of chunk 1
        assert chunks[0][-50:] in chunks[1]

    def test_empty_string(self):
        assert _chunk_text("") == [""]

    def test_exactly_chunk_size(self):
        text = "x" * 1400
        assert _chunk_text(text) == [text]


# ── _PHI_PATTERNS ──────────────────────────────────────────────────────────────

class TestPhiPatterns:
    def test_matches_dob(self):
        assert _PHI_PATTERNS.search("patient dob: 1980-01-01")

    def test_matches_ahcccs(self):
        assert _PHI_PATTERNS.search("AHCCCS member enrolled")

    def test_matches_diagnosis(self):
        assert _PHI_PATTERNS.search("Diagnosis: autism spectrum disorder")

    def test_matches_icd10(self):
        assert _PHI_PATTERNS.search("ICD-10 code F84.0")

    def test_matches_npi(self):
        assert _PHI_PATTERNS.search("Provider NPI number")

    def test_no_match_for_normal_text(self):
        assert not _PHI_PATTERNS.search("quarterly revenue report Q1 2026")

    def test_no_match_for_f3_content(self):
        assert not _PHI_PATTERNS.search("F3 Energy brand guidelines Pure tagline")

    def test_case_insensitive(self):
        assert _PHI_PATTERNS.search("DOB")
        assert _PHI_PATTERNS.search("medicaid")
        assert _PHI_PATTERNS.search("MEDICAID")


# ── _classify ─────────────────────────────────────────────────────────────────

class TestClassify:
    def _make_client(self, response_json: dict) -> MagicMock:
        msg = MagicMock()
        msg.content = [MagicMock()]
        msg.content[0].text = json.dumps(response_json)
        client = MagicMock()
        client.messages.create.return_value = msg
        return client

    def test_returns_parsed_classification(self):
        client = self._make_client(
            {"score": 9, "entity": "F3E", "summary": "Q1 contract", "discard_reason": ""}
        )
        result = _classify(client, "contract.pdf", "Tommy", "tommy@f3energy.com", "F3E", "content")
        assert result["score"] == 9
        assert result["entity"] == "F3E"

    def test_strips_markdown_fences(self):
        msg = MagicMock()
        msg.content = [MagicMock()]
        msg.content[0].text = "```json\n{\"score\": 7, \"entity\": \"OSN\", \"summary\": \"x\", \"discard_reason\": \"\"}\n```"
        client = MagicMock()
        client.messages.create.return_value = msg
        result = _classify(client, "file.txt", "Matt", "matt@osn.com", "OSN", "preview")
        assert result["score"] == 7

    def test_falls_back_on_api_error(self):
        client = MagicMock()
        client.messages.create.side_effect = Exception("network error")
        result = _classify(client, "file.txt", "User", "user@hjrglobal.com", "HJRG", "preview")
        # Fallback: score=5 so the file is kept
        assert result["score"] == 5

    def test_falls_back_on_json_parse_error(self):
        msg = MagicMock()
        msg.content = [MagicMock()]
        msg.content[0].text = "not json at all"
        client = MagicMock()
        client.messages.create.return_value = msg
        result = _classify(client, "file.txt", "User", "user@hjrglobal.com", "HJRG", "preview")
        assert result["score"] == 5


# ── _extract_google_doc ────────────────────────────────────────────────────────

class TestExtractGoogleDoc:
    def test_returns_decoded_bytes(self):
        service = MagicMock()
        service.files().export().execute.return_value = b"Contract text here"
        result = _extract_google_doc(service, "file123")
        assert result == "Contract text here"

    def test_returns_string_directly(self):
        service = MagicMock()
        service.files().export().execute.return_value = "Already a string"
        result = _extract_google_doc(service, "file123")
        assert result == "Already a string"

    def test_returns_empty_on_error(self):
        service = MagicMock()
        service.files().export.side_effect = Exception("api error")
        result = _extract_google_doc(service, "file123")
        assert result == ""


# ── _extract_google_sheet ──────────────────────────────────────────────────────

class TestExtractGoogleSheet:
    def test_returns_tabular_text(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Q1"
        ws.append(["Revenue", "Cost", "Net"])
        ws.append([100000, 80000, 20000])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        service = MagicMock()
        service.files().export().execute.return_value = xlsx_bytes
        result = _extract_google_sheet(service, "sheet123")
        assert "Q1" in result
        assert "Revenue" in result

    def test_returns_empty_on_error(self):
        service = MagicMock()
        service.files().export.side_effect = Exception("api error")
        result = _extract_google_sheet(service, "sheet123")
        assert result == ""


# ── _extract_pdf_bytes ─────────────────────────────────────────────────────────

class TestExtractPdfBytes:
    def test_returns_empty_gracefully_when_pdfplumber_missing(self):
        with patch.dict("sys.modules", {"pdfplumber": None}):
            result = _extract_pdf_bytes(b"%PDF-1.4 fake")
            assert result == ""

    def test_extracts_text_with_pdfplumber(self):
        mock_page = MagicMock()
        mock_page.extract_text.return_value = "Page content here"
        mock_pdf = MagicMock()
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_pdf.pages = [mock_page]

        with patch("pdfplumber.open", return_value=mock_pdf):
            result = _extract_pdf_bytes(b"%PDF fake bytes")
        assert "Page content here" in result


# ── _extract_xlsx_bytes ────────────────────────────────────────────────────────

class TestExtractXlsxBytes:
    def test_returns_tabular_text(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Value"])
        ws.append(["Item A", 42])
        buf = io.BytesIO()
        wb.save(buf)
        result = _extract_xlsx_bytes(buf.getvalue())
        assert "Name" in result
        assert "Item A" in result

    def test_returns_empty_on_bad_bytes(self):
        result = _extract_xlsx_bytes(b"not an xlsx file")
        assert result == ""


# ── _extract_content routing ───────────────────────────────────────────────────

class TestExtractContent:
    def test_skips_image_mime(self):
        service = MagicMock()
        result = _extract_content(service, {"id": "x", "mimeType": "image/png"})
        assert result == ""

    def test_skips_folder_mime(self):
        service = MagicMock()
        result = _extract_content(
            service,
            {"id": "x", "mimeType": "application/vnd.google-apps.folder"}
        )
        assert result == ""

    def test_routes_google_doc(self):
        service = MagicMock()
        service.files().export().execute.return_value = b"doc text"
        result = _extract_content(
            service,
            {"id": "x", "mimeType": "application/vnd.google-apps.document"}
        )
        assert result == "doc text"

    def test_routes_google_sheet(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        buf = io.BytesIO()
        wb.save(buf)

        service = MagicMock()
        service.files().export().execute.return_value = buf.getvalue()
        result = _extract_content(
            service,
            {"id": "x", "mimeType": "application/vnd.google-apps.spreadsheet"}
        )
        assert result != ""


# ── _ingest_file ───────────────────────────────────────────────────────────────

class TestIngestFile:
    def _make_kb(self):
        kb = MagicMock()
        kb.upsert_documents = MagicMock(return_value=1)
        return kb

    def test_ingests_single_chunk(self):
        kb = self._make_kb()
        file_meta = {"id": "file1", "name": "contract.pdf", "mimeType": "application/pdf", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 9, "entity": "F3E", "summary": "Q1 distribution contract", "discard_reason": ""}
        user = {"email": "tommy@f3energy.com", "name": "Tommy Anderson", "entity_default": "F3E"}
        content = "This is the contract text"
        n = _ingest_file(kb, file_meta, content, classification, user)
        assert n == 1
        kb.upsert_documents.assert_called_once()
        docs = kb.upsert_documents.call_args[0][0]
        doc = docs[0]
        assert doc.source == "drive_sweep"
        assert doc.entity == "F3E"
        assert doc.source_id == "file1"

    def test_lex_sub_entity_split(self):
        kb = self._make_kb()
        file_meta = {"id": "file2", "name": "llc-ops.docx", "mimeType": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 8, "entity": "LEX-LLC", "summary": "LLC operations manual", "discard_reason": ""}
        user = {"email": "shaun@lexingtonservices.com", "name": "Shaun Hawkins", "entity_default": "LEX"}
        n = _ingest_file(kb, file_meta, "LLC SOP content", classification, user)
        docs = kb.upsert_documents.call_args[0][0]
        doc = docs[0]
        assert doc.entity == "LEX"
        assert doc.sub_entity == "LEX-LLC"

    def test_multiple_chunks_for_long_content(self):
        kb = self._make_kb()
        kb.upsert_documents = MagicMock(return_value=4)
        file_meta = {"id": "file3", "name": "big.txt", "mimeType": "text/plain", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 7, "entity": "HJRG", "summary": "Big document", "discard_reason": ""}
        user = {"email": "harrison@hjrglobal.com", "name": "Harrison", "entity_default": "FNDR"}
        long_content = "x" * 5000
        n = _ingest_file(kb, file_meta, long_content, classification, user)
        assert n == 4
        # Full content passed as single Document — KB handles chunking internally
        kb.upsert_documents.assert_called_once()
        docs = kb.upsert_documents.call_args[0][0]
        assert len(docs) == 1
        assert docs[0].content == long_content

    def test_drive_link_format(self):
        kb = self._make_kb()
        file_meta = {"id": "abc123", "name": "doc.txt", "mimeType": "text/plain", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 6, "entity": "FNDR", "summary": "A doc", "discard_reason": ""}
        user = {"email": "harrison@hjrglobal.com", "name": "Harrison", "entity_default": "FNDR"}
        _ingest_file(kb, file_meta, "content here", classification, user)
        docs = kb.upsert_documents.call_args[0][0]
        assert "drive.google.com/file/d/abc123" in docs[0].deep_link


# ── non-canonical entity guard (audit W6-05, entity-firewall strengthening) ─────

class TestNonCanonicalEntityGuard:
    """A Haiku mis-classification that returns an off-menu code (e.g. 'F3' from a
    filename token) must NOT mint a non-canonical entity that no channel routes to;
    it falls back to the file owner's canonical default. Reproduces the audited
    OSN Val Vista receipt that was tagged entity='F3'.
    """

    def _make_kb(self):
        kb = MagicMock()
        kb.upsert_documents = MagicMock(return_value=1)
        return kb

    def test_non_canonical_entity_falls_back_to_owner_default(self):
        kb = self._make_kb()
        file_meta = {"id": "f3receipt", "name": "376.2 F3.pdf", "mimeType": "application/pdf", "modifiedTime": "2024-11-06T00:00:00Z"}
        classification = {"score": 8, "entity": "F3", "summary": "Val Vista receipt", "discard_reason": ""}
        user = {"email": "matt@onestopnutrition.com", "name": "Matt", "entity_default": "OSN"}
        _ingest_file(kb, file_meta, "425 SOUTH VAL VISTA DRIVE ...", classification, user)
        doc = kb.upsert_documents.call_args[0][0][0]
        assert doc.entity == "OSN"          # not the bogus 'F3'
        assert doc.sub_entity is None

    def test_non_canonical_entity_and_non_canonical_default_falls_to_fndr(self):
        kb = self._make_kb()
        file_meta = {"id": "x", "name": "weird.pdf", "mimeType": "application/pdf", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 6, "entity": "F3", "summary": "?", "discard_reason": ""}
        user = {"email": "someone@hjrglobal.com", "name": "Someone", "entity_default": "BOGUS"}
        _ingest_file(kb, file_meta, "content", classification, user)
        assert kb.upsert_documents.call_args[0][0][0].entity == "FNDR"

    def test_canonical_parent_preserved(self):
        kb = self._make_kb()
        file_meta = {"id": "o1", "name": "osn.pdf", "mimeType": "application/pdf", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 8, "entity": "OSN", "summary": "OSN doc", "discard_reason": ""}
        user = {"email": "matt@onestopnutrition.com", "name": "Matt", "entity_default": "OSN"}
        _ingest_file(kb, file_meta, "content", classification, user)
        assert kb.upsert_documents.call_args[0][0][0].entity == "OSN"

    def test_canonical_sub_entity_preserved_through_guard(self):
        kb = self._make_kb()
        file_meta = {"id": "cl1", "name": "hjrp-cl-lease.pdf", "mimeType": "application/pdf", "modifiedTime": "2026-01-01T00:00:00Z"}
        classification = {"score": 8, "entity": "HJRP-CL", "summary": "Cinema Lanes lease", "discard_reason": ""}
        user = {"email": "harrison@hjrglobal.com", "name": "Harrison", "entity_default": "FNDR"}
        _ingest_file(kb, file_meta, "content", classification, user)
        doc = kb.upsert_documents.call_args[0][0][0]
        assert doc.entity == "HJRP"           # parent kept (guard sees canonical HJRP post-split)
        assert doc.sub_entity == "HJRP-CL"    # sub-entity preserved

    def test_canonical_set_is_materializer_codes(self):
        # Cheap guard: _CANONICAL_ENTITIES must remain a view of the materializer's
        # ENTITY_CODES (catches someone re-hardcoding it to a divergent literal).
        from cora import drive_materializer
        from cora.connectors import drive_sweep
        assert drive_sweep._CANONICAL_ENTITIES == frozenset(drive_materializer.ENTITY_CODES)

    def test_classifier_prompt_codes_collapse_to_canonical_set(self):
        # The REAL drift risk (per the Slice A D-051 review): the guard rejects any
        # code Haiku returns that is not canonical-after-split. If the classifier
        # prompt's allowed entity list drifts from _CANONICAL_ENTITIES, the guard
        # would start downgrading a legit code (prompt adds one) or accept a code
        # Haiku can never emit (ENTITY_CODES adds one). Assert the prompt's allowed
        # codes, collapsed by the SAME sub-entity split _ingest_file uses, == the
        # guard set. Extracted from the prompt text so it can't silently drift.
        import re as _re
        from cora.connectors import drive_sweep
        prompt = drive_sweep._CLASSIFY_PROMPT
        block = prompt.split("entity from:", 1)[1].split("Respond with JSON", 1)[0]
        # Codes may contain digits (F3E, F3C) and hyphens (LEX-LLC, HJRP-CL).
        codes = set(_re.findall(r"\b[A-Z][A-Z0-9]+(?:-[A-Z0-9]+)*\b", block))
        assert codes, "could not parse the classifier prompt's allowed entity codes"
        collapsed = set()
        for c in codes:
            if "-" in c and c.split("-")[0] in ("LEX", "HJRP", "HJRPROD"):
                collapsed.add(c.split("-")[0])
            else:
                collapsed.add(c)
        assert collapsed == set(drive_sweep._CANONICAL_ENTITIES), (
            "classifier prompt entity codes (collapsed to parents) diverged from the "
            f"guard's canonical set. prompt->{sorted(collapsed)} guard->{sorted(drive_sweep._CANONICAL_ENTITIES)}"
        )


# ── sweep_user ────────────────────────────────────────────────────────────────

class TestSweepUser:
    def _make_kb(self):
        kb = MagicMock()
        kb.get_sync_state.return_value = None
        kb.set_sync_state = MagicMock()
        kb.upsert_documents = MagicMock(return_value=1)
        return kb

    def _make_anthropic(self, score: int = 8):
        msg = MagicMock()
        msg.content = [MagicMock()]
        msg.content[0].text = json.dumps({
            "score": score, "entity": "F3E", "summary": "test doc", "discard_reason": ""
        })
        client = MagicMock()
        client.messages.create.return_value = msg
        return client

    def _make_drive_service(self, files: list[dict]):
        service = MagicMock()
        service.files().list().execute.return_value = {"files": files, "nextPageToken": None}
        service.files().export().execute.return_value = b"document content that is long enough to pass the 150 char threshold and contain meaningful business information for the test"
        service.files().get_media().execute.return_value = b"file content" * 20
        return service

    def test_returns_stats_dict(self):
        user = {"email": "tommy@f3energy.com", "name": "Tommy", "entity_default": "F3E",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "f1", "name": "contract.pdf",
                 "mimeType": "application/pdf",
                 "modifiedTime": "2026-01-01T00:00:00Z", "size": "10000"}
            ])
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                mock_extract.return_value = "Meaningful business content " * 20
                stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                                   freshness_days=30, dry_run=False)

        assert "files_enumerated" in stats
        assert "chunks_ingested" in stats
        assert stats["files_enumerated"] >= 1

    def test_dry_run_does_not_call_upsert(self):
        user = {"email": "tommy@f3energy.com", "name": "Tommy", "entity_default": "F3E",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "f2", "name": "notes.txt", "mimeType": "text/plain",
                 "modifiedTime": "2026-01-01T00:00:00Z", "size": "5000"}
            ])
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                mock_extract.return_value = "Business notes " * 20
                sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                           freshness_days=30, dry_run=True)

        kb.upsert_documents.assert_not_called()

    def test_phi_guard_skips_lex_phi_content(self):
        user = {"email": "shaun@lexingtonservices.com", "name": "Shaun", "entity_default": "LEX",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "f3", "name": "client-record.pdf",
                 "mimeType": "application/pdf",
                 "modifiedTime": "2026-01-01T00:00:00Z", "size": "8000"}
            ])
            # Content that triggers PHI guard
            phi_content = "Patient DOB: 1990-01-01, AHCCCS member ID: 12345, diagnosis: F84.0" * 10
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                mock_extract.return_value = phi_content
                stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                                   freshness_days=30, dry_run=False)

        assert stats["phi_skipped"] >= 1
        kb.upsert_documents.assert_not_called()

    def test_dedup_skips_shared_files(self):
        user = {"email": "gaelan@f3energy.com", "name": "Gaelan", "entity_default": "F3E",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)
        # File ID already seen by a prior user
        seen = {"already_seen_file_id"}

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "already_seen_file_id", "name": "shared.docx",
                 "mimeType": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                 "modifiedTime": "2026-01-01T00:00:00Z", "size": "5000"}
            ])
            stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                               freshness_days=30, dry_run=False, seen_file_ids=seen)

        assert stats["dedup_skipped"] == 1
        kb.upsert_documents.assert_not_called()

    def test_noise_filtered_when_score_below_4(self):
        user = {"email": "tommy@f3energy.com", "name": "Tommy", "entity_default": "F3E",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        # Low score -- should be filtered
        anthropic_client = self._make_anthropic(score=2)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "f5", "name": "template.docx", "mimeType": "text/plain",
                 "modifiedTime": "2026-01-01T00:00:00Z", "size": "5000"}
            ])
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                mock_extract.return_value = "Content " * 30
                stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                                   freshness_days=30, dry_run=False)

        assert stats["noise_filtered"] >= 1
        kb.upsert_documents.assert_not_called()

    def test_build_service_failure_returns_empty_stats(self):
        user = {"email": "nonexistent@nowhere.com", "name": "X", "entity_default": "FNDR",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = MagicMock()

        with patch("cora.connectors.drive_sweep._build_drive_service",
                   side_effect=Exception("DWD not configured")):
            stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                               freshness_days=30, dry_run=False)

        assert stats["files_enumerated"] == 0
        kb.upsert_documents.assert_not_called()

    def test_cora_internal_doc_skipped_before_extract(self):
        # WS1-DRIVE: a Cora build/audit doc is skipped at ingest BEFORE extraction,
        # so it never reaches the KB and re-poisons RAG (the Minute Press leak vector).
        user = {"email": "harrison@hjrglobal.com", "name": "Harrison", "entity_default": "FNDR",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "fc1", "name": "2026-06-16_fndr_cora-rebuild-execution-log.md",
                 "mimeType": "text/markdown",
                 "modifiedTime": "2026-06-16T00:00:00Z", "size": "9000"}
            ])
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                                   freshness_days=30, dry_run=False)

        assert stats.get("cora_internal_skipped", 0) >= 1
        kb.upsert_documents.assert_not_called()
        mock_extract.assert_not_called()  # guard fires BEFORE extraction

    def test_legit_cora_adjacent_doc_not_skipped(self):
        # Negative control: a legit cora-ADJACENT business doc is NOT skipped.
        user = {"email": "harrison@hjrglobal.com", "name": "Harrison", "entity_default": "F3E",
                "enabled": True, "dwd_eligible": True, "drive_sweep": True}
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)

        with patch("cora.connectors.drive_sweep._build_drive_service") as mock_build:
            mock_build.return_value = self._make_drive_service([
                {"id": "fc2", "name": "f3-brand-assets-cora-reference.md",
                 "mimeType": "text/markdown",
                 "modifiedTime": "2026-06-16T00:00:00Z", "size": "9000"}
            ])
            with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
                mock_extract.return_value = "F3 brand reference content " * 20
                stats = sweep_user(user, "/fake/sa.json", kb, anthropic_client,
                                   freshness_days=30, dry_run=False)

        assert stats.get("cora_internal_skipped", 0) == 0
        kb.upsert_documents.assert_called()  # legit doc still ingested

    def test_founders_os_loop_skips_cora_internal(self):
        # The dominant real leak vector: the founders_os folder walk must guard too.
        from cora.connectors.drive_sweep import _process_single_folder_files
        kb = self._make_kb()
        anthropic_client = self._make_anthropic(score=9)
        service = self._make_drive_service([
            {"id": "fos1", "name": "2026-06-16_fndr_cora-forensic-findings-report.md",
             "mimeType": "text/markdown", "modifiedTime": "2026-06-16T00:00:00Z", "size": "9000"}
        ])
        stats = {"files_enumerated": 0, "files_extracted": 0, "chunks_ingested": 0,
                 "phi_skipped": 0, "noise_filtered": 0, "dedup_skipped": 0}
        with patch("cora.connectors.drive_sweep._extract_content") as mock_extract:
            _process_single_folder_files(
                service=service, folder_id="F", label="FNDR", effective_entity="FNDR",
                kb=kb, anthropic_client=anthropic_client, cutoff_str="2020-01-01T00:00:00Z",
                dry_run=False, is_lex=False, score_threshold=4, seen_file_ids=set(), stats=stats,
            )
        assert stats.get("cora_internal_skipped", 0) >= 1
        kb.upsert_documents.assert_not_called()
        mock_extract.assert_not_called()  # guard fires BEFORE extraction


# ── run_sweep ─────────────────────────────────────────────────────────────────

class TestRunSweep:
    def _accounts_yaml(self, tmp_path) -> str:
        import yaml
        data = {
            "accounts": [
                {"email": "harrison@hjrglobal.com", "name": "Harrison",
                 "enabled": True, "dwd_eligible": True, "drive_sweep": True,
                 "entity_default": "FNDR"},
                {"email": "tommy@f3energy.com", "name": "Tommy",
                 "enabled": True, "dwd_eligible": True, "drive_sweep": True,
                 "entity_default": "F3E"},
                # drive_sweep: false -- should be excluded
                {"email": "skip@hjrglobal.com", "name": "Skip",
                 "enabled": True, "dwd_eligible": True, "drive_sweep": False,
                 "entity_default": "FNDR"},
            ]
        }
        path = tmp_path / "accounts.yaml"
        path.write_text(yaml.dump(data), encoding="utf-8")
        return str(path)

    def test_sweeps_only_eligible_accounts(self, tmp_path):
        yaml_path = self._accounts_yaml(tmp_path)
        kb = MagicMock()
        kb.get_sync_state.return_value = None
        kb.set_sync_state = MagicMock()
        anthropic_client = MagicMock()

        with patch("cora.connectors.drive_sweep.sweep_user") as mock_sweep:
            mock_sweep.return_value = {
                "files_enumerated": 5, "files_extracted": 3,
                "chunks_ingested": 10, "phi_skipped": 0,
                "noise_filtered": 2, "dedup_skipped": 0,
            }
            stats = run_sweep(
                sa_json_path="/fake/sa.json",
                accounts_yaml_path=yaml_path,
                kb=kb,
                anthropic_client=anthropic_client,
            )

        # skip@ was excluded because drive_sweep: false
        assert mock_sweep.call_count == 2
        assert stats["accounts_swept"] == 2
        assert stats["chunks_ingested"] == 20  # 10 * 2 accounts

    def test_only_email_filter(self, tmp_path):
        yaml_path = self._accounts_yaml(tmp_path)
        kb = MagicMock()
        kb.get_sync_state.return_value = None
        kb.set_sync_state = MagicMock()
        anthropic_client = MagicMock()

        with patch("cora.connectors.drive_sweep.sweep_user") as mock_sweep:
            mock_sweep.return_value = {
                "files_enumerated": 2, "files_extracted": 1,
                "chunks_ingested": 3, "phi_skipped": 0,
                "noise_filtered": 1, "dedup_skipped": 0,
            }
            stats = run_sweep(
                sa_json_path="/fake/sa.json",
                accounts_yaml_path=yaml_path,
                kb=kb,
                anthropic_client=anthropic_client,
                only_email="tommy@f3energy.com",
            )

        assert mock_sweep.call_count == 1
        assert stats["accounts_swept"] == 1

    def test_seen_file_ids_shared_across_users(self, tmp_path):
        """Verify that seen_file_ids set is passed to each sweep_user call."""
        yaml_path = self._accounts_yaml(tmp_path)
        kb = MagicMock()
        kb.get_sync_state.return_value = None
        kb.set_sync_state = MagicMock()
        anthropic_client = MagicMock()
        captured_sets = []

        def capture_seen(user, sa_json_path, kb, anthropic_client,
                         freshness_days, dry_run, seen_file_ids):
            captured_sets.append(seen_file_ids)
            seen_file_ids.add(f"file_from_{user['email']}")
            return {"files_enumerated": 1, "files_extracted": 1,
                    "chunks_ingested": 1, "phi_skipped": 0,
                    "noise_filtered": 0, "dedup_skipped": 0}

        with patch("cora.connectors.drive_sweep.sweep_user", side_effect=capture_seen):
            run_sweep(
                sa_json_path="/fake/sa.json",
                accounts_yaml_path=yaml_path,
                kb=kb,
                anthropic_client=anthropic_client,
            )

        # Same set object shared between both calls
        assert len(captured_sets) == 2
        assert captured_sets[0] is captured_sets[1]
        # Second call should see the file_id added by first call
        assert any("harrison@hjrglobal.com" in fid for fid in captured_sets[1])


# ── W4-01: founders_os self-budget + resumable checkpoint ──────────────────────

import itertools  # noqa: E402
import time as _time  # noqa: E402

from cora.connectors import drive_sweep as _ds  # noqa: E402


class _FakeKB:
    """Dict-backed KB with the real checkpoint/sync_state contract."""

    def __init__(self, sync: dict | None = None):
        self.checkpoints: dict[str, dict] = {}
        self.sync: dict[str, int] = dict(sync or {})
        self.upserts: list = []
        self.sync_writes: list[tuple[str, int]] = []
        self.deleted_checkpoints: list[str] = []

    def upsert_documents(self, docs):
        docs = list(docs)
        self.upserts.extend(docs)
        return len(docs)

    def get_checkpoint(self, key):
        return self.checkpoints.get(key)

    def set_checkpoint(self, key, data):
        self.checkpoints[key] = data

    def delete_checkpoint(self, key):
        self.deleted_checkpoints.append(key)
        self.checkpoints.pop(key, None)

    def get_sync_state(self, key):
        v = self.sync.get(key)
        return (v, None) if v is not None else None

    def set_sync_state(self, key, last_sync_at, last_source_modified=None):
        self.sync[key] = last_sync_at
        self.sync_writes.append((key, last_sync_at))


class _FakeTreeService:
    """Drive service that serves an in-memory folder tree with (optionally
    empty) file lists. Subfolder queries return children; file queries return
    that folder's files. Empty file lists mean _process_single_folder_files
    drains in one page without needing extraction/classification fakes."""

    def __init__(self, subfolders: dict[str, list[dict]],
                 files: dict[str, list[dict]] | None = None):
        self._subfolders = subfolders
        self._files = files or {}

    def files(self):  # noqa: A003 — mimics googleapiclient surface
        return self

    def list(self, **kwargs):
        import re
        q = kwargs.get("q", "")
        m = re.search(r"'([^']+)' in parents", q)
        parent = m.group(1) if m else None
        is_folder_q = "mimeType='application/vnd.google-apps.folder'" in q
        outer = self

        class _Req:
            def execute(_self):
                if is_folder_q:
                    return {"files": outer._subfolders.get(parent, []),
                            "nextPageToken": None}
                return {"files": outer._files.get(parent, []),
                        "nextPageToken": None}

        return _Req()


class TestFoundersOsEntityFor:
    def test_exact_match(self):
        assert _ds._founders_os_entity_for("08-Lexington-Services") == "LEX"

    def test_prefix_match(self):
        assert _ds._founders_os_entity_for("02-F3-Energy-extra") == "F3E"

    def test_shared_maps_fndr(self):
        assert _ds._founders_os_entity_for("_shared") == "FNDR"

    def test_skip_folder_returns_none(self):
        assert _ds._founders_os_entity_for("_archive") is None

    def test_unmapped_returns_none(self):
        assert _ds._founders_os_entity_for("random-folder") is None


class TestSweepFolderTreeCheckpoint:
    """Real _sweep_folder_tree BFS with a fake service (empty folders)."""

    def test_full_walk_marks_tree_done_and_returns_true(self):
        # root -> [a, b]; a -> [c]
        svc = _FakeTreeService(subfolders={
            "root": [{"id": "a", "name": "a"}, {"id": "b", "name": "b"}],
            "a": [{"id": "c", "name": "c"}],
            "b": [], "c": [],
        })
        kb = _FakeKB()
        done = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="FNDR", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
        )
        assert done is True
        ck = kb.checkpoints["ck"]
        assert ck["tree_done"] is True
        assert set(ck["completed_folder_ids"]) == {"root", "a", "b", "c"}

    def test_tree_done_checkpoint_short_circuits(self):
        # A tree_done checkpoint means the whole subtree is already ingested.
        svc = _FakeTreeService(subfolders={"root": [{"id": "a", "name": "a"}]})
        kb = _FakeKB()
        kb.checkpoints["ck"] = {"completed_folder_ids": ["root", "a"], "tree_done": True}
        # If it tried to walk, it would hit the service; make list() explode.
        svc.list = MagicMock(side_effect=AssertionError("should not walk"))
        done = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="FNDR", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={}, checkpoint_key="ck",
        )
        assert done is True

    def test_budget_cut_persists_partial_and_returns_false(self):
        svc = _FakeTreeService(subfolders={
            "root": [{"id": "a", "name": "a"}, {"id": "b", "name": "b"}],
            "a": [], "b": [],
        })
        kb = _FakeKB()
        # Deadline already in the past -> after processing the first folder
        # (root) the loop's between-folder check trips and returns False.
        done = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="FNDR", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
            deadline_monotonic=_time.monotonic() - 1,
        )
        assert done is False
        ck = kb.checkpoints["ck"]
        assert ck["tree_done"] is False
        assert "root" in ck["completed_folder_ids"]  # progress recorded
        assert set(ck["completed_folder_ids"]) != {"root", "a", "b"}  # not all

    def test_resume_completes_over_two_runs(self):
        # Run 1: past deadline -> stops after root, checkpoints. Run 2: no
        # deadline -> resumes, skips root, finishes a & b -> tree_done.
        svc = _FakeTreeService(subfolders={
            "root": [{"id": "a", "name": "a"}, {"id": "b", "name": "b"}],
            "a": [], "b": [],
        })
        kb = _FakeKB()
        r1 = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="LEX", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
            deadline_monotonic=_time.monotonic() - 1,
        )
        assert r1 is False and kb.checkpoints["ck"]["tree_done"] is False
        r2 = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="LEX", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
            deadline_monotonic=None,
        )
        assert r2 is True
        assert kb.checkpoints["ck"]["tree_done"] is True
        assert set(kb.checkpoints["ck"]["completed_folder_ids"]) == {"root", "a", "b"}

    def test_skip_folder_ids_not_walked(self):
        # The root tree must not descend into a sub-entity folder ('llc').
        svc = _FakeTreeService(subfolders={
            "root": [{"id": "llc", "name": "llc"}, {"id": "misc", "name": "misc"}],
            "llc": [{"id": "llc-child", "name": "x"}], "misc": [],
        })
        kb = _FakeKB()
        done = _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="LEX", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=True, score_threshold=6,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
            skip_folder_ids=frozenset({"llc"}),
        )
        assert done is True
        completed = set(kb.checkpoints["ck"]["completed_folder_ids"])
        assert "llc" not in completed and "llc-child" not in completed
        assert completed == {"root", "misc"}

    def test_dry_run_writes_no_checkpoint(self):
        svc = _FakeTreeService(subfolders={"root": []})
        kb = _FakeKB()
        _ds._sweep_folder_tree(
            service=svc, folder_id="root", entity="FNDR", sub_entity=None,
            kb=kb, anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=True, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0}, checkpoint_key="ck",
        )
        assert kb.checkpoints == {}

    def test_folder_budget_cut_midpages_not_marked_complete(self):
        # A folder whose files span >1 page and gets budget-cut mid-folder must
        # NOT be recorded as completed (so it re-processes from page 1 next run).
        class _MultiPageService:
            def files(self):  # noqa: A003
                return self

            def list(self, **kwargs):
                import re
                q = kwargs.get("q", "")
                is_folder_q = "mimeType='application/vnd.google-apps.folder'" in q
                has_token = "pageToken" in kwargs
                outer_folder_q = is_folder_q

                class _Req:
                    def execute(_self):
                        if outer_folder_q:
                            return {"files": [], "nextPageToken": None}  # no subfolders
                        # File query: page 1 has a token (empty files), page 2 ends it.
                        if not has_token:
                            return {"files": [], "nextPageToken": "p2"}
                        return {"files": [], "nextPageToken": None}

                return _Req()

        kb = _FakeKB()
        done = _ds._sweep_folder_tree(
            service=_MultiPageService(), folder_id="root", entity="LEX",
            sub_entity=None, kb=kb, anthropic_client=None,
            cutoff_str="2020-01-01T00:00:00Z", dry_run=False, is_lex=False,
            score_threshold=4, seen_file_ids=set(), stats={"files_enumerated": 0},
            checkpoint_key="ck", deadline_monotonic=_time.monotonic() - 1,
        )
        assert done is False
        ck = kb.checkpoints["ck"]
        assert ck["tree_done"] is False
        assert "root" not in ck["completed_folder_ids"]  # mid-folder cut -> not done


class TestProcessSingleFolderBudget:
    def test_multi_page_stops_on_deadline_returns_false(self):
        # Two pages; deadline in the past -> after page 1 (has token) -> False.
        svc = MagicMock()
        page1 = {"files": [], "nextPageToken": "p2"}
        page2 = {"files": [], "nextPageToken": None}
        svc.files().list().execute.side_effect = [page1, page2]
        done = _ds._process_single_folder_files(
            service=svc, folder_id="F", label="LEX", effective_entity="LEX",
            kb=_FakeKB(), anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0},
            deadline_monotonic=_time.monotonic() - 1,
        )
        assert done is False

    def test_single_page_returns_true(self):
        svc = _FakeTreeService(subfolders={}, files={"F": []})
        done = _ds._process_single_folder_files(
            service=svc, folder_id="F", label="FNDR", effective_entity="FNDR",
            kb=_FakeKB(), anthropic_client=None, cutoff_str="2020-01-01T00:00:00Z",
            dry_run=False, is_lex=False, score_threshold=4,
            seen_file_ids=set(), stats={"files_enumerated": 0},
            deadline_monotonic=_time.monotonic() - 1,  # ignored — single page
        )
        assert done is True


class TestSweepFoundersOsOrchestration:
    """sweep_founders_os ordering + watermark atomicity (patched _sweep_folder_tree)."""

    def _patch_build(self):
        # Neutralise the Drive/Sheets service builders.
        return patch.multiple(
            "cora.connectors.drive_sweep",
            _build_sa_drive_service_direct=MagicMock(return_value="svc"),
            _build_sa_sheets_service_direct=MagicMock(return_value=None),
        )

    def test_neediest_first_ordering(self):
        # FNDR has no watermark; F3E stale (1000); OSN fresh (2000).
        top = [
            {"id": "f_osn", "name": "09-One-Stop-Nutrition"},
            {"id": "f_f3e", "name": "02-F3-Energy"},
            {"id": "f_fndr", "name": "00-Founder"},
        ]
        kb = _FakeKB(sync={
            "founders_os_F3E_f_f3e": 1000,
            "founders_os_OSN_f_osn": 2000,
        })
        order: list[str] = []

        def fake_tree(**kw):
            order.append(kw["entity"])
            return True

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders", return_value=top), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        assert order == ["FNDR", "F3E", "OSN"]  # no-wm, then stalest, then fresh
        # every entity completed -> all watermarks advanced
        assert {k for k, _ in kb.sync_writes} == {
            "founders_os_FNDR_f_fndr", "founders_os_F3E_f_f3e", "founders_os_OSN_f_osn"}

    def test_interrupted_entity_does_not_advance_watermark(self):
        top = [{"id": "f_fndr", "name": "00-Founder"},
               {"id": "f_lex", "name": "08-Lexington-Services"}]
        kb = _FakeKB()

        def fake_tree(**kw):
            # FNDR completes, LEX's tree is budget-cut (returns False).
            return kw["entity"] != "LEX"

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders",
                   side_effect=lambda svc, fid: top if fid == _ds.FOUNDERS_OS_ROOT_ID else []), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree):
            agg = _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        wm_keys = {k for k, _ in kb.sync_writes}
        assert "founders_os_FNDR_f_fndr" in wm_keys       # completed -> advanced
        assert "founders_os_LEX_f_lex" not in wm_keys      # interrupted -> NOT advanced
        assert agg["budget_interrupted"] is True
        # LEX's checkpoints must NOT be cleared on interrupt.
        assert not any("f_lex" in k for k in kb.deleted_checkpoints)

    def test_root_tree_receives_sub_entity_skip_ids(self):
        top = [{"id": "f_lex", "name": "08-Lexington-Services"}]
        lex_children = [{"id": "llc", "name": "llc"}, {"id": "misc", "name": "misc"}]
        kb = _FakeKB()
        calls: list[dict] = []

        def fake_tree(**kw):
            calls.append(kw)
            return True

        def fake_list(svc, fid):
            if fid == _ds.FOUNDERS_OS_ROOT_ID:
                return top
            if fid == "f_lex":
                return lex_children
            return []

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders", side_effect=fake_list), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        root_calls = [c for c in calls if c["sub_entity"] is None]
        assert len(root_calls) == 1
        assert "llc" in (root_calls[0].get("skip_folder_ids") or set())
        # LEX watermark advanced (all subtrees returned True)
        assert any(k == "founders_os_LEX_f_lex" for k, _ in kb.sync_writes)

    def test_multi_subtree_interrupt_leaves_watermark_untouched(self):
        # LEX: sub-entity 'llc' completes, then the root tree is budget-cut.
        top = [{"id": "f_lex", "name": "08-Lexington-Services"}]
        lex_children = [{"id": "llc", "name": "llc"}]
        kb = _FakeKB()

        def fake_tree(**kw):
            return kw["sub_entity"] is not None  # sub-entity True, root False

        def fake_list(svc, fid):
            if fid == _ds.FOUNDERS_OS_ROOT_ID:
                return top
            if fid == "f_lex":
                return lex_children
            return []

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders", side_effect=fake_list), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        assert not any(k == "founders_os_LEX_f_lex" for k, _ in kb.sync_writes)
        assert kb.deleted_checkpoints == []  # nothing cleared on interrupt

    def test_between_entity_budget_defers_remaining(self):
        top = [{"id": "f_fndr", "name": "00-Founder"},
               {"id": "f_hjrg", "name": "01-HJR-Global"}]
        kb = _FakeKB()

        def fake_tree(**kw):
            return True

        # monotonic: deadline base=0 (budget 1min -> deadline 60); entity1
        # check=1 (<60, proceed); entity2 check=1e9 (>=60 -> defer).
        clock = itertools.chain([0.0, 1.0], itertools.repeat(1e9))
        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders",
                   side_effect=lambda svc, fid: top if fid == _ds.FOUNDERS_OS_ROOT_ID else []), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree), \
             patch("cora.connectors.drive_sweep.time.monotonic", side_effect=lambda: next(clock)):
            agg = _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=1)

        assert agg["budget_interrupted"] is True
        assert agg["entities_deferred"] == 1
        assert [k for k, _ in kb.sync_writes] == ["founders_os_FNDR_f_fndr"]

    def test_deadline_threaded_into_all_tree_calls(self):
        # D-051 CONFIRMED #2: the within-tree budget cut only works if
        # sweep_founders_os passes deadline_monotonic into EVERY _sweep_folder_tree
        # call (sub-entity, root-after-sub, and no-sub). Dropping it silently
        # reinstates the PT2H SIGKILL (the W4-01 bug). Pin the threading.
        top = [{"id": "f_lex", "name": "08-Lexington-Services"},
               {"id": "f_fndr", "name": "00-Founder"}]
        kb = _FakeKB()
        calls: list[dict] = []

        def fake_tree(**kw):
            calls.append(kw)
            return True

        def fake_list(svc, fid):
            if fid == _ds.FOUNDERS_OS_ROOT_ID:
                return top
            if fid == "f_lex":
                return [{"id": "llc", "name": "llc"}]
            return []

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders", side_effect=fake_list), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", side_effect=fake_tree):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=120)

        # LEX -> sub-entity tree + root tree (2), FNDR -> 1 = 3 tree calls.
        assert len(calls) == 3
        assert all(c.get("deadline_monotonic") is not None for c in calls)

    def test_resume_watermark_uses_original_sweep_start(self):
        # D-051 CONFIRMED #1: on a resumed completion the watermark advances to
        # the ORIGINAL sweep start (the start marker), NOT the completing run's
        # clock — else a file dropped into an already-completed subtree between
        # runs falls permanently below the incremental cutoff and is lost.
        top = [{"id": "f_fndr", "name": "00-Founder"}]
        kb = _FakeKB()
        # A start marker from a prior (interrupted) run, pinned to an OLD time.
        kb.checkpoints["founders_os_startmark_founders_os_FNDR_f_fndr"] = {"started_at": 1000}

        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders",
                   side_effect=lambda svc, fid: top if fid == _ds.FOUNDERS_OS_ROOT_ID else []), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", return_value=True):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        # Watermark == the ORIGINAL start (1000), NOT ~now.
        assert kb.sync["founders_os_FNDR_f_fndr"] == 1000
        # Marker cleared on completion.
        assert "founders_os_startmark_founders_os_FNDR_f_fndr" not in kb.checkpoints

    def test_fresh_entity_watermark_recent_and_marker_cleared(self):
        # Single-run (no pre-existing marker): watermark ~= now, marker written
        # then cleared on completion.
        top = [{"id": "f_fndr", "name": "00-Founder"}]
        kb = _FakeKB()
        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders",
                   side_effect=lambda svc, fid: top if fid == _ds.FOUNDERS_OS_ROOT_ID else []), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", return_value=True):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        assert kb.sync["founders_os_FNDR_f_fndr"] > 1_700_000_000  # a recent epoch
        assert "founders_os_startmark_founders_os_FNDR_f_fndr" not in kb.checkpoints

    def test_interrupt_persists_start_marker_and_no_watermark(self):
        top = [{"id": "f_lex", "name": "08-Lexington-Services"}]
        kb = _FakeKB()
        with self._patch_build(), \
             patch("cora.connectors.drive_sweep._list_subfolders",
                   side_effect=lambda svc, fid: top if fid == _ds.FOUNDERS_OS_ROOT_ID else []), \
             patch("cora.connectors.drive_sweep._sweep_folder_tree", return_value=False):
            _ds.sweep_founders_os("/sa.json", kb, None, time_budget_min=None)

        mk = kb.checkpoints.get("founders_os_startmark_founders_os_LEX_f_lex")
        assert mk and isinstance(mk["started_at"], int)  # persisted for resume
        assert not any(k == "founders_os_LEX_f_lex" for k, _ in kb.sync_writes)

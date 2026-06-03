"""Targeted ingestion of the Shaun x Jen Lexington Dump Folder.

Folder ID: 1uU-nHtEz5bFNu-JTV4k5BidkfmKAqVfG
Owner: harrison@hjrglobal.com

Bypasses the noise classifier -- every file is ingested regardless of
Haiku score. All chunks are tagged entity=LEX, sub_entity=LEX-LLC,
source=drive_asset. PHI content is stored in KB; Cora's system prompt
guardrails prevent it from being surfaced in Slack.

Usage:
    .venv\\Scripts\\python.exe scripts\\ingest_dump_folder.py
    .venv\\Scripts\\python.exe scripts\\ingest_dump_folder.py --dry-run
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()
sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ingest-dump-folder")

# ── Target folder + file IDs ──────────────────────────────────────────────────

FOLDER_ID = "1uU-nHtEz5bFNu-JTV4k5BidkfmKAqVfG"
IMPERSONATE_EMAIL = "harrison@hjrglobal.com"

# All 20 files confirmed in the folder (harvested 2026-05-31)
TARGET_FILES = [
    {"id": "1o1wo1R5g081c_RUhIkGwF2ju7GhSbZhp", "name": "Billing Claims.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1DZT-Z7yYp9mLHgoymcDsSMX29jSFGnX6", "name": "Client Assignments.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1K3zrnizRFkDmqn_G681-rGoIypOiS5nr", "name": "Client Authorization Balances.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "10S11i7D0c-Hq-jfmTI5Xd6A5EoNr8zva", "name": "Client Policy Information.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1l4I78R5ctQtAcI5-5KmWyr-RoYsyLPP4TRMjVDUBJYA", "name": "Copy of Araceli",
     "mimeType": "application/vnd.google-apps.spreadsheet"},
    {"id": "175pxbEcYCksM4uZ5E77iUtPHuVdgfgH5H0I3oglWX4U", "name": "Copy of Gabe",
     "mimeType": "application/vnd.google-apps.spreadsheet"},
    {"id": "1BIfsobt7yTW8ICH3y6tEDNYnq9NLtOmRo6woGf9P7BY", "name": "Copy of Lexington Fleet Safety Policy",
     "mimeType": "application/vnd.google-apps.document"},
    {"id": "1QrtLg4_mqr1JRifdxPXC6aV2rmADOnHyb1xbIyCi37I", "name": "Copy of Lexington LLC Policies and Procedures August 2023",
     "mimeType": "application/vnd.google-apps.document"},
    {"id": "1zcI9ZWQXItN_hHlW3WvQMeW1CWyYF7PcyFF0dyvT9jM", "name": "Copy of LLC Employee Handbook - 6.23",
     "mimeType": "application/vnd.google-apps.document"},
    {"id": "1870Q7_NfxhxBLr2eNMbEUgiTheEpzGhwJJXGicPkS1M", "name": "Copy of Lucas",
     "mimeType": "application/vnd.google-apps.spreadsheet"},
    {"id": "1WvdWKoLmpgVFY6FjNi7XpW0bM_sPm_LhXRxTohCHujQ", "name": "Copy of Madison",
     "mimeType": "application/vnd.google-apps.spreadsheet"},
    {"id": "1cSw6yq5gV5F-rUIsdWhEmD_vD96G4Lv4", "name": "DDD Complete Provider Manual.pdf",
     "mimeType": "application/pdf"},
    {"id": "17iJOwFliwqENMYEcFWxxjzPbBImJcyGc", "name": "DDD OLCR Report.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1kbBhHfXftrrRGgX5_UoSVC0NZ_CLqX02", "name": "Employee Payrates.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1ynd5fhPSD7CU1gC-S-XA-ituTfPX_F5v", "name": "Guardian Assignments.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1MAAMp1KquAHDh3PonnKWFoQV5jwLxNzE", "name": "progressReport - Jalen Alicea August 2025.pdf",
     "mimeType": "application/pdf"},
    {"id": "1JtbPxdqlAbwmj-16_kv2DpDeZXxS5SE8", "name": "Provider Information.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "1x2w66DXteWvlEVs1MhDdUaBGPrLPlfkz", "name": "Providers over 16 hours.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"id": "16XDEG5RoPZaQAVX_2_8Run_DMeSYqH14", "name": "Rate_Book_050826.pdf",
     "mimeType": "application/pdf"},
    {"id": "1ckh4ASmczQEcA8M1aWwQ1lJzlfrMYFqa", "name": "Service Codes.xlsx",
     "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
]

# ── Auth ──────────────────────────────────────────────────────────────────────

def _build_drive_service(impersonate_email: str):
    """Build a Drive v3 service authenticated as impersonate_email via DWD."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_path or not Path(sa_path).exists():
        raise RuntimeError(f"GOOGLE_SERVICE_ACCOUNT_JSON not set or missing: {sa_path!r}")

    creds = service_account.Credentials.from_service_account_file(
        sa_path,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
        subject=impersonate_email,
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


# ── Content extraction (mirrors drive_sweep.py) ───────────────────────────────

_GOOGLE_DOC_MIME = "application/vnd.google-apps.document"
_GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
_GOOGLE_SLIDE_MIME = "application/vnd.google-apps.presentation"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MIME = "application/pdf"


def _extract_content(service, file_meta: dict) -> str:
    mime = file_meta["mimeType"]
    fid = file_meta["id"]

    try:
        if mime == _GOOGLE_DOC_MIME or mime == _GOOGLE_SLIDE_MIME:
            data = service.files().export(fileId=fid, mimeType="text/plain").execute()
            return data.decode("utf-8", errors="replace") if isinstance(data, bytes) else str(data)

        if mime == _GOOGLE_SHEET_MIME:
            data = service.files().export(fileId=fid, mimeType=_XLSX_MIME).execute()
            return _parse_xlsx(data)

        if mime == _PDF_MIME:
            data = service.files().get_media(fileId=fid).execute()
            return _parse_pdf(data)

        if mime == _XLSX_MIME:
            data = service.files().get_media(fileId=fid).execute()
            return _parse_xlsx(data)

        # Other binary -- skip
        return ""

    except Exception as exc:
        log.warning("extract failed for %s (%s): %s", file_meta["name"], fid, exc)
        return ""


def _parse_xlsx(data: bytes) -> str:
    import openpyxl, io as _io
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    rows_text.append("\t".join(cells))
            if rows_text:
                parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text[:500]))
        return "\n\n".join(parts)
    except Exception as exc:
        log.warning("xlsx parse failed: %s", exc)
        return ""


def _parse_pdf(data: bytes) -> str:
    try:
        import pdfplumber, io as _io
        with pdfplumber.open(_io.BytesIO(data)) as pdf:
            pages = []
            for page in pdf.pages[:80]:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
    except Exception as exc:
        log.warning("pdf parse failed: %s", exc)
        return ""


# ── Chunking ──────────────────────────────────────────────────────────────────

_CHUNK_SIZE = 1400
_CHUNK_OVERLAP = 150


def _chunk_text(text: str) -> list[str]:
    if len(text) <= _CHUNK_SIZE:
        return [text]
    chunks = []
    start = 0
    while start < len(text):
        end = start + _CHUNK_SIZE
        chunks.append(text[start:end])
        start = end - _CHUNK_OVERLAP
    return chunks


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true",
                        help="Extract and chunk but do not write to KB")
    args = parser.parse_args()

    from cora.knowledge_base.store import KnowledgeBase, Document
    db_path = Path(__file__).parent.parent / "data" / "cora_kb.db"
    kb = KnowledgeBase(db_path)

    log.info("Building Drive service (impersonating %s)", IMPERSONATE_EMAIL)
    service = _build_drive_service(IMPERSONATE_EMAIL)

    total_chunks = 0
    ingested_files = 0
    skipped_files = 0

    for file_meta in TARGET_FILES:
        name = file_meta["name"]
        fid = file_meta["id"]
        log.info("Processing: %s", name)

        content = _extract_content(service, file_meta)
        if not content or not content.strip():
            log.warning("  -> no content extracted, skipping")
            skipped_files += 1
            continue

        chunks = _chunk_text(content.strip())
        log.info("  -> %d chars, %d chunks", len(content), len(chunks))

        if args.dry_run:
            log.info("  [DRY RUN] would ingest %d chunks", len(chunks))
            total_chunks += len(chunks)
            ingested_files += 1
            continue

        # Build Document and upsert
        doc = Document(
            source="drive_asset",
            source_id=fid,
            entity="LEX",
            sub_entity="LEX-LLC",
            title=name,
            content="\n\n---\n\n".join(chunks),
            deep_link=f"https://drive.google.com/file/d/{fid}/view",
            metadata=json.dumps({
                "folder": "Shaun x Jen Lexington Dump Folder",
                "folder_id": FOLDER_ID,
                "mime_type": file_meta["mimeType"],
                "ingested_by": "ingest_dump_folder.py",
            }),
        )

        # Upsert each chunk individually so embeddings are per-chunk
        chunk_count = 0
        for i, chunk_text in enumerate(chunks):
            chunk_doc = Document(
                source="drive_asset",
                source_id=f"{fid}:chunk{i}",
                entity="LEX",
                sub_entity="LEX-LLC",
                title=name,
                content=chunk_text,
                deep_link=f"https://drive.google.com/file/d/{fid}/view",
                metadata={
                    "folder": "Shaun x Jen Lexington Dump Folder",
                    "folder_id": FOLDER_ID,
                    "chunk_index": i,
                    "total_chunks": len(chunks),
                    "mime_type": file_meta["mimeType"],
                },
            )
            kb.upsert_documents([chunk_doc])
            chunk_count += 1

        log.info("  -> ingested %d chunks for %s", chunk_count, name)
        total_chunks += chunk_count
        ingested_files += 1

    log.info("=" * 50)
    log.info("DONE: %d files ingested, %d skipped, %d total chunks",
             ingested_files, skipped_files, total_chunks)
    if args.dry_run:
        log.info("[DRY RUN] -- no data was written to KB")


if __name__ == "__main__":
    main()

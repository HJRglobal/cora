"""Recurring sync of the Shaun x Jen Lexington Dump Folder into the LEX KB.

Replaces the one-shot scripts/ingest_dump_folder.py (2026-06-01, hardcoded
20-file list). That one-shot missed everything added later -- notably the
"DDD Policies" shortcut (added 2026-06-04) holding the DDD Complete
Provider / Operations / Medical manuals + the EVV Documents folder, which is
why Cora could not answer Shaun's live-in caregiver EVV question
(#lex-leadership 2026-06-11, Asana 1215643646634974).

Behavior:
- Enumerates the dump folder RECURSIVELY, following folder shortcuts
  (the DDD Policies entry is a shortcut to a folder elsewhere in Drive).
- Watermark-based incremental: only files with Drive modifiedTime newer than
  the last sync are re-ingested (sync_state source "lex_dump_folder").
  --backfill forces everything.
- Entity tagging: entity=LEX, sub_entity=LEX-LLC for EVERYTHING in the dump
  folder, including the DDD Policies tree (Harrison directive 2026-06-11 PM,
  superseding the GM-level tagging shipped earlier the same day -- see D-046
  amendment). Rationale: the DDD policy consumers (Shaun/Jen/Jeff/Aaron) live
  in #llc-* channels per the 6/11 LLC routing directive, and the strict
  sub-entity filter excludes GM-level NULL chunks from those channels. The
  explicit LEX-LLC tag makes the manuals visible in #llc-* AND in GM #lex-*
  channels (GM scope sees all LEX chunks); only LTS/LBHS/LLA channels do not
  see them. The explicit tag also means store Step 0 auto-detection never
  fires (explicit sub_entity is never overridden).
- PHI guard: phi_guard.is_phi_risk runs per chunk and the per-file trip count
  is logged + stored in metadata.phi_risk_chunks. Published policy manuals
  trip the program-keyword patterns (ahcccs/medicaid/assessment) on nearly
  every chunk BY CONSTRUCTION -- they are manuals ABOUT those topics -- so
  the guard is an audit signal, not a scope-downgrade; everything is already
  LEX-LLC (the tightest custodian-gated scope). Surfacing remains governed by
  the prompt guardrails + PHI custodian gate.
- Large files: > MAX_FILE_MB skipped with a logged note. PDFs are parsed
  fully (no 80-page cap -- the old cap silently truncated the 460+ page
  Provider Manual past page 80).
- Legacy cleanup: the 2026-06-01 one-shot wrote chunk rows with
  source_id "{fid}:chunkN"; when this sync re-ingests a file it deletes
  those stale rows so the KB holds exactly one copy.

Usage:
    .venv\\Scripts\\python.exe scripts\\run_lex_dump_folder_sync.py [--dry-run] [--backfill]

Scheduled task: "Cora - LEX Dump Folder Sync" daily 4:45am AZ
(deployment/setup-lex-dump-folder-sync-task.ps1).
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

_REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO_ROOT))
sys.path.insert(0, str(_REPO_ROOT / "src"))

load_dotenv(_REPO_ROOT / ".env")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("lex-dump-folder-sync")

# -- Config -------------------------------------------------------------------

FOLDER_ID = "1uU-nHtEz5bFNu-JTV4k5BidkfmKAqVfG"  # Shaun x Jen Lexington Dump Folder
IMPERSONATE_EMAIL = "harrison@hjrglobal.com"
SYNC_STATE_SOURCE = "lex_dump_folder"
KB_SOURCE = "drive_asset"
MAX_FILE_MB = 60
MAX_PDF_PAGES = 2000
MAX_DEPTH = 4

# Folder names marking the published-policy subtree (DDD Policies + EVV
# Documents -- AHCCCS/DES published manuals, FAQs, compliance notices).
# Tracked as metadata provenance (policy_tree) only; tagging is uniform
# LEX-LLC per the 2026-06-11 PM directive.
_POLICY_TREE_FOLDER_NAMES = frozenset({"ddd policies", "evv documents"})

_GOOGLE_DOC_MIME = "application/vnd.google-apps.document"
_GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
_GOOGLE_SLIDE_MIME = "application/vnd.google-apps.presentation"
_GOOGLE_FOLDER_MIME = "application/vnd.google-apps.folder"
_GOOGLE_SHORTCUT_MIME = "application/vnd.google-apps.shortcut"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLSM_MIME = "application/vnd.ms-excel.sheet.macroenabled.12"
_PDF_MIME = "application/pdf"

_PARSEABLE_MIMES = {
    _GOOGLE_DOC_MIME, _GOOGLE_SHEET_MIME, _GOOGLE_SLIDE_MIME,
    _XLSX_MIME, _XLSM_MIME, _PDF_MIME,
}


# -- Auth ----------------------------------------------------------------------

def _build_drive_service(impersonate_email: str):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_path or not Path(sa_path).exists():
        raise RuntimeError(f"GOOGLE_SERVICE_ACCOUNT_JSON not set or missing: {sa_path!r}")

    creds = service_account.Credentials.from_service_account_file(
        sa_path,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
        subject=impersonate_email,
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


# -- Enumeration ----------------------------------------------------------------

def walk_folder(service, folder_id: str, *, policy_tree: bool = False,
                depth: int = 0, path: str = "") -> list[dict]:
    """Recursively list files under folder_id, following folder shortcuts.

    Returns flat list of file dicts with added keys: "path" (folder path within
    the dump tree) and "policy_tree" (True when the file sits inside the
    curated published-policy subtree -- see _POLICY_TREE_FOLDER_NAMES;
    provenance metadata only, tagging is uniform LEX-LLC).
    """
    if depth > MAX_DEPTH:
        log.warning("Max depth %d exceeded at %s -- not descending", MAX_DEPTH, path)
        return []

    files: list[dict] = []
    page_token = None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields=("nextPageToken, files(id,name,mimeType,size,modifiedTime,"
                    "shortcutDetails)"),
            pageSize=200,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()

        for f in resp.get("files", []):
            name = f.get("name", "")
            mime = f.get("mimeType", "")
            child_pt = policy_tree or name.strip().lower() in _POLICY_TREE_FOLDER_NAMES

            if mime == _GOOGLE_FOLDER_MIME:
                files.extend(walk_folder(
                    service, f["id"], policy_tree=child_pt,
                    depth=depth + 1, path=f"{path}/{name}",
                ))
            elif mime == _GOOGLE_SHORTCUT_MIME:
                details = f.get("shortcutDetails") or {}
                target_id = details.get("targetId")
                target_mime = details.get("targetMimeType", "")
                if target_id and target_mime == _GOOGLE_FOLDER_MIME:
                    files.extend(walk_folder(
                        service, target_id, policy_tree=child_pt,
                        depth=depth + 1, path=f"{path}/{name}",
                    ))
                elif target_id:
                    try:
                        target = service.files().get(
                            fileId=target_id,
                            fields="id,name,mimeType,size,modifiedTime",
                            supportsAllDrives=True,
                        ).execute()
                        target["path"] = path
                        target["policy_tree"] = child_pt
                        files.append(target)
                    except Exception as exc:
                        log.warning("shortcut target fetch failed for %s: %s", name, exc)
            else:
                f["path"] = path
                f["policy_tree"] = policy_tree
                files.append(f)

        page_token = resp.get("nextPageToken")
        if not page_token:
            return files


# -- Content extraction ----------------------------------------------------------

def _extract_content(service, file_meta: dict) -> str:
    mime = file_meta["mimeType"]
    fid = file_meta["id"]
    try:
        if mime in (_GOOGLE_DOC_MIME, _GOOGLE_SLIDE_MIME):
            data = service.files().export(fileId=fid, mimeType="text/plain").execute()
            return data.decode("utf-8", errors="replace") if isinstance(data, bytes) else str(data)
        if mime == _GOOGLE_SHEET_MIME:
            data = service.files().export(fileId=fid, mimeType=_XLSX_MIME).execute()
            return _parse_xlsx(data)
        if mime == _PDF_MIME:
            data = service.files().get_media(fileId=fid).execute()
            return _parse_pdf(data)
        if mime in (_XLSX_MIME, _XLSM_MIME):
            data = service.files().get_media(fileId=fid).execute()
            return _parse_xlsx(data)
        return ""
    except Exception as exc:
        log.warning("extract failed for %s (%s): %s", file_meta.get("name"), fid, exc)
        return ""


def _parse_xlsx(data: bytes) -> str:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    rows_text.append("\t".join(cells))
            if rows_text:
                parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text[:500]))
        return "\n\n".join(parts)
    except Exception as exc:
        log.warning("xlsx parse failed: %s", exc)
        return ""


def _parse_pdf(data: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            pages = []
            for page in pdf.pages[:MAX_PDF_PAGES]:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
    except Exception as exc:
        log.warning("pdf parse failed: %s", exc)
        return ""


# -- Tagging ---------------------------------------------------------------------

def resolve_sub_entity(file_meta: dict) -> tuple[str, bool]:
    """Return (sub_entity, policy_tree_flag) for a dump-folder file.

    UNIFORM LEX-LLC (Harrison directive 2026-06-11 PM, supersedes the GM-level
    tagging shipped earlier the same day): the DDD policy consumers live in
    #llc-* channels, where the strict sub-entity filter excludes NULL chunks.
    LEX-LLC chunks remain visible in GM #lex-* channels too. The policy_tree
    flag rides along as metadata provenance only.
    """
    return "LEX-LLC", bool(file_meta.get("policy_tree"))


def _parse_drive_time(value: str | None) -> int | None:
    if not value:
        return None
    try:
        return int(datetime.fromisoformat(value.replace("Z", "+00:00")).timestamp())
    except Exception:
        return None


# -- Main ------------------------------------------------------------------------

def run(dry_run: bool = False, backfill: bool = False) -> dict:
    from cora.knowledge_base.store import Document, KnowledgeBase
    from cora.phi_guard import is_phi_risk

    db_path = _REPO_ROOT / "data" / "cora_kb.db"
    kb = KnowledgeBase(db_path)

    state = kb.get_sync_state(SYNC_STATE_SOURCE)
    watermark = (state[1] or 0) if (state and not backfill) else 0
    log.info("Watermark: %s (backfill=%s)", watermark, backfill)

    log.info("Building Drive service (impersonating %s)", IMPERSONATE_EMAIL)
    service = _build_drive_service(IMPERSONATE_EMAIL)

    log.info("Enumerating dump folder %s recursively...", FOLDER_ID)
    all_files = walk_folder(service, FOLDER_ID)
    log.info("Found %d files (%d in the DDD policy tree)",
             len(all_files), sum(1 for f in all_files if f.get("policy_tree")))

    ingested = skipped_unchanged = skipped_large = skipped_empty = 0
    total_chunks = 0
    max_modified = watermark

    for f in all_files:
        name = f.get("name", "")
        fid = f["id"]
        mime = f.get("mimeType", "")
        modified_ts = _parse_drive_time(f.get("modifiedTime")) or 0
        max_modified = max(max_modified, modified_ts)

        if mime not in _PARSEABLE_MIMES:
            log.info("SKIP (unsupported mime %s): %s", mime, name)
            continue
        if modified_ts and modified_ts <= watermark:
            skipped_unchanged += 1
            continue
        size_mb = int(f.get("size") or 0) / (1024 * 1024)
        if size_mb > MAX_FILE_MB:
            log.warning("SKIP (>%dMB, %.1fMB): %s -- ingest manually if needed",
                        MAX_FILE_MB, size_mb, name)
            skipped_large += 1
            continue

        sub_entity, policy_tree = resolve_sub_entity(f)
        log.info("Processing: %s%s (%.1fMB) -> sub_entity=%s",
                 f.get("path", ""), "/" + name, size_mb, sub_entity)

        content = _extract_content(service, f)
        if not content or not content.strip():
            log.warning("  -> no content extracted, skipping")
            skipped_empty += 1
            continue

        # Chunk-level PHI audit (store will re-chunk identically; this uses the
        # same chunk size for an accurate count).
        from cora.knowledge_base.chunker import chunk_text
        chunks = chunk_text(content)
        phi_chunks = sum(1 for c in chunks if is_phi_risk(c))
        if phi_chunks:
            log.info("  -> PHI-risk patterns in %d/%d chunks%s",
                     phi_chunks, len(chunks),
                     " (expected for published policy text)" if policy_tree else "")

        if dry_run:
            log.info("  [DRY RUN] would ingest %d chunks", len(chunks))
            ingested += 1
            total_chunks += len(chunks)
            continue

        # Legacy cleanup: the 2026-06-01 one-shot wrote per-chunk rows with
        # source_id "{fid}:chunkN" -- delete them so this file isn't doubled.
        cur = kb._conn.cursor()
        cur.execute(
            "SELECT chunk_id FROM knowledge_chunks WHERE source = ? AND source_id LIKE ?",
            (KB_SOURCE, f"{fid}:chunk%"),
        )
        old_ids = [row[0] for row in cur.fetchall()]
        if old_ids:
            placeholders = ",".join("?" * len(old_ids))
            cur.execute(f"DELETE FROM knowledge_vec_bin WHERE chunk_id IN ({placeholders})", old_ids)
            cur.execute(f"DELETE FROM knowledge_vec_f32 WHERE chunk_id IN ({placeholders})", old_ids)
            cur.execute(f"DELETE FROM knowledge_chunks WHERE chunk_id IN ({placeholders})", old_ids)
            kb._conn.commit()
            log.info("  -> removed %d legacy one-shot chunks", len(old_ids))

        doc = Document(
            source=KB_SOURCE,
            source_id=fid,
            entity="LEX",
            sub_entity=sub_entity,
            title=name,
            content=content.strip(),
            date_modified=modified_ts or int(time.time()),
            deep_link=f"https://drive.google.com/file/d/{fid}/view",
            metadata={
                "folder": "Shaun x Jen Lexington Dump Folder",
                "folder_id": FOLDER_ID,
                "folder_path": f.get("path", ""),
                "mime_type": mime,
                "policy_tree": policy_tree,
                "phi_risk_chunks": phi_chunks,
                "ingested_by": "run_lex_dump_folder_sync.py",
            },
        )
        n = kb.upsert_documents([doc])
        log.info("  -> ingested %d chunks", n)
        ingested += 1
        total_chunks += n

    if not dry_run:
        kb.set_sync_state(SYNC_STATE_SOURCE, int(time.time()), max_modified)

    summary = {
        "files_found": len(all_files),
        "ingested": ingested,
        "skipped_unchanged": skipped_unchanged,
        "skipped_large": skipped_large,
        "skipped_empty": skipped_empty,
        "total_chunks": total_chunks,
    }
    log.info("DONE: %s%s", summary, " [DRY RUN]" if dry_run else "")
    return summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LEX Dump Folder recurring KB sync")
    parser.add_argument("--dry-run", action="store_true",
                        help="Enumerate + extract + chunk but write nothing")
    parser.add_argument("--backfill", action="store_true",
                        help="Ignore the watermark and re-ingest everything")
    args = parser.parse_args()
    run(dry_run=args.dry_run, backfill=args.backfill)

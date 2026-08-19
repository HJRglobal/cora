"""Turn a Cora-rendered finance report into an .xlsx the requester can open.

cq-c51123b0ad07 (Justin, 2026-08-10): "enable download of displayed data as
.xlsx". The staged prompt asked which screens count as "displayed data" and told
whoever built it not to guess. Scoped here to the surfaces the request came from
and the ones whose output is genuinely tabular by construction: the finance /
QBO report renderers.

WHY PARSE THE RENDERED TEXT RATHER THAN THE SOURCE DATA
------------------------------------------------------
Because the rendered text is the CONTRACT. D-095 says the LLM never computes
financials: `qbo_client.format_*_for_llm` echoes QBO's own strings verbatim into
`  • Label: value` lines and does no arithmetic. An export built from a
re-fetched or re-derived source could disagree with the message it claims to be
a copy of -- two numbers for one question, which on a money surface is worse
than no export at all. Parsing the exact string that was displayed makes the
sheet a TRANSCRIPTION: it cannot introduce a figure the reader did not see.

For the same reason this module does no arithmetic either -- no totals, no
percentages, no reformatting of a value. Values land as text in the cell, which
also stops Excel silently coercing "1-2" into a date or trimming a leading zero
off an account reference.
"""

from __future__ import annotations

import io
import logging
import re
import time

log = logging.getLogger("cora.table_export")

MAX_ROWS = 5000
MAX_CELL_CHARS = 2000
_SHEET_NAME_MAX = 31

# `  • Label: value` -- the shape every deterministic finance renderer emits.
_BULLET = re.compile(r"^\s*[•\-\*]\s+(?P<label>.+?):\s*(?P<value>.*)$")
# A section heading: a non-bullet line ending in ':' with nothing after it.
_HEADING = re.compile(r"^(?P<heading>[^•\-\*].*?):\s*$")


def rows_from_report(text: str) -> list[list[str]]:
    """Parse a rendered report into ``[[section, label, value], ...]``.

    Non-matching lines are kept as a single-cell row rather than dropped: a
    footer like "3 of 4 stores reported" is exactly the caveat that must not be
    lost when the numbers are lifted out of their message. An export that
    silently omits the qualifier next to a partial figure is how a partial
    number gets read as a complete one.
    """
    rows: list[list[str]] = [["Section", "Item", "Value"]]
    section = ""
    truncated = False
    first_content_line = True
    # The cap is checked on EVERY branch. The first cut checked it only in the
    # non-matching-line branch, so a report of pure bullets never tripped it and
    # build_xlsx simply sliced at MAX_ROWS with no marker -- and the marker it
    # did append pushed the list to MAX_ROWS+1, exactly where build_xlsx's own
    # [:MAX_ROWS] slice discarded it. A cap that drops its own truncation notice
    # is the failure this module's docstring says it prevents.
    for raw in (text or "").splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        if len(rows) >= MAX_ROWS - 1:      # leave room for the marker itself
            truncated = True
            break
        m = _BULLET.match(line)
        if m:
            first_content_line = False
            rows.append([section,
                         m.group("label").strip()[:MAX_CELL_CHARS],
                         m.group("value").strip()[:MAX_CELL_CHARS]])
            continue
        # The report's OWN header line ("Expense detail for HJRG (...):") ends in
        # a colon and would otherwise be captured as the first Section, stamping
        # the report title onto every row above the first real QBO section.
        h = _HEADING.match(line) if not first_content_line else None
        first_content_line = False
        if h:
            section = h.group("heading").strip()[:MAX_CELL_CHARS]
            rows.append([section, "", ""])
            continue
        rows.append([line.strip()[:MAX_CELL_CHARS], "", ""])
    if truncated:
        rows.append(["(truncated -- report longer than the export cap)", "", ""])
    return rows


def safe_sheet_name(title: str) -> str:
    """Excel rejects []:*?/\\ and anything over 31 chars, and openpyxl raises.

    cq-ad74f3908e8d is the sibling of this bug in the spreadsheet_build
    archetype, where an over-long tab name silently fell back to markdown. Here
    the name is sanitized rather than allowed to fail, because the tab name is
    cosmetic and the data is not.
    """
    name = re.sub(r"[\[\]:*?/\\]", "-", (title or "").strip()) or "Report"
    return name[:_SHEET_NAME_MAX]


def build_xlsx(rows: list[list[str]], sheet_title: str = "Report") -> bytes:
    """Render rows to xlsx bytes. Every cell is written as TEXT.

    Raises ImportError if openpyxl is unavailable; callers fall back to inline.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(sheet_title)
    widths: dict[int, int] = {}
    for r in rows[:MAX_ROWS]:
        cells = [("" if c is None else str(c))[:MAX_CELL_CHARS] for c in r]
        ws.append(cells)
        # openpyxl's value binder types a leading "=" as a FORMULA. The module
        # contract is that values land as TEXT -- a transcription must never
        # become something Excel evaluates.
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, str):
                cell.data_type = "s"
        for i, c in enumerate(cells, start=1):
            widths[i] = min(60, max(widths.get(i, 10), len(c) + 2))
    for i, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    if ws.max_row >= 1:
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(title: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (title or "report").lower()).strip("-") or "report"
    return f"{slug[:60]}-{int(time.time())}.xlsx"


def deliver_report_as_xlsx(slack_client, channel_id: str, title: str,
                           report_text: str, thread_ts: str | None = None
                           ) -> tuple[str, str]:
    """Build + upload. Returns ``(outcome, detail)`` from slack_file_upload.

    The report text is egress-sanitized BEFORE the workbook is built. A binary
    payload cannot be sanitized after the fact -- the WebClient egress patch only
    wraps chat_* sends, and the bytes are PUT straight to Slack's upload URL --
    so the guard has to run on the text while it is still text, or an xlsx
    becomes a way to ship exactly what sanitize_text exists to strip.
    """
    from .slack_file_upload import FAILED, upload_bytes

    try:
        from ..slack_egress import sanitize_text
        report_text = sanitize_text(report_text)
        title = sanitize_text(title)
    except Exception:  # noqa: BLE001
        log.exception("table_export: egress sanitize failed; not exporting")
        return FAILED, "sanitize failed"

    try:
        payload = build_xlsx(rows_from_report(report_text), title)
    except ImportError:
        log.warning("table_export: openpyxl unavailable -- cannot build xlsx")
        return FAILED, "openpyxl unavailable"
    except Exception as exc:  # noqa: BLE001
        log.warning("table_export: xlsx build failed: %s", exc)
        return FAILED, str(exc)

    return upload_bytes(
        slack_client, channel_id, export_filename(title), payload, title,
        thread_ts,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

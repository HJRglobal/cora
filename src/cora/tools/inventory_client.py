"""F3E inventory pulse — reads the latest Cotton 3PL weekly inventory report from Drive.

The report is an xlsx file named 'F3 Energy LLC - Weekly Inventory Report.xlsx'
stored in Google Drive.  It has 6 sheets; we consume three:

  UNIS       — Cotton 3PL warehouse inventory (Available / Allocated / On Hand per SKU)
  NIMBL      — Nimbl DTC fulfilment centre (lot-level rows → aggregate to SKU totals)
  117 office — In-house office stock (Available / Damaged per SKU)

Auth: reuses GOOGLE_SERVICE_ACCOUNT_JSON + CORA_DRIVE_IMPERSONATE from
drive_connector.py (drive.readonly scope covers file-content download).

Behavioural contract:
  - Source-opaque: never surface file IDs, sheet names, or Drive URLs in output
  - Report date shown as soft freshness label ("as of YYYY-MM-DD")
  - On Drive/parse error → UNKNOWN_RESPONSE pattern (same as financial_client.py)
"""

from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

# ── SKU metadata ──────────────────────────────────────────────────────────────
# Maps Cotton/Nimbl Item ID → (display name, brand bucket)
_SKU_META: dict[str, tuple[str, str]] = {
    "F3-Original":     ("Original Energy",        "Energy"),
    "F3-Citrus":       ("Citrus Clarity",          "Energy"),
    "F3-Tropical":     ("Tropical Theory",         "Energy"),
    "F3SL":            ("Strawberry Lemonade",     "Energy"),
    "F3VPE4":          ("Energy Variety Pack",     "Energy"),
    "F3-Original24pk": ("Original 24-pack",        "Energy"),
    "F3-Orange":       ("Orangesicle",             "Mood"),
    "F3-Peach":        ("Peach Paradise",          "Mood"),
    "F3SC":            ("Strawberries & Cream",    "Mood"),
    "F3PC":            ("Piña Colada",             "Mood"),
    "F3VPM4":          ("Mood Variety Pack",       "Mood"),
    "PURE-Original":   ("Pure Original",           "Pure"),
    "PURE-Citrus":     ("Pure Citrus Clarity",     "Pure"),
    "PURE-Tropical":   ("Pure Tropical Theory",    "Pure"),
    "PURE-SL":         ("Pure Strawberry Lemon.",  "Pure"),
}

# Brand display order
_BRAND_ORDER = ("Energy", "Mood", "Pure")

# Safety-stock thresholds (cases of 12)
_CRITICAL_CS = 50
_LOW_CS = 200


class InventoryClientError(Exception):
    """Raised when inventory data cannot be retrieved or parsed."""


# ── UNIS sheet column indices (0-based, from live file inspection) ────────────
_U_ITEM_ID   = 1
_U_AVAILABLE = 9
_U_ALLOCATED = 11
_U_ON_HAND   = 16
_U_GOODS_TYPE = 18

# ── NIMBL sheet column indices ────────────────────────────────────────────────
_N_ITEM_NUMBER = 2
_N_QUANTITY    = 4

# ── 117 office sheet column indices ──────────────────────────────────────────
_O_ITEM_ID  = 0
_O_AVAILABLE = 2
_O_DAMAGED   = 3

# ── Sheet names ───────────────────────────────────────────────────────────────
_SHEET_UNIS   = "UNIS"
_SHEET_NIMBL  = "NIMBL"
_SHEET_OFFICE = "117 office"

# ── Drive search ──────────────────────────────────────────────────────────────
_REPORT_FILENAME = "F3 Energy LLC - Weekly Inventory Report.xlsx"

# Env var that overrides the pinned canonical file id at runtime.
_INVENTORY_FILE_ID_ENV = "F3E_INVENTORY_FILE_ID"


def _canonical_inventory_file_id() -> str | None:
    """Return the pinned canonical fileId for the weekly inventory report, or None.

    Resolution order: env override -> data/maps/canonical-files.yaml. Fail-OPEN:
    any error returns None so the caller falls back to name-based search. The
    pinned id makes the lookup deterministic (immune to title collisions and
    modifiedTime ties).
    """
    env = os.environ.get(_INVENTORY_FILE_ID_ENV, "").strip()
    if env:
        return env
    try:
        import yaml  # noqa: PLC0415
        path = Path(__file__).resolve().parents[3] / "data" / "maps" / "canonical-files.yaml"
        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        fid = ((data.get("f3e_weekly_inventory") or {}).get("file_id") or "").strip()
        return fid or None
    except Exception as exc:  # missing / malformed yaml -> name-search fallback
        log.debug("canonical-files.yaml unreadable (%s) — using name search", exc)
        return None


# ────────────────────────────────────────────────────────────────────────────
# Drive helpers
# ────────────────────────────────────────────────────────────────────────────


def _build_service():
    """Return a Drive v3 service via the shared drive_connector auth."""
    from ..connectors.drive_connector import _build_drive_service  # noqa: PLC0415
    return _build_drive_service()


def _find_latest_file(service) -> tuple[str, str]:
    """Resolve the canonical weekly inventory report. Returns (file_id, modified_iso).

    Prefers the PINNED canonical fileId (deterministic). On 404 / trashed /
    unreadable, falls back to name-based search: most-recently-modified, largest
    among results sharing that modifiedTime (the largest has the most complete
    data — all three sections).
    """
    # Layer 1 — pinned canonical fileId (deterministic, no title-collision risk).
    canonical = _canonical_inventory_file_id()
    if canonical:
        try:
            meta = service.files().get(
                fileId=canonical, fields="id, modifiedTime, trashed"
            ).execute()
            if meta and not meta.get("trashed", False):
                log.info(
                    "Inventory: using pinned canonical file %s (modified %s)",
                    meta.get("id"), meta.get("modifiedTime"),
                )
                return meta["id"], meta.get("modifiedTime", "")
            log.warning(
                "Inventory canonical file %s missing/trashed — falling back to name search",
                canonical,
            )
        except Exception as exc:
            log.warning(
                "Inventory canonical file %s lookup failed (%s) — falling back to name search",
                canonical, exc,
            )

    # Layer 2 — name-based search fallback (legacy heuristic).
    try:
        resp = service.files().list(
            q=f"name = '{_REPORT_FILENAME}' and trashed = false",
            fields="files(id, modifiedTime, size)",
            orderBy="modifiedTime desc",
            pageSize=10,
        ).execute()
    except Exception as exc:
        raise InventoryClientError(f"Drive search failed: {exc}") from exc

    files = resp.get("files", [])
    if not files:
        raise InventoryClientError(
            f"No file named {_REPORT_FILENAME!r} found in Drive."
        )

    latest_time = files[0]["modifiedTime"]
    # Prefer largest file among those with the same (newest) timestamp
    candidates = [f for f in files if f["modifiedTime"] == latest_time]
    best = max(candidates, key=lambda f: int(f.get("size", 0)))
    log.info(
        "Inventory: using file %s (modified %s, %s bytes)",
        best["id"], best["modifiedTime"], best.get("size"),
    )
    return best["id"], best["modifiedTime"]


def _download_file(service, file_id: str) -> bytes:
    """Download raw xlsx bytes from Drive."""
    try:
        request = service.files().get_media(fileId=file_id)
        return request.execute()
    except Exception as exc:
        raise InventoryClientError(f"Drive download failed: {exc}") from exc


# ────────────────────────────────────────────────────────────────────────────
# xlsx parsing
# ────────────────────────────────────────────────────────────────────────────


def _int_val(v: Any) -> int:
    """Safely cast a cell value to int, defaulting to 0."""
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _parse_unis(wb) -> dict[str, dict[str, int]]:
    """Parse the UNIS sheet.  Returns {item_id: {available, allocated, on_hand, damaged}}.

    DAMAGE-type rows contribute only to 'damaged'; GOOD/QC rows contribute to
    available / allocated / on_hand.
    """
    if _SHEET_UNIS not in wb.sheetnames:
        log.warning("Sheet %r not found in workbook", _SHEET_UNIS)
        return {}

    ws = wb[_SHEET_UNIS]
    agg: dict[str, dict[str, int]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        item_id = str(row[_U_ITEM_ID] or "").strip()
        if not item_id:
            continue
        goods_type = str(row[_U_GOODS_TYPE] or "").strip().upper()

        if item_id not in agg:
            agg[item_id] = {"available": 0, "allocated": 0, "on_hand": 0, "damaged": 0}

        if goods_type == "DAMAGE":
            agg[item_id]["damaged"] += _int_val(row[_U_ON_HAND])
        else:
            agg[item_id]["available"] += _int_val(row[_U_AVAILABLE])
            agg[item_id]["allocated"] += _int_val(row[_U_ALLOCATED])
            agg[item_id]["on_hand"]   += _int_val(row[_U_ON_HAND])

    return agg


def _parse_nimbl(wb) -> dict[str, int]:
    """Parse the NIMBL sheet.  Returns {item_id: total_qty} (all lot rows summed)."""
    if _SHEET_NIMBL not in wb.sheetnames:
        log.warning("Sheet %r not found in workbook", _SHEET_NIMBL)
        return {}

    ws = wb[_SHEET_NIMBL]
    totals: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = str(row[_N_ITEM_NUMBER] or "").strip()
        if not item_id:
            continue
        qty = _int_val(row[_N_QUANTITY])
        totals[item_id] = totals.get(item_id, 0) + qty

    return totals


def _parse_office(wb) -> dict[str, dict[str, int]]:
    """Parse the 117 office sheet.  Returns {item_id: {available, damaged}}."""
    if _SHEET_OFFICE not in wb.sheetnames:
        log.warning("Sheet %r not found in workbook", _SHEET_OFFICE)
        return {}

    ws = wb[_SHEET_OFFICE]
    result: dict[str, dict[str, int]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = str(row[_O_ITEM_ID] or "").strip()
        if not item_id or item_id.lower() == "item id":
            continue
        result[item_id] = {
            "available": _int_val(row[_O_AVAILABLE]),
            "damaged":   _int_val(row[_O_DAMAGED]),
        }

    return result


def _parse_xlsx(data: bytes) -> tuple[
    dict[str, dict[str, int]],   # unis: {item_id -> {available, allocated, on_hand, damaged}}
    dict[str, int],              # nimbl: {item_id -> total_qty}
    dict[str, dict[str, int]],   # office: {item_id -> {available, damaged}}
]:
    """Parse all three relevant sheets from the xlsx bytes."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise InventoryClientError(
            "openpyxl is not installed — cannot parse inventory report"
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    unis   = _parse_unis(wb)
    nimbl  = _parse_nimbl(wb)
    office = _parse_office(wb)
    return unis, nimbl, office


# ────────────────────────────────────────────────────────────────────────────
# Formatting
# ────────────────────────────────────────────────────────────────────────────


def _flag(avail: int) -> str:
    """Return emoji flag for a given available-cases count."""
    if avail <= _CRITICAL_CS:
        return "🚨"
    if avail <= _LOW_CS:
        return "⚠️ "
    return "✅"


def _report_date(modified_iso: str) -> str:
    """Convert Drive modifiedTime ISO string to 'YYYY-MM-DD'."""
    try:
        dt = datetime.fromisoformat(modified_iso.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return "unknown"


def _build_inventory_rows(
    unis: dict[str, dict[str, int]],
    nimbl: dict[str, int],
    office: dict[str, dict[str, int]],
) -> dict[str, dict[str, Any]]:
    """Merge all three sources into a unified per-SKU dict."""
    all_ids: set[str] = set(unis) | set(nimbl) | set(office)
    merged: dict[str, dict[str, Any]] = {}

    for item_id in all_ids:
        u = unis.get(item_id, {"available": 0, "allocated": 0, "on_hand": 0, "damaged": 0})
        n = nimbl.get(item_id, 0)
        o = office.get(item_id, {"available": 0, "damaged": 0})
        merged[item_id] = {
            "item_id":       item_id,
            "unis_avail":    u["available"],
            "unis_alloc":    u["allocated"],
            "unis_damaged":  u["damaged"],
            "nimbl_qty":     n,
            "office_avail":  o["available"],
            "office_damaged": o["damaged"],
        }

    return merged


def _total_avail(row: dict[str, Any]) -> int:
    """Total available across all locations (Cotton + Nimbl + office)."""
    return row["unis_avail"] + row["nimbl_qty"] + row["office_avail"]


def format_inventory_pulse(
    unis: dict[str, dict[str, int]],
    nimbl: dict[str, int],
    office: dict[str, dict[str, int]],
    modified_iso: str,
) -> str:
    """Build the Slack-formatted inventory pulse string."""
    rows = _build_inventory_rows(unis, nimbl, office)
    if not rows:
        return "No inventory data found in the latest report."

    report_date = _report_date(modified_iso)
    lines: list[str] = []
    lines.append(f"*📦 F3 Inventory Pulse* _as of {report_date}_")
    lines.append("")

    # ── Alerts section ────────────────────────────────────────────────────────
    alerts: list[str] = []
    for item_id, row in sorted(rows.items()):
        meta = _SKU_META.get(item_id)
        if not meta:
            continue  # Unknown SKU — skip
        name, brand = meta
        total = _total_avail(row)
        flag = _flag(total)
        if flag.strip() != "✅":
            parts = []
            if row["unis_avail"]:
                parts.append(f"{row['unis_avail']:,} cs warehouse")
            if row["nimbl_qty"]:
                parts.append(f"{row['nimbl_qty']:,} cs Nimbl")
            if row["office_avail"]:
                parts.append(f"{row['office_avail']:,} cs office")
            location_str = ", ".join(parts) if parts else "0 cs"
            alerts.append(f"{flag}*{name}* ({brand}) — {location_str}  _{total:,} cs total_")

    if alerts:
        lines.append("*⚠️ Alerts*")
        lines.extend(alerts)
        lines.append("")

    # ── Per-brand breakdown ───────────────────────────────────────────────────
    for brand in _BRAND_ORDER:
        brand_rows = [
            (iid, r) for iid, r in sorted(rows.items())
            if _SKU_META.get(iid, ("", ""))[1] == brand
        ]
        if not brand_rows:
            continue

        lines.append(f"*F3 {brand}*")
        for item_id, row in brand_rows:
            name = _SKU_META[item_id][0]
            total = _total_avail(row)
            flag = _flag(total)

            detail = f"{flag}{name}: *{row['unis_avail']:,} cs* warehouse"
            extras: list[str] = []
            if row["unis_alloc"]:
                extras.append(f"{row['unis_alloc']:,} alloc")
            if row["nimbl_qty"]:
                extras.append(f"{row['nimbl_qty']:,} Nimbl")
            if row["office_avail"]:
                extras.append(f"{row['office_avail']:,} office")
            if row["unis_damaged"] or row["office_damaged"]:
                dmg = row["unis_damaged"] + row["office_damaged"]
                extras.append(f"{dmg:,} damaged")
            if extras:
                detail += f"  _({', '.join(extras)})_"
            lines.append(detail)
        lines.append("")

    lines.append(
        "_✅ Healthy  ⚠️  Low (≤200 cs)  🚨 Critical (≤50 cs)  |  cs = cases of 12_"
    )
    return "\n".join(lines)


# ────────────────────────────────────────────────────────────────────────────
# Location-filtered formatters (UNIS / NIMBL / office)
# ────────────────────────────────────────────────────────────────────────────




def _format_unis_for_llm(
    unis: dict[str, dict[str, int]],
    report_date: str,
    brand: str | None = None,
) -> str:
    """Format UNIS warehouse data as Slack text, optionally filtered by brand."""
    brand_label = f"F3 {brand.capitalize()}" if brand else "F3E"
    lines = [f"*{brand_label} warehouse inventory (UNIS) — as of {report_date}:*", ""]

    any_found = False
    for b in _BRAND_ORDER:
        if brand and b.lower() != brand.lower():
            continue
        b_items = sorted(
            (item_id for item_id in unis if item_id in _SKU_META and _SKU_META[item_id][1] == b)
        )
        if not b_items:
            continue
        if not brand:
            lines.append(f"*F3 {b}*")
        for item_id in b_items:
            name = _SKU_META[item_id][0]
            data = unis[item_id]
            avail = data["available"]
            alloc = data["allocated"]
            dmg = data["damaged"]
            flag = _flag(avail)
            entry = f"{flag}{name}: *{avail:,} cs* available"
            extras = []
            if alloc:
                extras.append(f"{alloc:,} allocated")
            if dmg:
                extras.append(f"{dmg:,} damaged")
            if extras:
                entry += f"  _({', '.join(extras)})_"
            lines.append(entry)
            any_found = True
        if not brand:
            lines.append("")

    if not any_found:
        lines.append("No stock on hand for this brand.")
    lines.append("_Weekly snapshot. cs = cases of 12._")
    return "\n".join(lines)


def _format_nimbl_weekly_for_llm(
    nimbl: dict[str, int],
    report_date: str,
    brand: str | None = None,
) -> str:
    """Format NIMBL lot-total data from the weekly Excel snapshot."""
    brand_label = f"F3 {brand.capitalize()}" if brand else "F3E"
    lines = [f"*{brand_label} Nimbl inventory — as of {report_date}:*", ""]

    any_found = False
    for b in _BRAND_ORDER:
        if brand and b.lower() != brand.lower():
            continue
        b_items = sorted(
            (item_id for item_id in nimbl if item_id in _SKU_META and _SKU_META[item_id][1] == b)
        )
        if not b_items:
            continue
        if not brand:
            lines.append(f"*F3 {b}*")
        for item_id in b_items:
            name = _SKU_META[item_id][0]
            qty = nimbl[item_id]
            flag = _flag(qty)
            lines.append(f"{flag}{name}: *{qty:,} cs*")
            any_found = True
        if not brand:
            lines.append("")

    if not any_found:
        lines.append("No stock on hand for this brand.")
    lines.append(
        "_Note: This is the weekly Excel snapshot. "
        "For real-time Nimbl stock, ask for ‘live Nimbl inventory’._"
    )
    return "\n".join(lines)


def _format_office_for_llm(
    office: dict[str, dict[str, int]],
    report_date: str,
    brand: str | None = None,
) -> str:
    """Format 117 office sheet data, optionally filtered by brand."""
    brand_label = f"F3 {brand.capitalize()}" if brand else "F3E"
    lines = [f"*{brand_label} office stock (117) — as of {report_date}:*", ""]

    any_found = False
    for b in _BRAND_ORDER:
        if brand and b.lower() != brand.lower():
            continue
        b_items = sorted(
            (item_id for item_id in office if item_id in _SKU_META and _SKU_META[item_id][1] == b)
        )
        if not b_items:
            continue
        if not brand:
            lines.append(f"*F3 {b}*")
        for item_id in b_items:
            name = _SKU_META[item_id][0]
            data = office[item_id]
            avail = data["available"]
            dmg = data["damaged"]
            flag = _flag(avail)
            entry = f"{flag}{name}: *{avail:,} cs* available"
            if dmg:
                entry += f"  _({dmg:,} damaged)_"
            lines.append(entry)
            any_found = True
        if not brand:
            lines.append("")

    if not any_found:
        lines.append("No stock on hand for this brand.")
    lines.append("_Weekly snapshot. cs = cases of 12._")
    return "\n".join(lines)


# ────────────────────────────────────────────────────────────────────────────
# Unknown-answer response (mirrors financial_client.UNKNOWN_RESPONSE)
# ────────────────────────────────────────────────────────────────────────────

UNKNOWN_RESPONSE = (
    "I don't have that right now. I will notify the team "
    "immediately to obtain the information and provide the correct and "
    "updated answer when you ask again."
)


# ────────────────────────────────────────────────────────────────────────────
# Public entry points
# ────────────────────────────────────────────────────────────────────────────

def _download_report() -> tuple[bytes, str]:
    """Download the latest inventory report from Drive.

    Returns:
        (file_bytes, modified_iso) tuple.

    Raises:
        InventoryClientError on Drive or auth errors.
    """
    service = _build_service()
    file_id, modified_iso = _find_latest_file(service)
    data = _download_file(service, file_id)
    return data, modified_iso


def get_f3e_inventory_pulse_text() -> str:
    """Return the full F3E inventory pulse as Slack-formatted text.

    Downloads the latest weekly Drive report, parses all three sheets
    (UNIS / NIMBL / 117 office), and returns a formatted summary.

    Returns UNKNOWN_RESPONSE on any Drive or parse error.
    """
    try:
        data, modified_iso = _download_report()
        unis, nimbl, office = _parse_xlsx(data)
        return format_inventory_pulse(unis, nimbl, office, modified_iso)
    except InventoryClientError as exc:
        log.error("inventory_pulse Drive error: %s", exc)
        return UNKNOWN_RESPONSE
    except Exception as exc:  # noqa: BLE001
        log.error("inventory_pulse unexpected error: %s", exc)
        return UNKNOWN_RESPONSE


def get_f3e_location_inventory_text(
    location: str,
    brand: str | None = None,
) -> str:
    """Return location-specific inventory as Slack-formatted text.

    Routes:
      - "nimbl" / "nimbl*"         → _format_nimbl_weekly_for_llm (weekly Excel)
      - "unis" / "warehouse" / "cotton" / "cotton 3pl" → _format_unis_for_llm
      - "office" / "117" / "117 office" → _format_office_for_llm

    Note: Caller handles Nimbl live-Shopify routing. This function always
    reads the weekly Excel snapshot for all three locations.

    Returns UNKNOWN_RESPONSE on Drive/parse errors.
    Returns an error string for unknown location names.
    """
    loc = location.strip().lower()
    try:
        data, modified_iso = _download_report()
        unis, nimbl, office = _parse_xlsx(data)
        report_date = _report_date(modified_iso)
    except InventoryClientError as exc:
        log.error("location_inventory Drive error location=%r: %s", location, exc)
        return UNKNOWN_RESPONSE
    except Exception as exc:  # noqa: BLE001
        log.error("location_inventory unexpected error location=%r: %s", location, exc)
        return UNKNOWN_RESPONSE

    if loc in ("unis", "warehouse", "cotton", "cotton 3pl"):
        return _format_unis_for_llm(unis, report_date, brand)
    if loc in ("nimbl",):
        return _format_nimbl_weekly_for_llm(nimbl, report_date, brand)
    if loc in ("office", "117", "117 office"):
        return _format_office_for_llm(office, report_date, brand)

    known = ["nimbl", "unis", "warehouse", "cotton", "office", "117"]
    return (
        f"Unknown location {location!r}. "
        f"Supported: {', '.join(known)}. "
        "Ask the user to clarify."
    )

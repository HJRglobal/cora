"""QBO monthly-report folder populator (cq-96adf03bcda3).

Replaces Justin's manual QBO export loop. Every month, for each provisioned QBO
realm, pull the prior month's Profit & Loss and Balance Sheet and write
naming-convention .xlsx files into the accounting archive on ``G:``.

WHY THIS IS SAFE TO AUTOMATE
----------------------------
Justin's loop was a pure export with zero transformation (decision F5, 2026-08-18),
which is exactly the shape a machine should own. Three properties keep it honest:

1. **No LLM anywhere.** D-095 -- Cora never computes financials. Every number
   written here is copied verbatim from the QBO report JSON by a deterministic
   walker. There is no summarization, no arithmetic, and no model call in this
   module or its script.

2. **Company-name assertion before every write.** The realm -> slug mapping lives
   in ``data/maps/qbo-monthly-report-slugs.yaml`` and carries the expected QBO
   company name. We read the realm's live ``companyinfo`` and refuse to write
   unless it matches. Filing OSN's P&L as ``llc`` is therefore not a bug that
   review has to catch -- it cannot occur.

3. **Never overwrite.** If a same-named file already exists (Justin or Hayden
   filed it by hand) we write a ``-cora`` suffixed sibling instead and report the
   comparison, so the manual upload becomes an optional parity cross-check rather
   than something this job can destroy.

THE FOLDER/FILE MONTH CONVENTION -- read this before changing dates
-------------------------------------------------------------------
The archive folder is the **filing** month; the files inside it carry the
**report** month, one month earlier. Verified against every populated folder:
``2026-06/`` holds ``2026-05_hjrg_pl.xlsx`` whose own period label is "May 2026".
Getting this backwards is the single easiest way to silently misfile a year of
statements, so ``filing_folder_for`` is a named function with its own test rather
than inline arithmetic. ``--month`` on the script always means the REPORT month.

Basis is pinned to Accrual (the C-track doctrine: QBO otherwise renders each
company in its own default basis, so figures across realms are not comparable)
AND verified from the response header -- pinning a parameter proves nothing if
the API ignores it.
"""

from __future__ import annotations

import calendar
import datetime
import io
import logging
import re
from pathlib import Path
from typing import Any, Callable

import yaml

log = logging.getLogger("cora.qbo_monthly_reports")

_REPO_ROOT = Path(__file__).resolve().parents[2]
_SLUG_MAP_PATH = _REPO_ROOT / "data" / "maps" / "qbo-monthly-report-slugs.yaml"

# Relative to the Founder-OS root, matching the live archive.
ARCHIVE_SUBPATH = ("01-HJR-Global", "accounting", "monthly-reports")

# Report kinds this job produces. The archive also contains `cf` (cash flow) and
# `ar`/`ap` aging for some months; those are deliberately out of scope -- the
# bundle asked for P&L + Balance Sheet, and CashFlow has no qbo_client reader yet.
KIND_PL = "pl"
KIND_BS = "bs"
KINDS = (KIND_PL, KIND_BS)

ACCOUNTING_METHOD = "Accrual"


# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
def load_slug_map(path: Path | None = None) -> dict[str, dict[str, str]]:
    """realm code -> {"slug", "company_name"}. Missing/Parse error -> {}."""
    p = path or _SLUG_MAP_PATH
    try:
        raw = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    except Exception as exc:  # noqa: BLE001 -- a bad map must not crash the job
        log.error("qbo_monthly_reports: slug map unreadable: %s", exc)
        return {}
    out: dict[str, dict[str, str]] = {}
    for realm, spec in (raw.get("realms") or {}).items():
        if not isinstance(spec, dict):
            continue
        slug = str(spec.get("slug") or "").strip()
        name = str(spec.get("company_name") or "").strip()
        if slug and name:
            out[str(realm).strip().upper()] = {
                "slug": slug,
                "company_name": name,
                # Absent => enabled. Only HRLLC ships disabled (personal expense
                # data into a KB-swept folder -- see the map's comment).
                "enabled": bool(spec.get("enabled", True)),
            }
    return out


def unmapped_slugs(path: Path | None = None) -> list[str]:
    """Archive slugs Cora cannot produce (no QBO token). Reported every run."""
    p = path or _SLUG_MAP_PATH
    try:
        raw = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    except Exception:  # noqa: BLE001
        return []
    return [str(s).strip() for s in (raw.get("unmapped_slugs") or []) if str(s).strip()]


def founder_os_root() -> Path:
    """Shared with delegated_worker so both honor FOUNDER_OS_ROOT in tests."""
    import os
    env = os.environ.get("FOUNDER_OS_ROOT", "").strip()
    return Path(env) if env else Path(r"G:\My Drive\HJR-Founder-OS")


# ─────────────────────────────────────────────────────────────────────────────
# Month arithmetic (see the convention note in the module docstring)
# ─────────────────────────────────────────────────────────────────────────────
def previous_month(today: datetime.date | None = None) -> str:
    """The month to REPORT on when the job fires: the last completed month."""
    day = today or datetime.date.today()
    first = day.replace(day=1)
    prev_end = first - datetime.timedelta(days=1)
    return f"{prev_end.year:04d}-{prev_end.month:02d}"


def month_bounds(report_month: str) -> tuple[str, str]:
    """('YYYY-MM-01', 'YYYY-MM-<last>') for a 'YYYY-MM' report month."""
    year, month = _split_month(report_month)
    last = calendar.monthrange(year, month)[1]
    return (f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last:02d}")


def filing_folder_for(report_month: str) -> str:
    """The archive folder a report month is FILED under = report month + 1.

    2026-05 reports live in the 2026-06 folder. Not inlined anywhere: this is the
    one piece of arithmetic that silently misfiles everything if inverted.
    """
    year, month = _split_month(report_month)
    return f"{year + 1:04d}-01" if month == 12 else f"{year:04d}-{month + 1:02d}"


def period_label(report_month: str) -> str:
    """QBO's own header style for a full month, e.g. 'May 2026'."""
    year, month = _split_month(report_month)
    return f"{calendar.month_name[month]} {year}"


_MONTH_RE = re.compile(r"^(\d{4})-(\d{2})$")


def _split_month(report_month: str) -> tuple[int, int]:
    """Strict YYYY-MM. Deliberately rejects what a human backfill typo produces.

    `--month` exists only for operator backfill, so a typo IS its main risk
    surface, and a loose parse was silently destructive in three separate ways
    (D-051): `2026-7` produced the filename `2026-7_llc_pl.xlsx`, which the
    collision check could never match against the real `2026-07_llc_pl.xlsx`, so
    the never-overwrite/parity machinery was bypassed and the archive gained a
    second non-conforming copy of the month; `26-07` resolved to a filing folder
    of `0026-08` and CREATED that directory on the shared Drive. Both are now
    refused rather than normalized, because silently "fixing" a typo hides the
    fact that the operator asked for the wrong month.
    """
    m = _MONTH_RE.match(str(report_month).strip())
    if not m:
        raise ValueError(
            f"report month must be zero-padded YYYY-MM, got {report_month!r}")
    year, month = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"month out of range in {report_month!r}")
    if not 2000 <= year <= 2100:
        raise ValueError(f"year out of range in {report_month!r}")
    return year, month


def assert_month_is_complete(report_month: str,
                             today: datetime.date | None = None) -> None:
    """Refuse a report month that has not finished.

    QBO answers a future as-of date with TODAY's balances, so `--month 2027-07`
    produced a Balance Sheet labelled "July 2027" holding August 2026 figures and
    filed it under a fabricated `2027-08/`. A P&L for the same window is empty and
    self-skips; the balance sheet is the one that lies confidently (D-051).
    """
    day = today or datetime.date.today()
    _, end = month_bounds(report_month)
    if datetime.date.fromisoformat(end) >= day.replace(day=1):
        raise ValueError(
            f"report month {report_month} is not a completed month as of "
            f"{day.isoformat()} -- QBO answers a future as-of date with today's "
            f"balances, which would file current figures under a future label")


def filename(report_month: str, slug: str, kind: str) -> str:
    _split_month(report_month)  # validate
    if kind not in KINDS:
        raise ValueError(f"unknown report kind {kind!r}")
    return f"{report_month}_{slug}_{kind}.xlsx"


def target_dir(report_month: str, root: Path | None = None) -> Path:
    base = root or founder_os_root()
    return base.joinpath(*ARCHIVE_SUBPATH, filing_folder_for(report_month))


def cora_variant_name(name: str) -> str:
    """`2026-07_hjrg_pl.xlsx` -> `2026-07_hjrg_pl-cora.xlsx`."""
    stem, _, ext = name.rpartition(".")
    return f"{stem}-cora.{ext}" if stem else f"{name}-cora"


# ─────────────────────────────────────────────────────────────────────────────
# Report -> rows (deterministic; no arithmetic on any value)
# ─────────────────────────────────────────────────────────────────────────────
def report_rows(report: dict[str, Any]) -> list[tuple[str, str]]:
    """Flatten a QBO report into ordered (label, value) pairs.

    Mirrors the shape of QBO's own .xlsx export: a two-column sheet where
    sections, their leaf accounts and their summary lines appear in document
    order. Values are carried through as the STRINGS QBO returned -- never parsed
    into floats and re-formatted, because a round-trip through float is a
    transformation and this job performs none (D-095).
    """
    out: list[tuple[str, str]] = []

    def cols(node: Any) -> list[str]:
        data = (node or {}).get("ColData") or []
        return [str((c or {}).get("value", "")) for c in data]

    def emit(cells: list[str]) -> None:
        if not cells:
            return
        label = cells[0]
        value = cells[1] if len(cells) > 1 else ""
        if label or value:
            out.append((label, value))

    def walk(rows: Any) -> None:
        for row in (rows or {}).get("Row") or []:
            if not isinstance(row, dict):
                continue
            if row.get("Header"):
                emit(cols(row["Header"]))
            if row.get("Rows"):
                walk(row["Rows"])
            if row.get("ColData"):
                emit(cols(row))
            if row.get("Summary"):
                emit(cols(row["Summary"]))

    walk(report.get("Rows") or {})
    return out


def report_basis(report: dict[str, Any]) -> str | None:
    return (report.get("Header") or {}).get("ReportBasis")


def report_period(report: dict[str, Any]) -> tuple[str | None, str | None]:
    """(StartPeriod, EndPeriod) as QBO echoed them."""
    header = report.get("Header") or {}
    return (header.get("StartPeriod"), header.get("EndPeriod"))


def report_date_macro(report: dict[str, Any]) -> str | None:
    """The date macro QBO actually applied, if any.

    This job ALWAYS asks with explicit dates, so a macro in the response means
    QBO ignored what we sent and substituted a period of its own -- which is
    precisely how the BalanceSheet `as_of_date` bug produced correctly-named
    files holding another month's figures (cq-157a961853c4).

    Checking the echoed EndPeriod alone cannot catch that class: QBO's default
    is "Last Month", and the scheduled run fires on the 2nd asking for the prior
    month, so the default and the request COINCIDE every single time. The
    mismatch is only ever visible on a backfill. The presence of a macro is the
    signature that survives the coincidence.
    """
    header = report.get("Header") or {}
    macro = header.get("DateMacro")
    if macro:
        return str(macro)
    for opt in header.get("Option") or []:
        if str((opt or {}).get("Name")) in ("DateMacro", "date_macro"):
            val = str((opt or {}).get("Value", "")).strip()
            if val:
                return val
    return None


def report_has_no_data(report: dict[str, Any]) -> bool:
    """QBO signals an empty report via a header Option rather than empty Rows."""
    for opt in (report.get("Header") or {}).get("Option") or []:
        if str((opt or {}).get("Name")) == "NoReportData":
            return str((opt or {}).get("Value", "")).lower() == "true"
    return False


def render_xlsx(
    report: dict[str, Any],
    *,
    company_name: str,
    report_title: str,
    period: str,
) -> bytes:
    """Build the .xlsx bytes in QBO's export layout.

    Row 1 company · row 2 report title · row 3 period · row 4 blank · row 5
    ['', 'Total'] · then the rows. Matches the historical files so a human
    diffing Cora's output against a manual export sees the same shape.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([company_name, ""])
    ws.append([report_title, ""])
    ws.append([period, ""])
    ws.append(["", ""])
    ws.append(["", "Total"])
    for label, value in report_rows(report):
        # Numeric cells, matching the ~350 existing archive files (their B7 is a
        # float, not text). A text amount produces a statement that looks right
        # and is arithmetically inert -- it cannot SUM or feed a workbook link
        # (D-051). This is a CELL TYPE, not a transformation: the value is QBO's
        # own, and anything that does not parse cleanly is written unchanged.
        num = _as_number(value)
        ws.append([label, value if num is None else num])
    # QBO's own UI export stamps a basis + generated-at footer, and it turned out
    # to matter: the archive's 2026-05 files were exported on 2026-05-22 -- MID
    # MONTH, before May closed -- so their figures are a snapshot of an open
    # month, not the closed month (HJRG management fees read 2,000 there vs
    # 79,580 once May actually closed). Without this line a reader cannot tell a
    # pre-close snapshot from a post-close statement. Cora fires after close.
    ws.append(["", ""])
    ws.append([footer_stamp(report), ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def footer_stamp(report: dict[str, Any],
                 now: datetime.datetime | None = None) -> str:
    """Provenance footer: basis, QBO's own report clock, and Cora's pull time.

    These are two different clocks and the footer used to print QBO's and label
    it Cora's (D-051). Since this stamp is the artifact that lets a reader tell a
    pre-close snapshot from a post-close statement, mislabelling its timestamp
    defeats its only purpose. `pulled by Cora` is also the marker
    :func:`is_cora_written` keys on, which is what makes a re-run idempotent --
    do not reword it without updating that constant.
    """
    basis = report_basis(report) or "Unspecified"
    qbo_time = (report.get("Header") or {}).get("Time") or "unknown"
    stamp = (now or datetime.datetime.now()).strftime("%Y-%m-%d %H:%M")
    return (f"{basis} Basis | QBO report time {qbo_time} | "
            f"{_CORA_STAMP} {stamp}")


# The REST report and the UI export disagree on wording for the same line: the
# API returns "Total Income"/"Total Expenses", Justin's export says "Total for
# Income"/"Total for Expenses". Normalizing away " for " lets one needle match a
# manual file and a Cora file, which is the whole point of the cross-check.
def _normalize_label(label: str) -> str:
    return " ".join(str(label).lower().replace(" for ", " ").split())


PARITY_KEYS = {KIND_PL: "total income", KIND_BS: "total assets"}


_CORA_STAMP = "pulled by Cora"


def is_cora_written(data: bytes) -> bool:
    """True when this .xlsx carries Cora's footer stamp.

    The only way to tell a Cora-written file from a manual export, and therefore
    the thing that makes a re-run idempotent rather than duplicating.
    """
    from openpyxl import load_workbook
    try:
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        for row in ws.iter_rows(values_only=True):
            if row and row[0] and _CORA_STAMP in str(row[0]):
                return True
    except Exception as exc:  # noqa: BLE001 -- unreadable => treat as not ours
        log.info("qbo_monthly_reports: stamp read failed: %s", exc)
    return False


def _as_number(raw: object) -> float | None:
    text = str(raw if raw is not None else "").strip().replace(",", "")
    text = text.replace("$", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _values_agree(a: object, b: object) -> bool | None:
    """None = could not compare. Numeric when both parse, else exact string.

    Tolerance is a hundredth of a cent: QBO stores totals like
    92339.00000000001, so an exact float compare is as wrong as a string one.
    """
    if a is None or b is None:
        return None
    na, nb = _as_number(a), _as_number(b)
    if na is not None and nb is not None:
        return abs(na - nb) < 0.0001
    return str(a).strip() == str(b).strip()


def read_top_value(data: bytes, label_contains: str) -> str | None:
    """Pull one labelled value out of an .xlsx (parity cross-check only).

    Used to describe a difference to a human; never to decide a write.
    """
    from openpyxl import load_workbook
    try:
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        needle = _normalize_label(label_contains)
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if needle in _normalize_label(row[0] or ""):
                return str(row[1]) if len(row) > 1 and row[1] is not None else ""
    except Exception as exc:  # noqa: BLE001 -- a cross-check must never fail a run
        log.info("qbo_monthly_reports: parity read failed: %s", exc)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Orchestration
# ─────────────────────────────────────────────────────────────────────────────
class Sources:
    """Injection seam so tests exercise the orchestrator with no network."""

    def __init__(
        self,
        *,
        provisioned: Callable[[], list[str]] | None = None,
        company_name: Callable[[str], str | None] | None = None,
        profit_loss: Callable[..., dict[str, Any]] | None = None,
        balance_sheet: Callable[..., dict[str, Any]] | None = None,
    ) -> None:
        self._provisioned = provisioned
        self._company_name = company_name
        self._profit_loss = profit_loss
        self._balance_sheet = balance_sheet

    def provisioned(self) -> list[str]:
        if self._provisioned:
            return self._provisioned()
        from .connectors.qbo_oauth import list_provisioned_entities
        return list_provisioned_entities()

    def company_name(self, entity: str) -> str | None:
        if self._company_name:
            return self._company_name(entity)
        from .tools import qbo_client
        payload = qbo_client._request(
            entity, "/v3/company/{realm_id}/companyinfo/{realm_id}")
        return ((payload or {}).get("CompanyInfo") or {}).get("CompanyName")

    def profit_loss(self, entity: str, start: str, end: str) -> dict[str, Any]:
        if self._profit_loss:
            return self._profit_loss(entity, start, end)
        from .tools import qbo_client
        return qbo_client.get_profit_loss(
            entity, start_date=start, end_date=end,
            accounting_method=ACCOUNTING_METHOD)

    def balance_sheet(self, entity: str, as_of: str) -> dict[str, Any]:
        if self._balance_sheet:
            return self._balance_sheet(entity, as_of)
        from .tools import qbo_client
        return qbo_client.get_balance_sheet(
            entity, as_of_date=as_of, accounting_method=ACCOUNTING_METHOD)


_TITLES = {KIND_PL: "Profit and Loss", KIND_BS: "Balance Sheet"}


def build_month(
    report_month: str,
    *,
    sources: Sources | None = None,
    root: Path | None = None,
    apply: bool = False,
    writer: Callable[[Path, bytes], None] | None = None,
    exists: Callable[[Path], bool] | None = None,
    reader: Callable[[Path], bytes] | None = None,
    slug_map: dict[str, dict[str, str]] | None = None,
) -> dict[str, Any]:
    """Populate one report month. Returns a summary dict (never raises for a
    single entity's failure -- one dead realm must not lose the other ten).
    """
    src = sources or Sources()
    smap = slug_map if slug_map is not None else load_slug_map()
    start, end = month_bounds(report_month)
    outdir = target_dir(report_month, root=root)

    if writer is None or exists is None or reader is None:
        from . import drive_io
        writer = writer or (lambda p, b: drive_io.write_bytes_atomic(p, b))
        exists = exists or (lambda p: bool(drive_io.exists(p)))
        reader = reader or (lambda p: drive_io.read_bytes(p))

    # NOTE for future consumers: this dict DOES carry financial values under
    # `parity` (manual_value/cora_value). The figure-free guarantee lives in
    # format_summary, NOT in the data -- anything that logs or JSON-dumps the
    # summary directly will leak them (HRLLC is personal-expense data).
    summary: dict[str, Any] = {
        "report_month": report_month,
        "filing_folder": filing_folder_for(report_month),
        "target_dir": str(outdir),
        "applied": bool(apply),
        "written": [],
        "parity": [],
        "skipped": [],
        "unmapped_slugs": unmapped_slugs(),
    }

    try:
        provisioned = src.provisioned()
    except Exception as exc:  # noqa: BLE001
        summary["error"] = f"could not list provisioned realms: {exc}"
        return summary

    for entity in sorted(provisioned):
        spec = smap.get(entity.upper())
        if not spec:
            summary["skipped"].append(
                {"entity": entity, "reason": "no slug mapping -- add it to "
                                             "qbo-monthly-report-slugs.yaml"})
            continue
        if not spec.get("enabled", True):
            summary["skipped"].append(
                {"entity": entity,
                 "reason": f"disabled in the slug map ({spec['slug']}) -- "
                           f"personal/sensitive books are opt-in"})
            continue

        # Identity assertion BEFORE any fetch or write.
        try:
            live_name = (src.company_name(entity) or "").strip()
        except Exception as exc:  # noqa: BLE001
            summary["skipped"].append(
                {"entity": entity, "reason": f"companyinfo read failed: {exc}"})
            continue
        if live_name.casefold() != spec["company_name"].casefold():
            summary["skipped"].append({
                "entity": entity,
                "reason": (f"company-name mismatch -- map expects "
                           f"{spec['company_name']!r}, QBO returned "
                           f"{live_name!r}; refusing to file under "
                           f"{spec['slug']!r}"),
            })
            continue

        for kind in KINDS:
            try:
                if kind == KIND_PL:
                    report = src.profit_loss(entity, start, end)
                else:
                    report = src.balance_sheet(entity, end)
            except Exception as exc:  # noqa: BLE001
                summary["skipped"].append(
                    {"entity": entity, "kind": kind, "reason": f"fetch failed: {exc}"})
                continue

            if report_has_no_data(report):
                summary["skipped"].append(
                    {"entity": entity, "kind": kind,
                     "reason": "QBO reports no data for this period"})
                continue

            basis = report_basis(report)
            if basis and basis.casefold() != ACCOUNTING_METHOD.casefold():
                # Pinning the param is not proof it was honored.
                summary["skipped"].append(
                    {"entity": entity, "kind": kind,
                     "reason": f"basis mismatch -- asked {ACCOUNTING_METHOD}, "
                               f"got {basis}"})
                continue

            # The same doctrine applied to the PERIOD, which is what the filename
            # and the row-3 label assert. A silently defaulted or shifted date
            # param yields a correctly-named file holding another period's
            # figures, with nothing to refuse it (D-051).
            want_start = start if kind == KIND_PL else None
            got_start, got_end = report_period(report)
            if (want_start and got_start and got_start != want_start) or \
                    (got_end and got_end != end):
                summary["skipped"].append(
                    {"entity": entity, "kind": kind,
                     "reason": f"period mismatch -- asked "
                               f"{want_start or end}..{end}, got "
                               f"{got_start or '?'}..{got_end or '?'}"})
                continue

            # An EndPeriod that MATCHES is not proof our dates were honored: the
            # scheduled run asks for the prior month and QBO's default is "Last
            # Month", so a wholly-ignored date param produces an identical echo
            # every time it fires. We always send explicit dates, so a macro in
            # the response means QBO substituted a period of its own.
            macro = report_date_macro(report)
            if macro:
                summary["skipped"].append(
                    {"entity": entity, "kind": kind,
                     "reason": f"QBO applied date macro {macro!r} instead of the "
                               f"explicit dates we sent -- the echoed period "
                               f"cannot be trusted to be the one requested"})
                continue

            data = render_xlsx(
                report,
                company_name=spec["company_name"],
                report_title=_TITLES[kind],
                period=period_label(report_month),
            )

            want = filename(report_month, spec["slug"], kind)
            path = outdir / want
            try:
                collided = bool(exists(path))
            except Exception as exc:  # noqa: BLE001 -- unknown => assume present
                summary["skipped"].append(
                    {"entity": entity, "kind": kind,
                     "reason": f"could not test for an existing file: {exc}"})
                continue

            if collided:
                # WHOSE file is it? A previous Cora run leaves its own stamp in
                # the footer. Without this test a re-run (realm 7 of 11 hit a QBO
                # 5xx, operator re-runs the month) read its OWN earlier output as
                # "the manual upload", wrote a -cora duplicate beside it, and on a
                # third run silently overwrote that -- unbounded duplication of
                # statements of record from the one guarantee this job makes.
                existing_bytes = None
                try:
                    existing_bytes = reader(path)
                except Exception as exc:  # noqa: BLE001
                    log.info("qbo_monthly_reports: could not read %s: %s",
                             path.name, exc)

                if existing_bytes is not None and is_cora_written(existing_bytes):
                    summary["skipped"].append(
                        {"entity": entity, "kind": kind,
                         "reason": f"{want} was already written by Cora -- "
                                   f"re-run is a no-op (delete it to regenerate)"})
                    continue

                variant = outdir / cora_variant_name(want)
                try:
                    if bool(exists(variant)):
                        summary["skipped"].append(
                            {"entity": entity, "kind": kind,
                             "reason": f"{want} exists (manual) and "
                                       f"{variant.name} is already there -- "
                                       f"refusing to overwrite either"})
                        continue
                except Exception as exc:  # noqa: BLE001 -- unknown => refuse
                    summary["skipped"].append(
                        {"entity": entity, "kind": kind,
                         "reason": f"could not test for {variant.name}: {exc}"})
                    continue

                path = variant
                note = {"entity": entity, "kind": kind, "manual_file": want,
                        "cora_file": path.name}
                try:
                    key = PARITY_KEYS[kind]
                    note["key"] = key
                    note["manual_value"] = (
                        read_top_value(existing_bytes, key)
                        if existing_bytes is not None else None)
                    note["cora_value"] = read_top_value(data, key)
                    # Compare NUMERICALLY. The archive stores 2000.0 and this
                    # renders "2000.00"; a string compare called every single
                    # collision a DIFFERS and sent a human hunting a discrepancy
                    # that did not exist (D-051). None still means "could not
                    # compare", which reads differently from "compared and
                    # differs" -- collapsing those two was the earlier bug here.
                    note["match"] = _values_agree(note["manual_value"],
                                                  note["cora_value"])
                except Exception as exc:  # noqa: BLE001
                    note["compare_error"] = str(exc)
                summary["parity"].append(note)

            if apply:
                try:
                    writer(path, data)
                except Exception as exc:  # noqa: BLE001
                    summary["skipped"].append(
                        {"entity": entity, "kind": kind,
                         "reason": f"write failed: {exc}"})
                    continue
            summary["written"].append(
                {"entity": entity, "kind": kind, "file": path.name,
                 "bytes": len(data)})

    return summary


def format_summary(summary: dict[str, Any]) -> str:
    """Figure-FREE run report.

    Deliberately carries filenames, counts and reasons but NO financial values:
    this text is safe to post anywhere, which keeps a routine ops digest from
    becoming the leak surface (HRLLC is personal-expense data, and the A5 lesson
    was two figure leaks into shared surfaces). Numbers live in the files.
    """
    lines: list[str] = []
    mode = "APPLIED" if summary.get("applied") else "DRY-RUN (no files written)"
    lines.append(f"QBO monthly reports -- {summary['report_month']} -> "
                 f"{summary['filing_folder']}/  [{mode}]")
    lines.append(f"target: {summary['target_dir']}")
    if summary.get("error"):
        lines.append(f"ERROR: {summary['error']}")
        return "\n".join(lines)

    written = summary.get("written") or []
    lines.append(f"files: {len(written)}")
    for w in written:
        lines.append(f"  + {w['file']}")
    for p in summary.get("parity") or []:
        if p.get("match") is True:
            verdict = f"matches the manual upload on {p.get('key')}"
        elif p.get("match") is False:
            verdict = f"DIFFERS from the manual upload on {p.get('key')}"
        else:
            verdict = "could not be compared"
        lines.append(f"  ! {p['manual_file']} already existed -> wrote "
                     f"{p['cora_file']}; {verdict}")
    for s in summary.get("skipped") or []:
        kind = f"/{s['kind']}" if s.get("kind") else ""
        lines.append(f"  - {s['entity']}{kind}: {s['reason']}")
    unmapped = summary.get("unmapped_slugs") or []
    if unmapped:
        lines.append(f"still a manual export ({len(unmapped)} entities, no QBO "
                     f"token): {', '.join(unmapped)}")
    return "\n".join(lines)
